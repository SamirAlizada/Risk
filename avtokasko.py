import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font
from openpyxl.styles import Alignment

# Fayl yolları
ucot_file = r"C:\Users\samir.alizade\Desktop\Automantion\Risk\Ucot.xlsx"
target_file = r"C:\Users\samir.alizade\Desktop\Automantion\Risk\AvtoKaskoPY.xlsx"

# 1. Ucot -> Simple sheet oxu
df = pd.read_excel(ucot_file, sheet_name="Simple")

# 2. I sütununa görə filter
filtered_df = df[df["I"] == "(04)AvtoKasko"]

# 3. Lazım olan sütunları seç
result_df = filtered_df[["II", "III", "VII", "XI"]].copy()

# 🔹 NaN dəyərləri 0 ilə əvəz et
result_df["VII"] = result_df["VII"].fillna(0)
result_df["XI"] = result_df["XI"].fillna(0)

print(f"📊 Total filtered rows: {len(result_df)}")

# 4. Target excel aç
book = load_workbook(target_file)
sheet = book["Forma8_1"]

# Forma8_1 sheet-də bütün merge-ləri aç
for merged_range in list(sheet.merged_cells.ranges):
    sheet.unmerge_cells(str(merged_range))

# 🔹 Təmizləmə
expected_rows = len(result_df)
end_row = 12 + expected_rows
total_row = end_row

print(f"🧹 12-ci sətirdən {total_row + 10}-cu sətirə qədər təmizləyirəm...")

for row in range(12, total_row + 10):
    for col in range(1, 10):
        sheet.cell(row=row, column=col).value = None

# 🔹 All Borders üçün stil
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 🔹 Normal font
cell_font = Font(name='A3 Times AzLat', size=14)

# 🔹 Bold font (yalnız yekun sətir üçün)
bold_font = Font(name='A3 Times AzLat', size=14, bold=True)

center_align = Alignment(horizontal='center', vertical='center')

# A11 başlığı
sheet["A11"] = "A"
sheet["A11"].border = thin_border
sheet["A11"].font = cell_font

start_row = 12

# 🔹 Toplamları saxlamaq üçün dəyişənlər
total_D = 0
total_E = 0
total_F = 0
total_G = 0
total_H = 0
total_I = 0

row_count = 0

for idx, row in result_df.iterrows():
    try:
        vii = float(row["VII"])
        xi = float(row["XI"])

        limit_15 = round(vii * 0.15, 2)
        final_xi = round(limit_15 if xi > limit_15 else xi, 2)

        # A sütunu: A1, A2 ...
        sheet[f"A{start_row}"] = f"A{start_row-11}"
        sheet[f"A{start_row}"].border = thin_border
        sheet[f"A{start_row}"].font = cell_font

        sheet[f"B{start_row}"] = row["II"]
        sheet[f"B{start_row}"].border = thin_border
        sheet[f"B{start_row}"].font = cell_font

        # C sütunu: tarix
        if pd.notna(row["III"]):
            date_value = pd.to_datetime(row["III"]).date()
            cell = sheet[f"C{start_row}"]
            cell.value = date_value
            cell.number_format = "DD.MM.YYYY"
        else:
            sheet[f"C{start_row}"] = ""
        sheet[f"C{start_row}"].border = thin_border
        sheet[f"C{start_row}"].font = cell_font

        val_D = round(vii, 2)
        val_E = round(final_xi, 2)
        val_F = round(vii - final_xi, 2)
        val_G = round(vii * 0.02, 2)
        val_H = round(final_xi * 0.02, 2)
        val_I = round(val_G - val_H, 2)

        # D–I sütunları
        for col, val in zip(['D','E','F','G','H','I'], [val_D, val_E, val_F, val_G, val_H, val_I]):
            sheet[f"{col}{start_row}"] = val
            sheet[f"{col}{start_row}"].border = thin_border
            sheet[f"{col}{start_row}"].font = cell_font

        sheet[f"A{start_row}"].alignment = center_align
        sheet[f"B{start_row}"].alignment = center_align
        sheet[f"C{start_row}"].alignment = center_align
        sheet[f"D{start_row}"].alignment = center_align
        sheet[f"E{start_row}"].alignment = center_align
        sheet[f"F{start_row}"].alignment = center_align
        sheet[f"G{start_row}"].alignment = center_align
        sheet[f"H{start_row}"].alignment = center_align
        sheet[f"I{start_row}"].alignment = center_align


        total_D += val_D
        total_E += val_E
        total_F += val_F
        total_G += val_G
        total_H += val_H
        total_I += val_I

        row_count += 1
        start_row += 1
        
    except Exception as e:
        print(f"❌ Error at row {idx}: {e}")
        continue

print(f"✅ {row_count} sətir yazıldı (expected: {len(result_df)})")

# 🔹 Toplamları yuvarlaqlaşdır
total_D = round(total_D, 2)
total_E = round(total_E, 2)
total_F = round(total_F, 2)
total_G = round(total_G, 2)
total_H = round(total_H, 2)
total_I = round(total_I, 2)

print(f"\n📊 TOPLAMLAR:")
print(f"Total D: {total_D}")
print(f"Total E: {total_E}")
print(f"Total F: {total_F}")
print(f"Total G: {total_G}")
print(f"Total H: {total_H}")
print(f"Total I: {total_I}")
print(f"Yekun sətiri: {start_row}")

# 🔹 Yekun sətirinə yaz və format et
sheet.cell(row=start_row, column=1).value = "AA1"
sheet.cell(row=start_row, column=1).border = thin_border
sheet.cell(row=start_row, column=1).font = cell_font
sheet.cell(row=start_row, column=1).alignment = center_align

sheet.cell(row=start_row, column=2).value = "Yekun BSH"
sheet.cell(row=start_row, column=2).border = thin_border
sheet.cell(row=start_row, column=2).font = cell_font
sheet.cell(row=start_row, column=2).alignment = center_align

# D–I sütunları Bold və ortalanmış
for col, val in zip(range(4,10), [total_D, total_E, total_F, total_G, total_H, total_I]):
    sheet.cell(row=start_row, column=col).value = val
    sheet.cell(row=start_row, column=col).border = thin_border
    sheet.cell(row=start_row, column=col).font = bold_font
    sheet.cell(row=start_row, column=col).alignment = center_align

print(f"\n✅ Yekun {start_row}-ci sətirə yazıldı")

# 6. Yadda saxla
book.save(target_file)

print("\n✅ Fayl saxlanıldı!")
