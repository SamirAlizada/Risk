import pandas as pd
import shutil
import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment

# ================== PATHS ==================
ucot_file = r"C:\Users\samir.alizade\Desktop\Automantion\Risk\UcotA.xlsx"
template_file = r"C:\Users\samir.alizade\Desktop\Automantion\Risk\ALL.xlsx"
output_folder = r"C:\Users\samir.alizade\Desktop\Automantion\Risk\Auto"

os.makedirs(output_folder, exist_ok=True)

# ================== STYLES ==================
thin = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
font = Font(name="A3 Times AZ Lat", size=14)
bold_font = Font(name="A3 Times AZ Lat", size=14, bold=True)
center = Alignment(horizontal="center", vertical="center")
ROW_HEIGHT = 28

# ================== READ UCOT ==================
df = pd.read_excel(ucot_file, sheet_name="Simple")
products = df["I"].dropna().unique()

# ================== MAIN LOOP ==================
for product in products:

    # 🔴 ƏSAS FİLTR: yalnız II boş olmayan sətirlər
    filtered = df[
        (df["I"] == product) &
        (df["II"].notna())
    ][["II", "III", "VII", "XI"]].copy()

    # Excel yaradılır (data olsa da, olmasa da)
    output_file = os.path.join(output_folder, f"{product}.xlsx")
    shutil.copy(template_file, output_file)

    wb = load_workbook(output_file)
    ws = wb["Forma8_1"]

    # Product adı
    ws["C8"] = product
    ws["C8"].font = font
    ws["C8"].alignment = center

    # Merge-ləri aç
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))

    # 🔹 Əgər KEÇİLƏ BİLƏN data yoxdursa → default fayl
    if filtered.empty:
        wb.save(output_file)
        print(f"⚠️ Data yoxdur (II hamısı boşdur): {output_file}")
        continue

    filtered["VII"] = filtered["VII"].fillna(0)
    filtered["XI"] = filtered["XI"].fillna(0)

    start_row = 12
    ws.insert_rows(start_row, amount=len(filtered))

    total_D = total_E = total_F = total_G = total_H = total_I = 0
    r = start_row

    for _, row in filtered.iterrows():
        vii = round(float(row["VII"]), 2)
        xi = round(float(row["XI"]), 2)

        limit = round(vii * 0.15, 2)
        final_xi = limit if xi > limit else xi

        # Default G H I
        val_G = val_H = val_I = 0

        if product in [
            "(04)AvtoKasko",
            "(08)Yuk",
            "(03)EmlakYanginDigerRisk",
            "(37)IcbariDashinmazEmlak"
        ]:
            if product == "(04)AvtoKasko":
                rate = 0.02
            elif product == "(08)Yuk":
                rate = 0.01
            else:
                rate = 0.20

            val_G = round(vii * rate, 2)
            val_H = round(final_xi * rate, 2)
            val_I = round(val_G - val_H, 2)

        values = [
            f"A{r-11}",
            row["II"],
            pd.to_datetime(row["III"]).date() if pd.notna(row["III"]) else "",
            vii,
            final_xi,
            round(vii - final_xi, 2),
            val_G,
            val_H,
            val_I
        ]

        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c)
            cell.value = v
            cell.font = font
            cell.border = thin
            cell.alignment = center
            if c == 3:
                cell.number_format = "DD.MM.YYYY"

        total_D += values[3]
        total_E += values[4]
        total_F += values[5]
        total_G += values[6]
        total_H += values[7]
        total_I += values[8]

        ws.row_dimensions[r].height = ROW_HEIGHT
        r += 1

    # ================== YEKUN ==================
    ws.cell(row=r, column=1).value = "AA1"
    ws.cell(row=r, column=2).value = "Yekun BSH"

    totals = [total_D, total_E, total_F, total_G, total_H, total_I]
    for i, v in enumerate(totals, start=4):
        c = ws.cell(row=r, column=i)
        c.value = round(v, 2)
        c.font = bold_font
        c.border = thin
        c.alignment = center

    ws.row_dimensions[r].height = ROW_HEIGHT

    # 4 sətir aşağı copy
    copy_row = r + 4
    for i, v in enumerate(totals, start=4):
        c = ws.cell(row=copy_row, column=i)
        c.value = round(v, 2)
        c.font = bold_font
        c.border = thin
        c.alignment = center

    ws.row_dimensions[copy_row].height = ROW_HEIGHT

    wb.save(output_file)
    print(f"✅ Hazırdır: {output_file}")
