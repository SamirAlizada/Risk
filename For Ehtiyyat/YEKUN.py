from openpyxl import load_workbook
import os
import shutil
import pandas as pd

UCOT_FILE = r"C:\Users\samir.alizade\Desktop\Automantion\Risk\UcotA.xlsx"
TEMPLATE_FILE = r"C:\Users\samir.alizade\Desktop\Automantion\Risk\ALL.xlsx"
OUTPUT_FOLDER = r"C:\Users\samir.alizade\Desktop\Automantion\Risk\AutoTEST"
PREVIOUS_FOLDER = r"C:\Users\samir.alizade\Desktop\Automantion\Risk\Previous"
YEKUN_TEMPLATE = r"C:\Users\samir.alizade\Desktop\Automantion\Risk\Yekun Reserv.xlsx"
REFERENCE_DATE = "2026-01-01"

# Type1 və Type2 məhsul siyahıları
TYPE1_PRODUCTS = [
    "(01)FerdiQeza",
    "(02)Tibbi",
    "(03)EmlakYanginDigerRisk",
    "(04)AvtoKasko",
    "(05)DemiryolNeqliyyVasitesi",
    "(06)HavaNeqliyyKasko",
    "(07)SuNeqliyyKasko",
    "(08)Yuk",
    "(09)KendTeserrufBitki",
    "(10)KendTeserrufHeyvan",
    "(11)IshcilerinDeleduzlug",
    "(12)PulvePulSenedSaxtalash",
    "(22)Kredit",
    "(23)Ipoteka",
    "(24)EmlakinDeyerdenDushmesi",
    "(25)IshinDayanmasiRiski",
    "(27)SernishinIcbari",
    "(28)IcbariEkoloji",
    "(29)YanginIcbari",
    "(30)DeputatlarinIcbari",
    "(31)TibbiPersonalinAIDSden",
    "(32)HerbiQulluqcularinIcbari",
    "(33)HuquqMuhafifeIcbari",
    "(34)DovletQulluqcuIcbari",
    "(35)DiplomatlarinIcbari",
    "(37)IcbariDashinmazEmlak",
    "(39)IcbariNVSMMS",
    "(40)IcbariSernishinFerdiQeza",
    "(41)Sefer",
    "(42)Titul",
    "(43)HuquqiXerc"
]

TYPE2_PRODUCTS = [
    "(19)PesheMesuliyy",
    "(20)IshegoturenMesuliyy",
    "(21)UmumiMulkiMesuliyy",
    "(13)AvtoKonulluMesuliyy",
    "(14)DemiryolNeqliySahibMesuliyy",
    "(15)HavaNeqliySahibMesuliyy",
    "(16)SuNeqliySahibMesuliyy",
    "(17)YukDashiyanMesuliyy",
    "(18)MulkiMuqavileUzreMesuliyy",
    "(26)AvtoIcbariMesuliyy",
    "(36)AuditorPesheMesuliyyIcbari",
    "(38)IcbariDashinmazEmlakMesul"
]

# Forma8_14 hesablama tələb edən məhsullar (F ilə başlayanlar üçün)
FORMA8_14_PRODUCTS = [
    "(04)AvtoKasko",
    "(08)Yuk",
    "(03)EmlakYanginDigerRisk",
    "(37)IcbariDashinmazEmlak"
]

def run_yekun_reserv(
    yekun_template: str,
    output_folder: str,
    processed_excels_folder: str,
    reference_date: str
):
    """
    Yekun Reserv.xlsx faylını kopyalayıb məlumatları doldurur
    
    Args:
        yekun_template: Orijinal "Yekun Reserv.xlsx" faylının yolu
        output_folder: Yeni "Yekun Reserv_processed.xlsx" saxlanılacaq folder
        processed_excels_folder: Product Excel fayllarının olduğu folder (AutoTEST)
    """
    
    print("=" * 60)
    print("▶ YEKUN RESERV PROSESİ BAŞLADI")
    print("=" * 60)
    
    # ==================== YOXLAMA ====================
    if not os.path.exists(yekun_template):
        raise FileNotFoundError(f"Yekun Reserv template tapılmadı: {yekun_template}")
    
    if not os.path.exists(processed_excels_folder):
        raise FileNotFoundError(f"Processed excels folder tapılmadı: {processed_excels_folder}")
    
    # ==================== KOPYALAMA ====================
    output_file = os.path.join(output_folder, "Yekun Reserv_processed.xlsx")
    
    print(f"  📋 Template kopyalanır...")
    print(f"     Source: {os.path.basename(yekun_template)}")
    print(f"     Target: {os.path.basename(output_file)}")
    
    shutil.copy2(yekun_template, output_file)
    print(f"  ✅ Kopyalama tamamlandı")
    
    # ==================== EXCEL AÇMA ====================
    wb = load_workbook(output_file)
    
    # İlk sheet-i götür (adətən "Sheet1" və ya ilk aktiv sheet)
    ws = wb.active
    
    print(f"\n  📊 Sheet: {ws.title}")

    # ================== E7-Ə TARİX YAZMA ==================
    ref_date = pd.to_datetime(reference_date)
    formatted_date = ref_date.strftime("%d.%m.%Y")
    ws["E7"].value = formatted_date
    print(f"    ✓ E7-ə tarix yazıldı: {formatted_date}")
    
    # ==================== PRODUCT SƏTİRLƏRİNİ TAP ====================
    product_rows = []
    
    print(f"\n  🔍 A sütununda 'A' ilə başlayan sətirləri axtarır...")
    
    for row in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        cell_d = ws.cell(row=row, column=4)
        
        # A sütunu "A" ilə başlayırsa və D sütununda product adı varsa
        if cell_a.value and isinstance(cell_a.value, str):
            a_val = cell_a.value.strip().upper()
            
            # A1, A2, A3... formatında olmalıdır (AA1, AB1 deyil)
            # və "Aralıq", "Yekun" kimi sözlər olmamalıdır
            if a_val.startswith("A") and len(a_val) <= 3:
                product_name = cell_d.value
                
                if product_name and isinstance(product_name, str):
                    # "Aralıq yekun", "Yekun" kimi sözləri skip et
                    if "aralıq" in product_name.lower() or "yekun" in product_name.lower():
                        print(f"    ⊘ Sətir {row} skip edildi: {product_name}")
                        continue
                    
                    product_rows.append({
                        "row": row,
                        "product": product_name.strip(),
                        "a_value": cell_a.value
                    })
                    print(f"    ✓ Sətir {row}: A={cell_a.value}, Product={product_name}")
    
    print(f"\n  📌 Toplam {len(product_rows)} product sətri tapıldı (A ilə başlayan)")
    
    # ==================== B İLƏ BAŞLAYAN SƏTİRLƏRİ TAP ====================
    product_rows_b = []
    
    print(f"\n  🔍 A sütununda 'B' ilə başlayan sətirləri axtarır...")
    
    for row in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        cell_d = ws.cell(row=row, column=4)
        
        # A sütunu "B" ilə başlayırsa və D sütununda product adı varsa
        if cell_a.value and isinstance(cell_a.value, str):
            a_val = cell_a.value.strip().upper()
            
            # B1, B2, B3... formatında olmalıdır (BB1, BC1 deyil)
            # və "Aralıq", "Yekun" kimi sözlər olmamalıdır
            if a_val.startswith("B") and len(a_val) <= 3 and not a_val.startswith("BB"):
                product_name = cell_d.value
                
                if product_name and isinstance(product_name, str):
                    # "Aralıq yekun", "Yekun" kimi sözləri skip et
                    if "aralıq" in product_name.lower() or "yekun" in product_name.lower():
                        print(f"    ⊘ Sətir {row} skip edildi: {product_name}")
                        continue
                    
                    product_rows_b.append({
                        "row": row,
                        "product": product_name.strip(),
                        "a_value": cell_a.value
                    })
                    print(f"    ✓ Sətir {row}: A={cell_a.value}, Product={product_name}")
    
    print(f"\n  📌 Toplam {len(product_rows_b)} product sətri tapıldı (B ilə başlayan)")
    
    # ==================== C İLƏ BAŞLAYAN SƏTİRLƏRİ TAP ====================
    product_rows_c = []
    
    print(f"\n  🔍 A sütununda 'C' ilə başlayan sətirləri axtarır...")
    
    for row in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        cell_d = ws.cell(row=row, column=4)
        
        # A sütunu "C" ilə başlayırsa və D sütununda product adı varsa
        if cell_a.value and isinstance(cell_a.value, str):
            a_val = cell_a.value.strip().upper()
            
            # C1, C2, C3... formatında olmalıdır (CC1, CD1 deyil)
            # və "Aralıq", "Yekun" kimi sözlər olmamalıdır
            if a_val.startswith("C") and len(a_val) <= 3 and not a_val.startswith("CC"):
                product_name = cell_d.value
                
                if product_name and isinstance(product_name, str):
                    # "Aralıq yekun", "Yekun" kimi sözləri skip et
                    if "aralıq" in product_name.lower() or "yekun" in product_name.lower():
                        print(f"    ⊘ Sətir {row} skip edildi: {product_name}")
                        continue
                    
                    product_rows_c.append({
                        "row": row,
                        "product": product_name.strip(),
                        "a_value": cell_a.value
                    })
                    print(f"    ✓ Sətir {row}: A={cell_a.value}, Product={product_name}")
    
    print(f"\n  📌 Toplam {len(product_rows_c)} product sətri tapıldı (C ilə başlayan)")
    
    # ==================== F İLƏ BAŞLAYAN SƏTİRLƏRİ TAP ====================
    product_rows_f = []
    
    print(f"\n  🔍 A sütununda 'F' ilə başlayan sətirləri axtarır...")
    
    for row in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        cell_d = ws.cell(row=row, column=4)
        
        # A sütunu "F" ilə başlayırsa və D sütununda product adı varsa
        if cell_a.value and isinstance(cell_a.value, str):
            a_val = cell_a.value.strip().upper()
            
            # F1, F2, F3... formatında olmalıdır (FF1, FG1 deyil)
            # və "Aralıq", "Yekun" kimi sözlər olmamalıdır
            if a_val.startswith("F") and len(a_val) <= 3 and not a_val.startswith("FF"):
                product_name = cell_d.value
                
                if product_name and isinstance(product_name, str):
                    # "Aralıq yekun", "Yekun" kimi sözləri skip et
                    if "aralıq" in product_name.lower() or "yekun" in product_name.lower():
                        print(f"    ⊘ Sətir {row} skip edildi: {product_name}")
                        continue
                    
                    product_rows_f.append({
                        "row": row,
                        "product": product_name.strip(),
                        "a_value": cell_a.value
                    })
                    print(f"    ✓ Sətir {row}: A={cell_a.value}, Product={product_name}")
    
    print(f"\n  📌 Toplam {len(product_rows_f)} product sətri tapıldı (F ilə başlayan)")
    
    if len(product_rows) == 0 and len(product_rows_b) == 0 and len(product_rows_c) == 0 and len(product_rows_f) == 0:
        print(f"  ⚠ Heç bir product sətri tapılmadı!")
        wb.save(output_file)
        wb.close()
        return
    
    # ==================== HƏR PRODUCT ÜÇÜN MƏLUMAT DOLDURMA (A ilə başlayan) ====================
    print(f"\n{'='*60}")
    print(f"▶ A İLƏ BAŞLAYAN PRODUCTLAR (Forma8_2, 8_6, 8_5)")
    print(f"{'='*60}")
    
    success_count = 0
    
    for item in product_rows:
        row_num = item["row"]
        product = item["product"]
        
        print(f"\n  ▶ Sətir {row_num}: {product}")
        
        # Product Excel faylını tap
        product_file = os.path.join(processed_excels_folder, f"{product}.xlsx")
        
        if not os.path.exists(product_file):
            print(f"    ⚠ Excel tapılmadı: {product}.xlsx")
            continue
        
        try:
            # Product Excel-i aç
            # QEYD: data_only=True ilə formullar hesablanmır, ona görə manual hesablayacağıq
            wb_product = load_workbook(product_file, data_only=False)
            
            # ==================== FORMA8_2-DƏN F TOPLAMI ====================
            forma8_2_f_total = None
            
            if "Forma8_2" in wb_product.sheetnames:
                ws_8_2 = wb_product["Forma8_2"]
                
                print(f"    🔍 Forma8_2-də toplam tapılır...")
                
                # B sütununda "Yekun BSH" olan sətiri tap
                for r in range(1, ws_8_2.max_row + 1):
                    cell_b = ws_8_2.cell(row=r, column=2)
                    cell_f = ws_8_2.cell(row=r, column=6)
                    
                    if cell_b.value and isinstance(cell_b.value, str):
                        if "yekun" in cell_b.value.lower() and "bsh" in cell_b.value.lower():
                            # Bu toplam sətridir
                            # F sütununda formula var, onu parse edək
                            if cell_f.value and isinstance(cell_f.value, str) and cell_f.value.startswith("="):
                                # Formula: =ROUND(SUM(F12:F214),2)
                                # SUM-dakı ranqı tap və manual hesabla
                                formula = cell_f.value
                                
                                # Range-i tap: F12:F214
                                import re
                                match = re.search(r'F(\d+):F(\d+)', formula)
                                
                                if match:
                                    start_r = int(match.group(1))
                                    end_r = int(match.group(2))
                                    
                                    total = 0.0
                                    for calc_r in range(start_r, end_r + 1):
                                        calc_cell = ws_8_2.cell(row=calc_r, column=6)
                                        
                                        # Hər hüceyrədə də formula var, onu da hesabla
                                        if calc_cell.value and isinstance(calc_cell.value, str) and calc_cell.value.startswith("="):
                                            # Formula: =ROUND((D12-E12)/D12*C12,2)
                                            c_val = ws_8_2.cell(row=calc_r, column=3).value or 0
                                            d_val = ws_8_2.cell(row=calc_r, column=4).value or 1
                                            e_val = ws_8_2.cell(row=calc_r, column=5).value or 0
                                            
                                            try:
                                                c_val = float(c_val)
                                                d_val = float(d_val)
                                                e_val = float(e_val)
                                                
                                                if d_val > 0:
                                                    row_result = ((d_val - e_val) / d_val) * c_val
                                                    total += row_result
                                            except (ValueError, TypeError):
                                                pass
                                    
                                    forma8_2_f_total = round(total, 2)
                                    print(f"    ✓ Forma8_2 F toplam (F{r}): {forma8_2_f_total:.2f} [manual hesablandı]")
                                    break
                            
                            # Əgər formula deyilsə, birbaşa dəyəri götür
                            elif cell_f.value is not None:
                                try:
                                    forma8_2_f_total = float(cell_f.value)
                                    print(f"    ✓ Forma8_2 F toplam (F{r}): {forma8_2_f_total:.2f}")
                                    break
                                except (ValueError, TypeError):
                                    pass
                
                if forma8_2_f_total is None:
                    print(f"    ⚠ Forma8_2-də 'Yekun BSH' sətiri tapılmadı")
            else:
                print(f"    ⚠ Forma8_2 sheet-i yoxdur")
            
            # ==================== FORMA8_6-DAN E16 ====================
            forma8_6_e16 = None
            
            if "Forma8_6" in wb_product.sheetnames:
                ws_8_6 = wb_product["Forma8_6"]
                
                cell_e16 = ws_8_6.cell(row=16, column=5)
                
                if cell_e16.value is not None:
                    try:
                        forma8_6_e16 = float(cell_e16.value)
                        print(f"    ✓ Forma8_6 E16: {forma8_6_e16:.2f}")
                    except (ValueError, TypeError):
                        print(f"    ⚠ Forma8_6 E16 rəqəm deyil: {cell_e16.value}")
                else:
                    print(f"    ⚠ Forma8_6 E16 boşdur")
            else:
                print(f"    ⚠ Forma8_6 sheet-i yoxdur")
            
            # ==================== E SÜTUNUNA YAZMA ====================
            if forma8_2_f_total is not None and forma8_6_e16 is not None:
                e_value = forma8_2_f_total + forma8_6_e16
                e_value = round(e_value, 2)
                
                ws.cell(row=row_num, column=5).value = e_value
                print(f"    ✓ E{row_num} = {forma8_2_f_total:.2f} + {forma8_6_e16:.2f} = {e_value:.2f}")
            else:
                print(f"    ⚠ E{row_num} hesablana bilmədi (məlumat tam deyil)")
            
            # ==================== FORMA8_5-DƏN F TOPLAMI ====================
            forma8_5_f_total = None
            
            if "Forma8_5" in wb_product.sheetnames:
                ws_8_5 = wb_product["Forma8_5"]
                
                print(f"    🔍 Forma8_5-də F sütununda ÜMUMİ CƏM tapılır...")
                
                # Ümumi cəm sətrini tap (ən son bold toplam)
                for r in range(ws_8_5.max_row, 0, -1):
                    cell_f = ws_8_5.cell(row=r, column=6)
                    
                    # Bold və dəyər varsa
                    if cell_f.value is not None and cell_f.font and cell_f.font.bold:
                        try:
                            forma8_5_f_total = float(cell_f.value)
                            print(f"    ✓ Forma8_5 F ümumi cəm (F{r}): {forma8_5_f_total:.2f}")
                            break
                        except (ValueError, TypeError):
                            pass
                
                if forma8_5_f_total is None:
                    print(f"    ⚠ Forma8_5-də F toplam tapılmadı")
            else:
                print(f"    ⚠ Forma8_5 sheet-i yoxdur")
            
            # ==================== F SÜTUNUNA YAZMA ====================
            if forma8_5_f_total is not None:
                ws.cell(row=row_num, column=6).value = forma8_5_f_total
                print(f"    ✓ F{row_num} = {forma8_5_f_total:.2f}")
            else:
                print(f"    ⚠ F{row_num} yazıla bilmədi (Forma8_5 toplam yoxdur)")
            
            wb_product.close()
            success_count += 1
            
        except Exception as e:
            print(f"    ❌ Xəta: {e}")
    
    # ==================== F İLƏ BAŞLAYAN PRODUCTLAR ÜÇÜN MƏLUMAT DOLDURMA ====================
    print(f"\n{'='*60}")
    print(f"▶ F İLƏ BAŞLAYAN PRODUCTLAR (Forma8_14)")
    print(f"{'='*60}")
    
    success_count_f = 0
    skipped_count_f = 0
    
    for item in product_rows_f:
        row_num = item["row"]
        product = item["product"]
        
        print(f"\n  ▶ Sətir {row_num}: {product}")
        
        # Yalnız xüsusi 4 məhsul üçün işlə
        if product not in FORMA8_14_PRODUCTS:
            print(f"    ⊘ Forma8_14 hesablaması tələb etməyən məhsul, skip edilir")
            skipped_count_f += 1
            continue
        
        # Product Excel faylını tap
        product_file = os.path.join(processed_excels_folder, f"{product}.xlsx")
        
        if not os.path.exists(product_file):
            print(f"    ⚠ Excel tapılmadı: {product}.xlsx")
            continue
        
        try:
            # Product Excel-i aç (iki dəfə - format və dəyər üçün)
            wb_product_format = load_workbook(product_file, data_only=False)
            wb_product = load_workbook(product_file, data_only=True)
            
            # ==================== FORMA8_14-DƏN E22 ====================
            forma8_14_e22 = None
            
            if "Forma8_14" in wb_product.sheetnames:
                ws_8_14 = wb_product["Forma8_14"]
                ws_8_14_format = wb_product_format["Forma8_14"] if "Forma8_14" in wb_product_format.sheetnames else None
                
                cell_e22 = ws_8_14.cell(row=22, column=5)  # E22
                
                # DEBUG: Formatdan da yoxla
                if ws_8_14_format:
                    cell_e22_format = ws_8_14_format.cell(row=22, column=5)
                    print(f"    🔍 Debug Forma8_14 E22:")
                    print(f"       Format wb: {cell_e22_format.value} (type: {type(cell_e22_format.value)})")
                    print(f"       Data wb: {cell_e22.value} (type: {type(cell_e22.value)})")
                
                if cell_e22.value is not None:
                    try:
                        forma8_14_e22 = float(cell_e22.value)
                        print(f"    ✓ Forma8_14 E22: {forma8_14_e22:.2f}")
                    except (ValueError, TypeError):
                        print(f"    ⚠ Forma8_14 E22 rəqəm deyil: {cell_e22.value}")
                else:
                    # Əgər data_only=True ilə None gəlirsə, format wb-dən oxu
                    if ws_8_14_format:
                        cell_e22_format = ws_8_14_format.cell(row=22, column=5)
                        if cell_e22_format.value is not None:
                            try:
                                forma8_14_e22 = float(cell_e22_format.value)
                                print(f"    ✓ Forma8_14 E22: {forma8_14_e22:.2f} [format wb-dən]")
                            except (ValueError, TypeError):
                                print(f"    ⚠ Forma8_14 E22 rəqəm deyil: {cell_e22_format.value}")
                        else:
                            print(f"    ⚠ Forma8_14 E22 boşdur (hər iki wb-də)")
                    else:
                        print(f"    ⚠ Forma8_14 E22 boşdur")
            else:
                print(f"    ⚠ Forma8_14 sheet-i yoxdur")
            
            # ==================== E SÜTUNUNA YAZMA ====================
            if forma8_14_e22 is not None:
                ws.cell(row=row_num, column=5).value = forma8_14_e22
                print(f"    ✓ E{row_num} = {forma8_14_e22:.2f}")
            else:
                print(f"    ⚠ E{row_num} yazıla bilmədi (Forma8_14 E22 yoxdur)")
            
            wb_product.close()
            wb_product_format.close()
            success_count_f += 1
            
        except Exception as e:
            print(f"    ❌ Xəta: {e}")
    
    # ==================== C İLƏ BAŞLAYAN PRODUCTLAR ÜÇÜN MƏLUMAT DOLDURMA ====================
    print(f"\n{'='*60}")
    print(f"▶ C İLƏ BAŞLAYAN PRODUCTLAR (Forma8_3, 8_11)")
    print(f"{'='*60}")
    
    success_count_c = 0
    
    for item in product_rows_c:
        row_num = item["row"]
        product = item["product"]
        
        print(f"\n  ▶ Sətir {row_num}: {product}")
        
        # Product tipini müəyyən et
        is_type1 = product in TYPE1_PRODUCTS
        is_type2 = product in TYPE2_PRODUCTS
        
        if is_type1:
            target_row = 25
            print(f"    → Type1 məhsul: G25 istifadə ediləcək")
        elif is_type2:
            target_row = 33
            print(f"    → Type2 məhsul: G33 istifadə ediləcək")
        else:
            print(f"    ⚠ Məhsul tipi müəyyən edilə bilmədi, skip edilir")
            continue
        
        # Product Excel faylını tap
        product_file = os.path.join(processed_excels_folder, f"{product}.xlsx")
        
        if not os.path.exists(product_file):
            print(f"    ⚠ Excel tapılmadı: {product}.xlsx")
            continue
        
        try:
            # Product Excel-i aç (iki dəfə - format və dəyər üçün)
            wb_product_format = load_workbook(product_file, data_only=False)  # Format yoxlamaq üçün
            wb_product = load_workbook(product_file, data_only=True)  # Dəyər oxumaq üçün
            
            # ==================== FORMA8_3-DƏN G25/G33 ====================
            forma8_3_value = None
            
            if "Forma8_3" in wb_product.sheetnames:
                ws_8_3 = wb_product["Forma8_3"]
                ws_8_3_format = wb_product_format["Forma8_3"] if "Forma8_3" in wb_product_format.sheetnames else None
                
                cell_g = ws_8_3.cell(row=target_row, column=7)  # G sütunu (data_only=True)
                
                # DEBUG: Formatdan da yoxla
                if ws_8_3_format:
                    cell_g_format = ws_8_3_format.cell(row=target_row, column=7)
                    print(f"    🔍 Debug Forma8_3 G{target_row}:")
                    print(f"       Format wb: {cell_g_format.value} (type: {type(cell_g_format.value)})")
                    print(f"       Data wb: {cell_g.value} (type: {type(cell_g.value)})")
                
                if cell_g.value is not None:
                    try:
                        forma8_3_value = float(cell_g.value)
                        print(f"    ✓ Forma8_3 G{target_row}: {forma8_3_value:.2f}")
                    except (ValueError, TypeError):
                        print(f"    ⚠ Forma8_3 G{target_row} rəqəm deyil: {cell_g.value}")
                else:
                    # Əgər data_only=True ilə None gəlirsə, format wb-dən oxu
                    if ws_8_3_format:
                        cell_g_format = ws_8_3_format.cell(row=target_row, column=7)
                        if cell_g_format.value is not None:
                            try:
                                forma8_3_value = float(cell_g_format.value)
                                print(f"    ✓ Forma8_3 G{target_row}: {forma8_3_value:.2f} [format wb-dən]")
                            except (ValueError, TypeError):
                                print(f"    ⚠ Forma8_3 G{target_row} rəqəm deyil: {cell_g_format.value}")
                        else:
                            print(f"    ⚠ Forma8_3 G{target_row} boşdur (hər iki wb-də)")
                    else:
                        print(f"    ⚠ Forma8_3 G{target_row} boşdur")
            else:
                print(f"    ⚠ Forma8_3 sheet-i yoxdur")
            
            # ==================== E SÜTUNUNA YAZMA ====================
            if forma8_3_value is not None:
                ws.cell(row=row_num, column=5).value = forma8_3_value
                print(f"    ✓ E{row_num} = {forma8_3_value:.2f}")
            else:
                print(f"    ⚠ E{row_num} yazıla bilmədi (Forma8_3 G{target_row} yoxdur)")
            
            # ==================== FORMA8_11-DƏN G25/G33 ====================
            forma8_11_value = None
            
            if "Forma8_11" in wb_product.sheetnames:
                ws_8_11 = wb_product["Forma8_11"]
                ws_8_11_format = wb_product_format["Forma8_11"] if "Forma8_11" in wb_product_format.sheetnames else None
                
                cell_g = ws_8_11.cell(row=target_row, column=7)  # G sütunu (data_only=True)
                
                # DEBUG: Formatdan da yoxla
                if ws_8_11_format:
                    cell_g_format = ws_8_11_format.cell(row=target_row, column=7)
                    print(f"    🔍 Debug Forma8_11 G{target_row}:")
                    print(f"       Format wb: {cell_g_format.value} (type: {type(cell_g_format.value)})")
                    print(f"       Data wb: {cell_g.value} (type: {type(cell_g.value)})")
                
                if cell_g.value is not None:
                    try:
                        forma8_11_value = float(cell_g.value)
                        print(f"    ✓ Forma8_11 G{target_row}: {forma8_11_value:.2f}")
                    except (ValueError, TypeError):
                        print(f"    ⚠ Forma8_11 G{target_row} rəqəm deyil: {cell_g.value}")
                else:
                    # Əgər data_only=True ilə None gəlirsə, format wb-dən oxu
                    if ws_8_11_format:
                        cell_g_format = ws_8_11_format.cell(row=target_row, column=7)
                        if cell_g_format.value is not None:
                            try:
                                forma8_11_value = float(cell_g_format.value)
                                print(f"    ✓ Forma8_11 G{target_row}: {forma8_11_value:.2f} [format wb-dən]")
                            except (ValueError, TypeError):
                                print(f"    ⚠ Forma8_11 G{target_row} rəqəm deyil: {cell_g_format.value}")
                        else:
                            print(f"    ⚠ Forma8_11 G{target_row} boşdur (hər iki wb-də)")
                    else:
                        print(f"    ⚠ Forma8_11 G{target_row} boşdur")
            else:
                print(f"    ⚠ Forma8_11 sheet-i yoxdur")
            
            # ==================== F SÜTUNUNA YAZMA ====================
            if forma8_11_value is not None:
                ws.cell(row=row_num, column=6).value = forma8_11_value
                print(f"    ✓ F{row_num} = {forma8_11_value:.2f}")
            else:
                print(f"    ⚠ F{row_num} yazıla bilmədi (Forma8_11 G{target_row} yoxdur)")
            
            wb_product.close()
            wb_product_format.close()
            success_count_c += 1
            
        except Exception as e:
            print(f"    ❌ Xəta: {e}")
    
    # ==================== B İLƏ BAŞLAYAN PRODUCTLAR ÜÇÜN MƏLUMAT DOLDURMA ====================
    print(f"\n{'='*60}")
    print(f"▶ B İLƏ BAŞLAYAN PRODUCTLAR (Forma8_9, 8_13)")
    print(f"{'='*60}")
    
    success_count_b = 0
    
    for item in product_rows_b:
        row_num = item["row"]
        product = item["product"]
        
        print(f"\n  ▶ Sətir {row_num}: {product}")
        
        # Product Excel faylını tap
        product_file = os.path.join(processed_excels_folder, f"{product}.xlsx")
        
        if not os.path.exists(product_file):
            print(f"    ⚠ Excel tapılmadı: {product}.xlsx")
            continue
        
        try:
            # Product Excel-i aç
            wb_product = load_workbook(product_file, data_only=False)
            
            # ==================== FORMA8_9-DAN E16 ====================
            forma8_9_e16 = None
            
            if "Forma8_9" in wb_product.sheetnames:
                ws_8_9 = wb_product["Forma8_9"]
                
                cell_e16 = ws_8_9.cell(row=16, column=5)
                
                if cell_e16.value is not None:
                    try:
                        forma8_9_e16 = float(cell_e16.value)
                        print(f"    ✓ Forma8_9 E16: {forma8_9_e16:.2f}")
                    except (ValueError, TypeError):
                        print(f"    ⚠ Forma8_9 E16 rəqəm deyil: {cell_e16.value}")
                else:
                    print(f"    ⚠ Forma8_9 E16 boşdur")
            else:
                print(f"    ⚠ Forma8_9 sheet-i yoxdur")
            
            # ==================== E SÜTUNUNA YAZMA ====================
            if forma8_9_e16 is not None:
                ws.cell(row=row_num, column=5).value = forma8_9_e16
                print(f"    ✓ E{row_num} = {forma8_9_e16:.2f}")
            else:
                print(f"    ⚠ E{row_num} yazıla bilmədi (Forma8_9 E16 yoxdur)")
            
            # ==================== FORMA8_13-DƏN E16 ====================
            forma8_13_e16 = None
            
            if "Forma8_13" in wb_product.sheetnames:
                ws_8_13 = wb_product["Forma8_13"]
                
                cell_e16 = ws_8_13.cell(row=16, column=5)
                
                if cell_e16.value is not None:
                    try:
                        forma8_13_e16 = float(cell_e16.value)
                        print(f"    ✓ Forma8_13 E16: {forma8_13_e16:.2f}")
                    except (ValueError, TypeError):
                        print(f"    ⚠ Forma8_13 E16 rəqəm deyil: {cell_e16.value}")
                else:
                    print(f"    ⚠ Forma8_13 E16 boşdur")
            else:
                print(f"    ⚠ Forma8_13 sheet-i yoxdur")
            
            # ==================== F SÜTUNUNA YAZMA ====================
            if forma8_13_e16 is not None:
                ws.cell(row=row_num, column=6).value = forma8_13_e16
                print(f"    ✓ F{row_num} = {forma8_13_e16:.2f}")
            else:
                print(f"    ⚠ F{row_num} yazıla bilmədi (Forma8_13 E16 yoxdur)")
            
            wb_product.close()
            success_count_b += 1
            
        except Exception as e:
            print(f"    ❌ Xəta: {e}")
    
    # ==================== SAXLAMA ====================
    wb.save(output_file)
    wb.close()
    
    print("\n" + "=" * 60)
    print(f"✅ YEKUN RESERV PROSESİ TAMAMLANDI")
    print(f"   - A ilə başlayan: {success_count}/{len(product_rows)} product")
    print(f"   - B ilə başlayan: {success_count_b}/{len(product_rows_b)} product")
    print(f"   - C ilə başlayan: {success_count_c}/{len(product_rows_c)} product")
    print(f"   - F ilə başlayan: {success_count_f}/{len(product_rows_f)} product (skip: {skipped_count_f})")
    print(f"   - Fayl: {os.path.basename(output_file)}")
    print("=" * 60)

# Yekun Reserv prosesi
print("\n▶ Yekun Reserv işlənir...")
try:
    run_yekun_reserv(
        yekun_template=YEKUN_TEMPLATE,
        output_folder=OUTPUT_FOLDER,
        processed_excels_folder=OUTPUT_FOLDER,
        reference_date=REFERENCE_DATE
    )
    print("✅ Yekun Reserv tamamlandı")
except Exception as e:
    print(f"❌ Yekun Reserv xətası: {e}")