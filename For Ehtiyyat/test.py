import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

# =======================
# Fayl yolları
# =======================
source_file = r"C:\Users\samir.alizade\Desktop\Automantion\Risk\UcotA.xlsx"
output_file = r"C:\Users\samir.alizade\Desktop\Automantion\Risk\TEST.xlsx"

# =======================
# Parametrlər
# =======================
start_row = 13
end_row = 24
start_col_index = 5  # E sütunu

# =======================
# TEST.xlsx faylını aç
# =======================
wb = load_workbook(output_file)
ws = wb.active

# C8-dən product oxu
product = ws['C8'].value

# =======================
# Mənbə fayldan DataFrame-ləri oxu
# =======================
zerer_df = pd.read_excel(source_file, sheet_name='Zerer')
subraq_df = pd.read_excel(source_file, sheet_name='Subraqasiya')

# =======================
# Tarix sütunlarını datetime tipinə çevir
# =======================
zerer_df['Sığоrtа hаdisəsinin bаş verdiyi tarixi'] = pd.to_datetime(
    zerer_df['Sığоrtа hаdisəsinin bаş verdiyi tarixi'], errors='coerce'
)
zerer_df['Sığorta ödənişi Tаriхi'] = pd.to_datetime(
    zerer_df['Sığorta ödənişi Tаriхi'], errors='coerce'
)

subraq_df['Sığоrtа hаdisəsinin bаş vеrdiyi tаriх'] = pd.to_datetime(
    subraq_df['Sığоrtа hаdisəsinin bаş vеrdiyi tаriх'], errors='coerce'
)
subraq_df['Subroqasiya gəlirinin daxil olduğu tarix'] = pd.to_datetime(
    subraq_df['Subroqasiya gəlirinin daxil olduğu tarix'], errors='coerce'
)

# =======================
# Üçbucaq formatında doldur
# =======================
for row in range(start_row, end_row + 1):
    # Cari sətirin D tarixi (hadisə intervali üçün)
    D_cell_row_value = ws[f"D{row}"].value
    D_cell_row = pd.to_datetime(D_cell_row_value, errors='coerce')
    
    if pd.isna(D_cell_row):
        # Bu sətir üçün bütün sütunları X et
        for col_offset in range(12):  # E-dən P-yə qədər (12 sütun)
            ws.cell(row=row, column=start_col_index + col_offset, value='X')
        continue
    
    col_offset = 0
    
    # Sağa doğru: D13, D14, D15... (ödəniş tarixi üçün)
    for payment_row in range(row, end_row + 1):
        current_col = start_col_index + col_offset
        
        # Ödəniş tarixini oxu
        D_cell_payment_value = ws[f"D{payment_row}"].value
        D_cell_payment = pd.to_datetime(D_cell_payment_value, errors='coerce')
        
        if pd.isna(D_cell_payment):
            ws.cell(row=row, column=current_col, value='X')
            col_offset += 1
            continue
        
        # Zərər hesabla
        zerer_filter = (
            (zerer_df['Sığоrtа hаdisəsinin bаş verdiyi tarixi'] < D_cell_row) &
            (zerer_df['Sığоrtа hаdisəsinin bаş verdiyi tarixi'] >= D_cell_row - relativedelta(months=3)) &
            (zerer_df['Sığorta ödənişi Tаriхi'] < D_cell_payment) &
            (zerer_df['Sığоrtаnın sinifləri'] == product)
        )
        zerer_sum = zerer_df.loc[zerer_filter, 'Sığorta ödənişi Cəmi'].sum()
        
        # Subroqasiya hesabla
        subraq_filter = (
            (subraq_df['Sığоrtа hаdisəsinin bаş vеrdiyi tаriх'] < D_cell_row) &
            (subraq_df['Sığоrtа hаdisəsinin bаş vеrdiyi tаriх'] >= D_cell_row - relativedelta(months=3)) &
            (subraq_df['Subroqasiya gəlirinin daxil olduğu tarix'] < D_cell_payment) &
            (subraq_df['Sığоrtаnın sinifləri'] == product)
        )
        subraq_sum = subraq_df.loc[subraq_filter, 'Ödənilmiş subroqasiya gəlirinin məbləği'].sum()
        
        # Nəticə
        result = zerer_sum - subraq_sum
        
        # Excel-ə yaz
        ws.cell(row=row, column=current_col, value=result)
        
        col_offset += 1
    
    # ⬅️ Qalan sütunlara X yaz
    while col_offset < 12:  # E-dən P-yə qədər ümumi 12 sütun
        ws.cell(row=row, column=start_col_index + col_offset, value='X')
        col_offset += 1

# =======================
# Excel faylını yadda saxla
# =======================
wb.save(output_file)
print("✅ Üçbucaq formatı düzgün dolduruldu (X-lərlə birlikdə)!")