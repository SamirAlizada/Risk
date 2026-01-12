from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
import pandas as pd

# Rate konfiqurasiyası
PRODUCT_RATES = {
    "(04)AvtoKasko": 0.02,
    "(08)Yuk": 0.01,
    "(03)EmlakYanginDigerRisk": 0.20,
    "(37)IcbariDashinmazEmlak": 0.20
}

# Hesablama tələb edən məhsullar
CALCULATED_PRODUCTS = {
    "(04)AvtoKasko",
    "(08)Yuk",
    "(03)EmlakYanginDigerRisk",
    "(37)IcbariDashinmazEmlak"
}

def get_rate(product: str) -> float:
    """Product üçün rate qaytarır"""
    return PRODUCT_RATES.get(product, 0)

def run_forma8_2(excel_file: str, ucot_file: str, reference_date: str):
    """Forma8_2 doldurur: Forma8_1-dən məlumat götürüb hesablamalar aparır"""
    
    wb = load_workbook(excel_file)
    
    # Sheet-lərin mövcudluğunu yoxla
    if "Forma8_1" not in wb.sheetnames:
        raise ValueError(f"{excel_file} faylında 'Forma8_1' sheet-i yoxdur!")
    if "Forma8_2" not in wb.sheetnames:
        raise ValueError(f"{excel_file} faylında 'Forma8_2' sheet-i yoxdur!")
    
    ws_src = wb["Forma8_1"]
    ws_dst = wb["Forma8_2"]

    # Referans tarixini çevir
    ref_date = pd.to_datetime(reference_date) if reference_date else None

    # ================== UCOT OXUMA ==================
    df_ucot = pd.read_excel(ucot_file, sheet_name="Simple")
    
    # Lazımi sütunları götür və tarix çevirməsini et
    required_cols = ["II", "I", "IV", "V"]
    missing_cols = [col for col in required_cols if col not in df_ucot.columns]
    if missing_cols:
        raise ValueError(f"UCOT faylında bu sütunlar tapılmadı: {missing_cols}")
    
    df_ucot = df_ucot[required_cols].copy()
    df_ucot["IV"] = pd.to_datetime(df_ucot["IV"], errors='coerce')
    df_ucot["V"] = pd.to_datetime(df_ucot["V"], errors='coerce')

    # ================== STYLES ==================
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    font = Font(name="A3 Times AZ Lat", size=10)
    bold_font = Font(name="A3 Times AZ Lat", size=10, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    ROW_HEIGHT = 28

    # ================== FORMA8_1-DƏN DATA OXUMA ==================
    src_row = 12
    policy_data = []  # [(policy_number, f_value), ...]

    while True:
        val_b = ws_src.cell(row=src_row, column=2).value
        
        # Boş və ya yekun sətirində dayan
        if not val_b or (isinstance(val_b, str) and ("Yekun" in val_b or "AA" in val_b)):
            break
        
        val_f = ws_src.cell(row=src_row, column=6).value
        
        # Excel formulunu hesabla
        if isinstance(val_f, str) and val_f.startswith('='):
            val_f = ws_src.cell(row=src_row, column=6).value
        
        # Numeric dəyərə çevir
        try:
            val_f = float(val_f) if val_f is not None else 0
        except (ValueError, TypeError):
            val_f = 0
        
        policy_data.append((val_b, val_f))
        src_row += 1

    data_count = len(policy_data)
    
    if data_count == 0:
        print(f"  ⚠ Forma8_1-də heç bir data tapılmadı")
        wb.save(excel_file)
        return 0.0  # ✅ 0 return et

    print(f"  📊 {data_count} policy tapıldı")

    # ================== MERGE-LƏRİ AÇ ==================
    for m in list(ws_dst.merged_cells.ranges):
        ws_dst.unmerge_cells(str(m))
    
    # ================== MƏHSUL ADINI YAZ (C8) ==================
    first_policy = policy_data[0][0]
    ucot_row = df_ucot[df_ucot["II"] == first_policy]
    
    if not ucot_row.empty:
        product = ucot_row.iloc[0]["I"]
        ws_dst["C8"].value = product
        ws_dst["C8"].font = font
        ws_dst["C8"].alignment = center
        print(f"  📦 Product: {product}")
    else:
        product = None
        print(f"  ⚠ Product məlumatı tapılmadı")

    # ================== D6-YA TARİX YAZMA ==================
    ref_date = pd.to_datetime(reference_date)
    formatted_date = ref_date.strftime("%d.%m.%Y")
    ws_dst["D6"].value = formatted_date
    print(f"    ✓ D6-ya tarix yazıldı: {formatted_date}")

    # ================== SƏTIR ƏLAVƏSİ ==================
    start_row = 12
    ws_dst.insert_rows(start_row, amount=data_count)

    # ================== DATA YAZIMI ==================
    r = start_row
    total_f_calculated = 0  # ✅ Manual hesablama

    for counter, (policy_num, f_value) in enumerate(policy_data, start=1):
        
        # A sütunu - sıra nömrəsi
        cell_a = ws_dst.cell(row=r, column=1)
        cell_a.value = f"A{counter}"
        cell_a.font = font
        cell_a.border = thin
        cell_a.alignment = center

        # B sütunu - policy nömrəsi
        cell_b = ws_dst.cell(row=r, column=2)
        cell_b.value = policy_num
        cell_b.font = font
        cell_b.border = thin
        cell_b.alignment = center

        # C sütunu - Forma8_1-dən gələn F dəyəri
        cell_c = ws_dst.cell(row=r, column=3)
        cell_c.value = f_value
        cell_c.font = font
        cell_c.border = thin
        cell_c.alignment = center

        # UCOT-dan məlumat götür
        ucot_row = df_ucot[df_ucot["II"] == policy_num]
        
        if not ucot_row.empty:
            iv_date = ucot_row.iloc[0]["IV"]
            v_date = ucot_row.iloc[0]["V"]
            product = ucot_row.iloc[0]["I"]

            # D sütunu - müddət (gün sayı)
            if pd.notna(iv_date) and pd.notna(v_date):
                day_diff = (v_date - iv_date).days
                d_value = day_diff + 1
            else:
                d_value = 1
            
            # E sütunu - referans tarixə qədər olan günlər
            if ref_date and pd.notna(iv_date):
                e_value = (ref_date - iv_date).days
            else:
                e_value = None
        else:
            d_value = 1
            e_value = None
            product = None

        # D sütununu yaz
        cell_d = ws_dst.cell(row=r, column=4)
        cell_d.value = d_value
        cell_d.font = font
        cell_d.border = thin
        cell_d.alignment = center

        # E sütununu yaz (varsa)
        if e_value is not None:
            cell_e = ws_dst.cell(row=r, column=5)
            cell_e.value = e_value
            cell_e.font = font
            cell_e.border = thin
            cell_e.alignment = center

        # F sütunu - formula YAZ
        cell_f = ws_dst.cell(row=r, column=6)
        cell_f.value = f"=ROUND((D{r}-E{r})/D{r}*C{r},2)"
        cell_f.font = font
        cell_f.border = thin
        cell_f.alignment = center
        
        # ✅ F-i manual hesabla
        if e_value is not None and d_value > 0:
            f_calc = round((d_value - e_value) / d_value * f_value, 2)
            total_f_calculated += f_calc

        # G sütunu - formula: C * rate
        rate = get_rate(product) if product else 0
        cell_g = ws_dst.cell(row=r, column=7)
        cell_g.value = f"=ROUND(C{r}*{rate},2)"
        cell_g.font = font
        cell_g.border = thin
        cell_g.alignment = center

        # H sütunu - formula (yalnız müəyyən məhsullar üçün)
        cell_h = ws_dst.cell(row=r, column=8)
        if product in CALCULATED_PRODUCTS:
            cell_h.value = f"=ROUND((D{r}-E{r})/D{r}*G{r},2)"
        else:
            cell_h.value = 0
        cell_h.font = font
        cell_h.border = thin
        cell_h.alignment = center

        ws_dst.row_dimensions[r].height = ROW_HEIGHT
        r += 1

    # ================== TOPLAM SƏTİRİ ==================
    total_row = r
    
    # A və B sütunları
    cell_total_a = ws_dst.cell(row=total_row, column=1)
    cell_total_a.value = "AA1"
    cell_total_a.font = bold_font
    cell_total_a.border = thin
    cell_total_a.alignment = center
    
    cell_total_b = ws_dst.cell(row=total_row, column=2)
    cell_total_b.value = "Yekun BSH"
    cell_total_b.font = bold_font
    cell_total_b.border = thin
    cell_total_b.alignment = center

    # C toplam
    cell_total_c = ws_dst.cell(row=total_row, column=3)
    cell_total_c.value = f"=ROUND(SUM(C{start_row}:C{total_row-1}),2)"
    cell_total_c.font = bold_font
    cell_total_c.border = thin
    cell_total_c.alignment = center

    # F toplam
    cell_total_f = ws_dst.cell(row=total_row, column=6)
    cell_total_f.value = f"=ROUND(SUM(F{start_row}:F{total_row-1}),2)"
    cell_total_f.font = bold_font
    cell_total_f.border = thin
    cell_total_f.alignment = center

    # G toplam
    cell_total_g = ws_dst.cell(row=total_row, column=7)
    cell_total_g.value = f"=ROUND(SUM(G{start_row}:G{total_row-1}),2)"
    cell_total_g.font = bold_font
    cell_total_g.border = thin
    cell_total_g.alignment = center

    # H toplam
    cell_total_h = ws_dst.cell(row=total_row, column=8)
    cell_total_h.value = f"=ROUND(SUM(H{start_row}:H{total_row-1}),2)"
    cell_total_h.font = bold_font
    cell_total_h.border = thin
    cell_total_h.alignment = center

    # ================== 4 SƏTİR AŞAĞIYA KOPYALAMA ==================
    copy_row = total_row + 4
    
    # F kopyası
    cell_f_copy = ws_dst.cell(row=copy_row, column=6)
    cell_f_copy.value = f"=F{total_row}"
    cell_f_copy.font = bold_font
    cell_f_copy.border = thin
    cell_f_copy.alignment = center

    # H kopyası
    cell_h_copy = ws_dst.cell(row=copy_row, column=8)
    cell_h_copy.value = f"=H{total_row}"
    cell_h_copy.font = bold_font
    cell_h_copy.border = thin
    cell_h_copy.alignment = center

    wb.save(excel_file)
    wb.close()
    
    print(f"  ✅ Forma8_2 tamamlandı - Total F: {total_f_calculated:.2f}")
    
    return total_f_calculated  # ✅ Hesablanmış dəyəri return et