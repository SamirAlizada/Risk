from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
import pandas as pd

# Rate konfiqurasiyası (G sütunu üçün)
PRODUCT_RATES = {
    "(04)AvtoKasko": 0.01,
    "(08)Yuk": 0.01,
    "(03)EmlakYanginDigerRisk": 0.20,
    "(37)IcbariDashinmazEmlak": 0.20
}

def get_rate(product: str) -> float:
    """Product üçün rate qaytarır"""
    return PRODUCT_RATES.get(product, 0)

def run_forma8_5(excel_file: str, ucot_file: str, reference_date: str):
    """Forma8_5 doldurur: Forma8_4-dən data götürüb yeni hesablamalar aparır"""
    
    wb = load_workbook(excel_file)
    
    # Sheet-lərin mövcudluğunu yoxla
    if "Forma8_1" not in wb.sheetnames:
        raise ValueError(f"{excel_file} faylında 'Forma8_1' sheet-i yoxdur!")
    if "Forma8_4" not in wb.sheetnames:
        raise ValueError(f"{excel_file} faylında 'Forma8_4' sheet-i yoxdur!")
    if "Forma8_5" not in wb.sheetnames:
        raise ValueError(f"{excel_file} faylında 'Forma8_5' sheet-i yoxdur!")
    
    ws_8_1 = wb["Forma8_1"]
    ws_8_4 = wb["Forma8_4"]
    ws = wb["Forma8_5"]
    
    # Product adını Forma8_1-dən götür
    product = ws_8_1["C8"].value if ws_8_1["C8"].value else None
    
    if not product:
        print(f"  ⚠ {excel_file}: Forma8_1-də product məlumatı yoxdur (C8 boşdur)")
        wb.save(excel_file)
        return
    
    # Forma8_5-in C8-nə yaz
    ws["C8"].value = product
    ws["C8"].font = Font(name="A3 Times AZ Lat", size=10)
    ws["C8"].alignment = Alignment(horizontal="center", vertical="center")

    # ================== D6-YA TARİX YAZMA ==================
    ref_date = pd.to_datetime(reference_date)
    formatted_date = ref_date.strftime("%d.%m.%Y")
    ws["D6"].value = formatted_date
    print(f"    ✓ D6-ya tarix yazıldı: {formatted_date}")
    
    # ================== UCOT OXUMA ==================
    df_ucot = pd.read_excel(ucot_file, sheet_name="Simple")
    
    # Lazımi sütunları yoxla
    required_cols = ["I", "XXXVIII", "XL", "XLI"]
    missing_cols = [col for col in required_cols if col not in df_ucot.columns]
    if missing_cols:
        raise ValueError(f"UCOT faylında bu sütunlar tapılmadı: {missing_cols}")
    
    # Tarix sütunlarını çevir
    df_ucot["XL"] = pd.to_datetime(df_ucot["XL"], errors='coerce')
    df_ucot["XLI"] = pd.to_datetime(df_ucot["XLI"], errors='coerce')
    
    # Referans tarixi
    ref_date = pd.to_datetime(reference_date) if reference_date else None
    
    # ================== FORMA8_4-DƏN DATA OXUMA ==================
    src_row = 12
    data_rows = []
    
    print(f"    Debug - Forma8_4-dən oxuma:")
    
    while True:
        val_a = ws_8_4.cell(row=src_row, column=1).value
        val_b = ws_8_4.cell(row=src_row, column=2).value
        
        # Boş sətirdə dayan
        if not val_a and not val_b:
            break
        
        # ✅ Skip etməli sətirləri müəyyən et
        skip = False
        
        # A sütununda "qrup təkrarsığortaçılar" varsa skip
        if val_a and isinstance(val_a, str) and "qrup" in val_a.lower():
            skip = True
        
        # B sütununda "qrup təkrarsığortaçılar" varsa skip
        if val_b and isinstance(val_b, str) and "qrup" in val_b.lower():
            skip = True
        
        # A sütununda toplam sətirləri (AA1, BB1, CC1, DD1, ZZ1)
        if val_a and isinstance(val_a, str):
            if val_a in ["AA1", "BB1", "CC1", "DD1", "ZZ1"] or "yekun" in val_a.lower() or "aralıq" in val_a.lower():
                skip = True
        
        # B sütununda "Yekun" və ya "Aralıq" varsa skip
        if val_b and isinstance(val_b, str):
            if "yekun" in val_b.lower() or "aralıq" in val_b.lower():
                skip = True
        
        if skip:
            src_row += 1
            continue
        
        # ✅ Valid data sətri
        if val_b:
            val_f = ws_8_4.cell(row=src_row, column=6).value
            val_f = val_f if val_f is not None else 0
            
            data_rows.append((val_b, val_f, src_row))
        
        src_row += 1
    
    if not data_rows:
        print(f"  ⚠ {product}: Forma8_4-də data yoxdur")
        wb.save(excel_file)
        return
    
    print(f"    Forma8_4-dən {len(data_rows)} sətir tapıldı")
    
    # ================== STYLES ==================
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    font = Font(name="A3 Times AZ Lat", size=10)
    bold_font = Font(name="A3 Times AZ Lat", size=10, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    ROW_HEIGHT = 28
    
    # ================== MERGE-LƏRİ AÇ ==================
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))
    
    # ================== QRUPLARA BÖLÜŞDÜRMƏ ==================
    groups = {1: [], 2: [], 3: [], 4: []}
    
    print(f"    Debug - Qruplara bölüşdürmə:")
    
    for b_val, f_val, row_num in data_rows:
        a_val = ws_8_4.cell(row=row_num, column=1).value
        
        if a_val and isinstance(a_val, str):
            # A ile başlayanlar - Qrup 1
            if a_val.startswith("A") and len(a_val) <= 3 and a_val[1:].isdigit():
                groups[1].append((b_val, f_val))
                print(f"      {a_val} = {b_val} -> Qrup 1")
            # B ile başlayanlar - Qrup 2
            elif a_val.startswith("B") and len(a_val) <= 3 and a_val[1:].isdigit():
                groups[2].append((b_val, f_val))
                print(f"      {a_val} = {b_val} -> Qrup 2")
            # C ile başlayanlar - Qrup 3
            elif a_val.startswith("C") and len(a_val) <= 3 and a_val[1:].isdigit():
                groups[3].append((b_val, f_val))
                print(f"      {a_val} = {b_val} -> Qrup 3")
            # D ile başlayanlar - Qrup 4
            elif a_val.startswith("D") and len(a_val) <= 3 and a_val[1:].isdigit():
                groups[4].append((b_val, f_val))
                print(f"      {a_val} = {b_val} -> Qrup 4")
    
    # Her qrupun sayını göster
    print(f"    Qrup nəticələri:")
    for g_num in [1, 2, 3, 4]:
        print(f"      Qrup {g_num}: {len(groups[g_num])} sətir")
    
    # ✅ Boş qrupları çıxarmırıq - hamısını saxlayırıq
    # groups = {k: v for k, v in groups.items() if v}  # ❌ Bu sətr artıq lazım deyil
    
    # ================== QRUPLARI YAZMAQ ==================
    current_row = 12
    rate = get_rate(product)
    
    # Ümumi toplamlar
    grand_totals = {"C": 0.0, "F": 0.0, "G": 0.0, "H": 0.0}
    
    # Qrup prefix-ləri
    group_prefix = {1: "A", 2: "B", 3: "C", 4: "D"}
    
    # ✅ HƏMIŞƏ 4 QRUPU İŞLƏ (1, 2, 3, 4)
    for group_num in [1, 2, 3, 4]:
        group_data = groups[group_num]
        group_size = len(group_data)
        
        if group_size > 0:
            print(f"    Qrup {group_num}: {group_size} sətir yazılır...")
        else:
            print(f"    Qrup {group_num}: boşdur (0 ilə toplam yazılır)")
        
        # Qrup toplamları
        group_totals = {"C": 0.0, "F": 0.0, "G": 0.0, "H": 0.0}
        
        # ✅ Əgər qrupda data varsa
        if group_size > 0:
            # Hər qrup üçün ayrıca insert rows
            ws.insert_rows(current_row, amount=group_size)
            
            for idx, (b_val, f_val) in enumerate(group_data, start=1):
                r = current_row + idx - 1
                
                # A sütunu
                cell_a = ws.cell(row=r, column=1)
                prefix = group_prefix[group_num]
                cell_a.value = f"{prefix}{idx}"
                cell_a.font = font
                cell_a.border = thin
                cell_a.alignment = center
                
                # B sütunu
                cell_b = ws.cell(row=r, column=2)
                cell_b.value = b_val
                cell_b.font = font
                cell_b.border = thin
                cell_b.alignment = center
                
                # C sütunu
                val_c = round(float(f_val), 2)
                cell_c = ws.cell(row=r, column=3)
                cell_c.value = val_c
                cell_c.font = font
                cell_c.border = thin
                cell_c.alignment = center
                
                # UCOT-dan məlumat tap
                ucot_row = df_ucot[df_ucot["XXXVIII"] == b_val]
                
                if not ucot_row.empty:
                    xl_date = ucot_row.iloc[0]["XL"]
                    xli_date = ucot_row.iloc[0]["XLI"]
                    
                    if pd.notna(xl_date) and pd.notna(xli_date):
                        day_diff = (xli_date - xl_date).days
                        val_d = day_diff + 1
                    else:
                        val_d = 1
                    
                    if ref_date and pd.notna(xl_date):
                        e_day_diff = (ref_date - xl_date).days
                        val_e = e_day_diff + 1
                    else:
                        val_e = None
                else:
                    val_d = 1
                    val_e = None
                
                # D sütunu
                cell_d = ws.cell(row=r, column=4)
                cell_d.value = val_d
                cell_d.font = font
                cell_d.border = thin
                cell_d.alignment = center
                
                # E sütunu
                if val_e is not None:
                    cell_e = ws.cell(row=r, column=5)
                    cell_e.value = val_e
                    cell_e.font = font
                    cell_e.border = thin
                    cell_e.alignment = center
                
                # F sütunu
                if val_e is not None and val_d > 0:
                    val_f = round((val_d - val_e) / val_d * val_c, 2)
                else:
                    val_f = val_c
                
                cell_f = ws.cell(row=r, column=6)
                cell_f.value = val_f
                cell_f.font = font
                cell_f.border = thin
                cell_f.alignment = center
                
                # G sütunu
                val_g = round(val_c * rate, 2)
                cell_g = ws.cell(row=r, column=7)
                cell_g.value = val_g
                cell_g.font = font
                cell_g.border = thin
                cell_g.alignment = center
                
                # H sütunu
                if val_e is not None and val_d > 0:
                    val_h = round((val_d - val_e) / val_d * val_g, 2)
                else:
                    val_h = val_g
                
                cell_h = ws.cell(row=r, column=8)
                cell_h.value = val_h
                cell_h.font = font
                cell_h.border = thin
                cell_h.alignment = center
                
                # Toplamları yığ
                group_totals["C"] += val_c
                group_totals["F"] += val_f
                group_totals["G"] += val_g
                group_totals["H"] += val_h
                
                ws.row_dimensions[r].height = ROW_HEIGHT
        
        # ✅ Qrup toplam sətri (data olsa da, olmasada - həmişə yazılır)
        total_row = current_row + group_size
        
        if group_size > 0:
            print(f"      ✓ Qrup {group_num} toplam: C={group_totals['C']:.2f}, F={group_totals['F']:.2f}")
        
        # C toplam
        cell_total_c = ws.cell(row=total_row, column=3)
        cell_total_c.value = round(group_totals["C"], 2)
        cell_total_c.font = bold_font
        cell_total_c.border = thin
        cell_total_c.alignment = center
        
        # F toplam
        cell_total_f = ws.cell(row=total_row, column=6)
        cell_total_f.value = round(group_totals["F"], 2)
        cell_total_f.font = bold_font
        cell_total_f.border = thin
        cell_total_f.alignment = center
        
        # G toplam
        cell_total_g = ws.cell(row=total_row, column=7)
        cell_total_g.value = round(group_totals["G"], 2)
        cell_total_g.font = bold_font
        cell_total_g.border = thin
        cell_total_g.alignment = center
        
        # H toplam
        cell_total_h = ws.cell(row=total_row, column=8)
        cell_total_h.value = round(group_totals["H"], 2)
        cell_total_h.font = bold_font
        cell_total_h.border = thin
        cell_total_h.alignment = center
        
        # Ümumi toplama əlavə et
        grand_totals["C"] += group_totals["C"]
        grand_totals["F"] += group_totals["F"]
        grand_totals["G"] += group_totals["G"]
        grand_totals["H"] += group_totals["H"]
        
        ws.row_dimensions[total_row].height = ROW_HEIGHT
        
        # Növbəti qrup üçün 2 sətir aşağıdan başla
        current_row += group_size + 2
    
    # ✅ ÜMUMİ CƏM (4-cü qrupdan sonra - həmişə)
    # current_row indi 4-cü qrupun toplamından 2 sətir aşağıdadır
    # Biz 1 sətir aşağı istəyirik, ona görə -1 edirik
    grand_total_row = current_row - 1
    
    print(f"    ÜMUMİ CƏM (sətir {grand_total_row}): C={grand_totals['C']:.2f}, F={grand_totals['F']:.2f}, G={grand_totals['G']:.2f}, H={grand_totals['H']:.2f}")
    
    # C ümumi cəm
    cell_grand_c = ws.cell(row=grand_total_row, column=3)
    cell_grand_c.value = round(grand_totals["C"], 2)
    cell_grand_c.font = bold_font
    cell_grand_c.border = thin
    cell_grand_c.alignment = center
    
    # F ümumi cəm
    cell_grand_f = ws.cell(row=grand_total_row, column=6)
    cell_grand_f.value = round(grand_totals["F"], 2)
    cell_grand_f.font = bold_font
    cell_grand_f.border = thin
    cell_grand_f.alignment = center
    
    # G ümumi cəm
    cell_grand_g = ws.cell(row=grand_total_row, column=7)
    cell_grand_g.value = round(grand_totals["G"], 2)
    cell_grand_g.font = bold_font
    cell_grand_g.border = thin
    cell_grand_c.alignment = center
    
    # H ümumi cəm
    cell_grand_h = ws.cell(row=grand_total_row, column=8)
    cell_grand_h.value = round(grand_totals["H"], 2)
    cell_grand_h.font = bold_font
    cell_grand_h.border = thin
    cell_grand_h.alignment = center
    
    ws.row_dimensions[grand_total_row].height = ROW_HEIGHT
    
    wb.save(excel_file)
    print(f"  ✅ {product}: 4 qrup işləndi")