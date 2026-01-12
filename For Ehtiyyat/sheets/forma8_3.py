from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta

# Forma8_3(1) istifadə edən məhsullar (3 il - 12 period)
FORMA8_3_TYPE1_PRODUCTS = [
    "(01)FerdiQeza",
    "(02)Tibbi",
    "(03)EmlakYanginDigerRisk",
    "(04)AvtoKasko",
    "(05)DemiryolNeqliyyVasitesi",
    "(06)HavaNeqliyyKasko",
    "(07)SuNeqliyyKasko",
    "(08)Yuk",
    "(09)KendTeserrufBitki",
    "(10)KendTeserrufHeyvan",
    "(11)IshcilerinDeleduzlug",
    "(12)PulvePulSenedSaxtalash",
    "(22)Kredit",
    "(23)Ipoteka",
    "(24)EmlakinDeyerdenDushmesi",
    "(25)IshinDayanmasiRiski",
    "(27)SernishinIcbari",
    "(28)IcbariEkoloji",
    "(29)YanginIcbari",
    "(30)DeputatlarinIcbari",
    "(31)TibbiPersonalinAIDSden",
    "(32)HerbiQulluqcularinIcbari",
    "(33)HuquqMuhafifeIcbari",
    "(34)DovletQulluqcuIcbari",
    "(35)DiplomatlarinIcbari",
    "(37)IcbariDashinmazEmlak",
    "(39)IcbariNVSMMS",
    "(40)IcbariSernishinFerdiQeza",
    "(41)Sefer",
    "(42)Titul",
    "(43)HuquqiXerc"
]

# Forma8_3(2) istifadə edən məhsullar (5 il - 20 period)
FORMA8_3_TYPE2_PRODUCTS = [
    "(19)PesheMesuliyy",
    "(20)IshegoturenMesuliyy",
    "(21)UmumiMulkiMesuliyy",
    "(13)AvtoKonulluMesuliyy",
    "(14)DemiryolNeqliySahibMesuliyy",
    "(15)HavaNeqliySahibMesuliyy",
    "(16)SuNeqliySahibMesuliyy",
    "(17)YukDashiyanMesuliyy",
    "(18)MulkiMuqavileUzreMesuliyy",
    "(26)AvtoIcbariMesuliyy",
    "(36)AuditorPesheMesuliyyIcbari",
    "(38)IcbariDashinmazEmlakMesul"
]

def run_forma8_3(excel_file: str, ucot_file: str, reference_date: str):
    """Forma8_3 doldurur: Product-a görə düzgün sheet-i seçib Borc sheet-indən məlumat götürür"""
    
    wb = load_workbook(excel_file)
    
    # Sheet-lərin mövcudluğunu yoxla
    if "Forma8_1" not in wb.sheetnames:
        raise ValueError(f"{excel_file} faylında 'Forma8_1' sheet-i yoxdur!")
    
    ws_8_1 = wb["Forma8_1"]
    
    # Product adını Forma8_1-dən götür
    product = ws_8_1["C8"].value if ws_8_1["C8"].value else None
    
    if not product:
        print(f"  ⚠ {excel_file}: Forma8_1-də product məlumatı yoxdur (C8 boşdur)")
        wb.save(excel_file)
        return
    
    print(f"    Product: {product}")
    
    # ================== PRODUCT-A GÖRƏ TİP TƏYİN ET ==================
    use_type1 = product in FORMA8_3_TYPE1_PRODUCTS
    use_type2 = product in FORMA8_3_TYPE2_PRODUCTS
    
    if not use_type1 and not use_type2:
        print(f"  ⚠ {product}: Forma8_3 tip məlumatı tapılmadı, skip edilir")
        wb.save(excel_file)
        return
    
    if use_type1:
        print(f"    → Type1: Forma8_3(1) - 3 il (12 period)")
        sheet_to_keep = "Forma8_3(1)"
        sheet_to_delete = "Forma8_3(2)"
        num_periods = 12
        start_row = 24
        total_row = 25
    else:  # use_type2
        print(f"    → Type2: Forma8_3(2) - 5 il (20 period)")
        sheet_to_keep = "Forma8_3(2)"
        sheet_to_delete = "Forma8_3(1)"
        num_periods = 20
        start_row = 32
        total_row = 33
    
    # ================== SHEET-LƏRİ İDARƏ ET ==================
    if sheet_to_delete in wb.sheetnames:
        del wb[sheet_to_delete]
        print(f"    ✓ {sheet_to_delete} silindi")
    
    if sheet_to_keep not in wb.sheetnames:
        print(f"  ⚠ {sheet_to_keep} sheet-i tapılmadı!")
        wb.save(excel_file)
        return
    
    ws = wb[sheet_to_keep]
    ws.title = "Forma8_3"
    print(f"    ✓ {sheet_to_keep} → Forma8_3 adına dəyişdirildi")
    
    # ================== C8-Ə PRODUCT YAZMA ==================
    ws["C8"].value = product
    ws["C8"].font = Font(name="A3 Times AZ Lat", size=14)
    ws["C8"].alignment = Alignment(horizontal="center", vertical="center")
    
    # ================== UCOT-DAN BORC SHEET OXUMA ==================
    print(f"    Borc sheet-indən məlumat oxuyur...")
    
    try:
        df_borc = pd.read_excel(ucot_file, sheet_name="Borc")
    except Exception as e:
        print(f"    ⚠ Borc sheet oxuna bilmədi: {e}")
        wb.save(excel_file)
        return
    
    required_cols = ["I", "II", "III"]
    missing_cols = [col for col in required_cols if col not in df_borc.columns]
    if missing_cols:
        print(f"    ⚠ Borc sheet-ində bu sütunlar tapılmadı: {missing_cols}")
        wb.save(excel_file)
        return
    
    df_product = df_borc[df_borc["I"] == product].copy()
    
    if df_product.empty:
        print(f"    ⚠ Borc sheet-ində {product} üçün data yoxdur")
        df_product = pd.DataFrame(columns=["I", "II", "III"])
    
    if not df_product.empty:
        df_product["II"] = pd.to_datetime(df_product["II"], errors='coerce')
        df_product["III"] = pd.to_numeric(df_product["III"], errors='coerce').fillna(0)
    
    ref_date = pd.to_datetime(reference_date)
    
    # ================== STYLES ==================
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    font_style = Font(name="A3 Times AZ Lat", size=14)
    bold_font = Font(name="A3 Times AZ Lat", size=14, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    
    # ================== D SÜTUNU - PERIOD HESABLAMA ==================
    print(f"    {num_periods} period hesablanır (D{start_row} → D13)...")
    
    grand_total_d = 0.0
    
    for i in range(num_periods):
        period_end = ref_date - relativedelta(months=i*3)
        period_start = period_end - relativedelta(months=3)
        
        if not df_product.empty:
            mask = (df_product["II"] >= period_start) & (df_product["II"] < period_end)
            period_sum = df_product.loc[mask, "III"].sum()
        else:
            period_sum = 0
        
        row = start_row - i
        ws.cell(row=row, column=4).value = round(period_sum, 2)
        ws.cell(row=row, column=4).font = font_style
        ws.cell(row=row, column=4).border = thin
        ws.cell(row=row, column=4).alignment = center
        
        grand_total_d += period_sum
    
    # D sütunu toplam
    ws.cell(row=total_row, column=4).value = round(grand_total_d, 2)
    ws.cell(row=total_row, column=4).font = bold_font
    ws.cell(row=total_row, column=4).border = thin
    ws.cell(row=total_row, column=4).alignment = center
    
    print(f"    ✓ D sütunu toplam D{total_row} = {grand_total_d:.2f}")
    
    # ================== F VƏ G SÜTUNLARI HESABLAMA ==================
    print(f"    F və G sütunları hesablanır...")
    
    grand_total_f = 0.0
    grand_total_g = 0.0
    
    for i in range(num_periods):
        row = start_row - i
        
        # D və E dəyərlərini götür (E-ni yalnız oxuyuruq)
        val_d = ws.cell(row=row, column=4).value or 0
        val_e = ws.cell(row=row, column=5).value or 0
        
        # F = (D + E) * 0.03
        val_f = round((val_d + val_e) * 0.03, 2)
        ws.cell(row=row, column=6).value = val_f
        ws.cell(row=row, column=6).font = font_style
        ws.cell(row=row, column=6).border = thin
        ws.cell(row=row, column=6).alignment = center
        
        # G = D + E + F
        val_g = round(val_d + val_e + val_f, 2)
        ws.cell(row=row, column=7).value = val_g
        ws.cell(row=row, column=7).font = font_style
        ws.cell(row=row, column=7).border = thin
        ws.cell(row=row, column=7).alignment = center
        
        grand_total_f += val_f
        grand_total_g += val_g
    
    # F sütunu toplam
    ws.cell(row=total_row, column=6).value = round(grand_total_f, 2)
    ws.cell(row=total_row, column=6).font = bold_font
    ws.cell(row=total_row, column=6).border = thin
    ws.cell(row=total_row, column=6).alignment = center
    
    # G sütunu toplam
    ws.cell(row=total_row, column=7).value = round(grand_total_g, 2)
    ws.cell(row=total_row, column=7).font = bold_font
    ws.cell(row=total_row, column=7).border = thin
    ws.cell(row=total_row, column=7).alignment = center
    
    print(f"    ✓ F sütunu toplam F{total_row} = {grand_total_f:.2f}")
    print(f"    ✓ G sütunu toplam G{total_row} = {grand_total_g:.2f}")
    
    wb.save(excel_file)
    print(f"  ✅ {product}: Forma8_3 tamamlandı")