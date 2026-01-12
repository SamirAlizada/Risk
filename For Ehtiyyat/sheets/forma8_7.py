from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os
import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta

# Forma8_7(1) istifadə edən məhsullar
FORMA8_7_TYPE1_PRODUCTS = [
    "(01)FerdiQeza",
    "(02)Tibbi",
    "(03)EmlakYanginDigerRisk",
    "(04)AvtoKasko",
    "(05)DemiryolNeqliyyVasitesi",
    "(06)HavaNeqliyyKasko",
    "(07)SuNeqliyyKasko",
    "(08)Yuk",
    "(09)KendTeserrufBitki",
    "(10)KendTeserrufHeyvan",
    "(11)IshcilerinDeleduzlug",
    "(12)PulvePulSenedSaxtalash",
    "(22)Kredit",
    "(23)Ipoteka",
    "(24)EmlakinDeyerdenDushmesi",
    "(25)IshinDayanmasiRiski",
    "(27)SernishinIcbari",
    "(28)IcbariEkoloji",
    "(29)YanginIcbari",
    "(30)DeputatlarinIcbari",
    "(31)TibbiPersonalinAIDSden",
    "(32)HerbiQulluqcularinIcbari",
    "(33)HuquqMuhafifeIcbari",
    "(34)DovletQulluqcuIcbari",
    "(35)DiplomatlarinIcbari",
    "(37)IcbariDashinmazEmlak",
    "(39)IcbariNVSMMS",
    "(40)IcbariSernishinFerdiQeza",
    "(41)Sefer",
    "(42)Titul",
    "(43)HuquqiXerc"
]

# Forma8_7(2) istifadə edən məhsullar
FORMA8_7_TYPE2_PRODUCTS = [
    "(19)PesheMesuliyy",
    "(20)IshegoturenMesuliyy",
    "(21)UmumiMulkiMesuliyy",
    "(13)AvtoKonulluMesuliyy",
    "(14)DemiryolNeqliySahibMesuliyy",
    "(15)HavaNeqliySahibMesuliyy",
    "(16)SuNeqliySahibMesuliyy",
    "(17)YukDashiyanMesuliyy",
    "(18)MulkiMuqavileUzreMesuliyy",
    "(26)AvtoIcbariMesuliyy",
    "(36)AuditorPesheMesuliyyIcbari",
    "(38)IcbariDashinmazEmlakMesul"
]

# 0.01 əmsalı tətbiq ediləcək məhsullar
COEFFICIENT_0_01_PRODUCTS = [
    "(04)AvtoKasko",
    "(08)Yuk",
    "(03)EmlakYanginDigerRisk",
    "(37)IcbariDashinmazEmlak"
]

def run_forma8_7(excel_file: str, previous_folder: str, reference_date: str, total_f_from_forma8_2: float = None):
    """Forma8_7 hazırlayır: Product-a görə düzgün sheet-i seçir və əvvəlki datanı kopyalayır"""
    
    wb = load_workbook(excel_file)
    
    # Sheet-lərin mövcudluğunu yoxla
    if "Forma8_1" not in wb.sheetnames:
        raise ValueError(f"{excel_file} faylında 'Forma8_1' sheet-i yoxdur!")
    
    ws_8_1 = wb["Forma8_1"]
    
    # Product adını Forma8_1-dən götür
    product = ws_8_1["C8"].value if ws_8_1["C8"].value else None
    
    if not product:
        print(f"  ⚠ {excel_file}: Forma8_1-də product məlumatı yoxdur (C8 boşdur)")
        wb.save(excel_file)
        return
    
    print(f"    Product: {product}")
    
    # ================== PRODUCT-A GÖRƏ TİP TƏYİN ET ==================
    use_type1 = product in FORMA8_7_TYPE1_PRODUCTS
    use_type2 = product in FORMA8_7_TYPE2_PRODUCTS
    
    if not use_type1 and not use_type2:
        print(f"  ⚠ {product}: Forma8_7 tip məlumatı tapılmadı, skip edilir")
        wb.save(excel_file)
        return
    
    if use_type1:
        print(f"    → Type1: Forma8_7(1)")
        sheet_to_keep = "Forma8_7(1)"
        sheet_to_delete = "Forma8_7(2)"
        prev_start_row = 14
        prev_end_row = 24
        new_start_row = 13
        sum_row = 24  # Type1 üçün D24
        copy_from_row = 23  # F23
        copy_to_row = 24    # E24
        total_f_target_row = 24  # Type1 üçün F24
        g_calculation_row = 24  # Type1 üçün G24
        calculation_row = 24  # Type1 üçün H24, I24, J24, K24
    else:  # use_type2
        print(f"    → Type2: Forma8_7(2)")
        sheet_to_keep = "Forma8_7(2)"
        sheet_to_delete = "Forma8_7(1)"
        prev_start_row = 14
        prev_end_row = 32
        new_start_row = 13
        sum_row = 32  # Type2 üçün D32
        copy_from_row = 31  # F31
        copy_to_row = 32    # E32
        total_f_target_row = 32  # Type2 üçün F32
        g_calculation_row = 32  # Type2 üçün G32
        calculation_row = 32  # Type2 üçün H32, I32, J32, K32
    
    # ================== SHEET-LƏRİ İDARƏ ET ==================
    if sheet_to_delete in wb.sheetnames:
        del wb[sheet_to_delete]
        print(f"    ✓ {sheet_to_delete} silindi")
    
    if sheet_to_keep not in wb.sheetnames:
        print(f"  ⚠ {sheet_to_keep} sheet-i tapılmadı!")
        wb.save(excel_file)
        return
    
    ws = wb[sheet_to_keep]
    ws.title = "Forma8_7"
    print(f"    ✓ {sheet_to_keep} → Forma8_7 adına dəyişdirildi")
    
    # ================== STYLES ==================
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    font_style = Font(name="A3 Times AZ Lat", size=10)
    bold_font = Font(name="A3 Times AZ Lat", size=10, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    
    # ================== C8-Ə PRODUCT YAZMA ==================
    ws["C8"].value = product
    ws["C8"].font = font_style
    ws["C8"].alignment = center

    # ================== D6-YA TARİX YAZMA ==================
    ref_date = pd.to_datetime(reference_date)
    formatted_date = ref_date.strftime("%d.%m.%Y")
    ws["D6"].value = formatted_date
    print(f"    ✓ D6-ya tarix yazıldı: {formatted_date}")
    
    # ================== ƏVVƏLKİ EXCEL-DƏN DATA KOPYALAMA ==================
    previous_file = os.path.join(previous_folder, f"{product}.xlsx")
    
    if not os.path.exists(previous_file):
        print(f"    ⚠ Əvvəlki Excel tapılmadı: {previous_file}")
    else:
        try:
            print(f"    Əvvəlki Excel-dən data kopyalanır: {os.path.basename(previous_file)}")
            wb_prev = load_workbook(previous_file)
            
            if "Forma8_7" not in wb_prev.sheetnames:
                print(f"    ⚠ Əvvəlki Excel-də Forma8_7 sheet-i yoxdur")
            else:
                ws_prev = wb_prev["Forma8_7"]
                
                rows_copied = 0
                for i, prev_row in enumerate(range(prev_start_row, prev_end_row + 1)):
                    new_row = new_start_row + i
                    
                    for col in range(4, 12):  # D-dən K-ya
                        prev_cell = ws_prev.cell(row=prev_row, column=col)
                        new_cell = ws.cell(row=new_row, column=col)
                        
                        new_cell.value = prev_cell.value
                        
                        if prev_cell.font:
                            new_cell.font = prev_cell.font.copy()
                        if prev_cell.border:
                            new_cell.border = prev_cell.border.copy()
                        if prev_cell.fill:
                            new_cell.fill = prev_cell.fill.copy()
                        if prev_cell.alignment:
                            new_cell.alignment = prev_cell.alignment.copy()
                        if prev_cell.number_format:
                            new_cell.number_format = prev_cell.number_format
                    
                    rows_copied += 1
                
                print(f"    ✓ {rows_copied} sətir kopyalandı (D{prev_start_row}:K{prev_end_row} → D{new_start_row}:K{new_start_row + rows_copied - 1})")
        except Exception as e:
            print(f"    ⚠ Kopyalama xətası: {e}")
    
    # ================== FORMA8_1-DƏN SON 3 AYLIK TOPLAM ==================
    print(f"    Forma8_1-dən son 3 aylıq toplam hesablanır...")
    
    # Reference date-dən 3 ay əvvəl
    ref_date = pd.to_datetime(reference_date)
    period_end = ref_date
    period_start = ref_date - relativedelta(months=3)
    
    print(f"      Period: {period_start.date()} - {period_end.date()}")
    
    # Forma8_1-dən data oxu
    forma8_1_data = []
    
    # Forma8_1-də 12-ci sətirdən başlayıb data oxuyuruq
    row = 12
    while True:
        val_a = ws_8_1.cell(row=row, column=1).value
        val_c = ws_8_1.cell(row=row, column=3).value
        val_f = ws_8_1.cell(row=row, column=6).value
        
        # Boş və ya toplam sətirlərində dayan
        if not val_a or (isinstance(val_a, str) and ("AA" in val_a or "Yekun" in val_a)):
            break
        
        # Tarix və F dəyəri
        if val_c and val_f:
            forma8_1_data.append({
                'date': val_c,
                'value': val_f
            })
        
        row += 1
    
    # DataFrame yarat
    if forma8_1_data:
        df = pd.DataFrame(forma8_1_data)
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
        df['value'] = pd.to_numeric(df['value'], errors='coerce').fillna(0)
        
        # Period filtri
        mask = (df['date'] >= period_start) & (df['date'] < period_end)
        period_sum = df.loc[mask, 'value'].sum()
    else:
        period_sum = 0
    
    # D sütununa yaz
    ws.cell(row=sum_row, column=4).value = round(period_sum, 2)
    ws.cell(row=sum_row, column=4).font = font_style
    ws.cell(row=sum_row, column=4).border = thin
    ws.cell(row=sum_row, column=4).alignment = center
    
    print(f"      ✓ D{sum_row} = {period_sum:.2f}")
    
    # ================== F→E KOPYALAMA ==================
    print(f"    F{copy_from_row} → E{copy_to_row} kopyalanır...")
    
    source_cell = ws.cell(row=copy_from_row, column=6)  # F sütunu (column=6)
    target_cell = ws.cell(row=copy_to_row, column=5)    # E sütunu (column=5)
    
    # Dəyəri kopyala
    target_cell.value = source_cell.value
    
    # Formatı kopyala
    if source_cell.font:
        target_cell.font = source_cell.font.copy()
    if source_cell.border:
        target_cell.border = source_cell.border.copy()
    if source_cell.fill:
        target_cell.fill = source_cell.fill.copy()
    if source_cell.alignment:
        target_cell.alignment = source_cell.alignment.copy()
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    
    print(f"      ✓ E{copy_to_row} = {target_cell.value}")
    
    # ================== FORMA8_2-DƏN GƏLƏN TOTAL_F YAZMA ==================
    if total_f_from_forma8_2 is not None:
        print(f"    Forma8_2-dən total_f yazılır → F{total_f_target_row}")
        
        target_f_cell = ws.cell(row=total_f_target_row, column=6)  # F sütunu
        target_f_cell.value = round(total_f_from_forma8_2, 2)
        target_f_cell.font = bold_font
        target_f_cell.border = thin
        target_f_cell.alignment = center
        
        print(f"      ✓ F{total_f_target_row} = {total_f_from_forma8_2:.2f}")
    
    # ================== G SÜTUNU HESABLAMA: D + E - F ==================
    print(f"    G{g_calculation_row} hesablanır: D + E - F")
    
    # D, E, F dəyərlərini oxu
    d_value = ws.cell(row=g_calculation_row, column=4).value or 0
    e_value = ws.cell(row=g_calculation_row, column=5).value or 0
    f_value = ws.cell(row=g_calculation_row, column=6).value or 0
    
    # Numeric dəyərə çevir
    try:
        d_value = float(d_value)
    except (ValueError, TypeError):
        d_value = 0
    
    try:
        e_value = float(e_value)
    except (ValueError, TypeError):
        e_value = 0
    
    try:
        f_value = float(f_value)
    except (ValueError, TypeError):
        f_value = 0
    
    # G hesabla: D + E - F
    g_value = d_value + e_value - f_value
    
    # G sütununa yaz
    target_g_cell = ws.cell(row=g_calculation_row, column=7)  # G sütunu
    target_g_cell.value = round(g_value, 2)
    target_g_cell.font = bold_font
    target_g_cell.border = thin
    target_g_cell.alignment = center
    
    print(f"      ✓ G{g_calculation_row} = {g_value:.2f}")
    
    # ================== H, I, J, K SÜTUNLARI HESABLAMA ==================
    print(f"    H, I, J, K sütunları hesablanır (row {calculation_row})")
    
    # 0.01 əmsalının tətbiq edilib-edilməyəcəyini yoxla
    apply_coefficient = product in COEFFICIENT_0_01_PRODUCTS
    
    if apply_coefficient:
        print(f"      → {product} üçün 0.01 əmsalı tətbiq edilir")
        coefficient = 0.01
    else:
        print(f"      → {product} üçün 0.01 əmsalı tətbiq edilmir (0.00 yazılacaq)")
        coefficient = 0
    
    # H = D * coefficient
    h_value = d_value * coefficient
    h_cell = ws.cell(row=calculation_row, column=8)  # H sütunu
    h_cell.value = round(h_value, 2)
    h_cell.font = bold_font
    h_cell.border = thin
    h_cell.alignment = center
    print(f"      ✓ H{calculation_row} = {h_value:.2f}")
    
    # I = E * coefficient
    i_value = e_value * coefficient
    i_cell = ws.cell(row=calculation_row, column=9)  # I sütunu
    i_cell.value = round(i_value, 2)
    i_cell.font = bold_font
    i_cell.border = thin
    i_cell.alignment = center
    print(f"      ✓ I{calculation_row} = {i_value:.2f}")
    
    # J = F * coefficient
    j_value = f_value * coefficient
    j_cell = ws.cell(row=calculation_row, column=10)  # J sütunu
    j_cell.value = round(j_value, 2)
    j_cell.font = bold_font
    j_cell.border = thin
    j_cell.alignment = center
    print(f"      ✓ J{calculation_row} = {j_value:.2f}")
    
    # K = G * coefficient
    k_value = g_value * coefficient
    k_cell = ws.cell(row=calculation_row, column=11)  # K sütunu
    k_cell.value = round(k_value, 2)
    k_cell.font = bold_font
    k_cell.border = thin
    k_cell.alignment = center
    print(f"      ✓ K{calculation_row} = {k_value:.2f}")
    
    wb.save(excel_file)
    print(f"  ✅ {product}: Forma8_7 tamamlandı")