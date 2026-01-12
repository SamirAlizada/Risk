from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import pandas as pd
from dateutil.relativedelta import relativedelta
import os

# Forma8_14 hesablama tələb edən məhsullar
FORMA8_14_CALCULATION_PRODUCTS = [
    "(04)AvtoKasko",
    "(08)Yuk",
    "(03)EmlakYanginDigerRisk",
    "(37)IcbariDashinmazEmlak"
]

def run_forma8_14(excel_file: str, reference_date: str, ucot_file: str, previous_folder: str):
    """Forma8_14 hazırlayır: Xüsusi 4 məhsul üçün hesablama, qalanları default saxlayır"""
    
    wb = load_workbook(excel_file)
    
    # Sheet-lərin mövcudluğunu yoxla
    if "Forma8_1" not in wb.sheetnames:
        raise ValueError(f"{excel_file} faylında 'Forma8_1' sheet-i yoxdur!")
    
    ws_8_1 = wb["Forma8_1"]
    
    # Product adını Forma8_1-dən götür
    product = ws_8_1["C8"].value if ws_8_1["C8"].value else None
    
    if not product:
        print(f"  ⚠ {excel_file}: Forma8_1-də product məlumatı yoxdur (C8 boşdur)")
        wb.save(excel_file)
        return
    
    print(f"    Product: {product}")
    
    # ================== FORMA8_14 SHEET YOXLA ==================
    if "Forma8_14" not in wb.sheetnames:
        print(f"  ⚠ Forma8_14 sheet-i tapılmadı!")
        wb.save(excel_file)
        return
    
    ws = wb["Forma8_14"]
    print(f"    ✓ Forma8_14 sheet-i tapıldı")
    
    # ================== STYLES ==================
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    font_style = Font(name="A3 Times AZ Lat", size=10)
    bold_font = Font(name="A3 Times AZ Lat", size=10, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    
    # ================== C8-Ə PRODUCT YAZMA ==================
    ws["C8"].value = product
    ws["C8"].font = font_style
    ws["C8"].alignment = center
    
    print(f"    ✓ C8-ə product yazıldı: {product}")

    # ================== D6-YA TARİX YAZMA ==================
    ref_date = pd.to_datetime(reference_date)
    formatted_date = ref_date.strftime("%d.%m.%Y")
    ws["D6"].value = formatted_date
    print(f"    ✓ D6-ya tarix yazıldı: {formatted_date}")
    
    # ================== XÜSUSI 4 MƏHSUL ÜÇÜN HESABLAMA ==================
    if product in FORMA8_14_CALCULATION_PRODUCTS:
        print(f"    → Xüsusi məhsul: hesablama aparılır")
        
        # ================== ƏVVƏLKİ EXCEL-DƏN E22 → E12 KOPYALAMA ==================
        previous_file = os.path.join(previous_folder, f"{product}.xlsx")

        if not os.path.exists(previous_file):
            print(f"    ⚠ Əvvəlki Excel tapılmadı: {previous_file}")
            print(f"    ⚠ E12 boş qalacaq")
        else:
            try:
                print(f"    Əvvəlki Excel-dən E22 kopyalanır: {os.path.basename(previous_file)}")
                wb_prev = load_workbook(previous_file, data_only=True)
                
                if "Forma8_14" not in wb_prev.sheetnames:
                    print(f"    ⚠ Əvvəlki Excel-də Forma8_14 sheet-i yoxdur")
                    print(f"    ⚠ E12 boş qalacaq")
                else:
                    ws_prev = wb_prev["Forma8_14"]
                    
                    # E22-dən sadəcə dəyəri oxu
                    prev_e22_value = ws_prev.cell(row=22, column=5).value
                    
                    # E12-yə SADƏCƏ DƏYƏRİ yaz (format toxunmur)
                    ws.cell(row=12, column=5).value = prev_e22_value
                    
                    print(f"      ✓ E22 → E12 kopyalandı (dəyər: {prev_e22_value})")
                    wb_prev.close()
            
            except Exception as e:
                print(f"    ⚠ Kopyalama xətası: {e}")
        
        # ================== FORMA8_7-DƏN K24 → E13 KOPYALAMA ==================
        print(f"    Forma8_7-dən K24 kopyalanır...")

        if "Forma8_7" not in wb.sheetnames:
            print(f"    ⚠ Forma8_7 sheet-i tapılmadı, E13 boş qalacaq")
        else:
            ws_8_7 = wb["Forma8_7"]
            
            # K24-dən oxu
            source_k24_value = ws_8_7.cell(row=24, column=11).value  # K24
            
            # E13-ə SADƏCƏ DƏYƏRİ yaz (format toxunmur)
            ws.cell(row=13, column=5).value = source_k24_value
            
            print(f"      ✓ K24 → E13 kopyalandı (dəyər: {source_k24_value})")

        # ================== FORMA8_10-DƏN K24 → E14 KOPYALAMA ==================
        print(f"    Forma8_10-dən K24 kopyalanır...")

        if "Forma8_10" not in wb.sheetnames:
            print(f"    ⚠ Forma8_10 sheet-i tapılmadı, E14 boş qalacaq")
        else:
            ws_8_10 = wb["Forma8_10"]
            
            # K24-dən oxu
            source_k24_value = ws_8_10.cell(row=24, column=11).value  # K24
            
            # E14-ə SADƏCƏ DƏYƏRİ yaz (format toxunmur)
            ws.cell(row=14, column=5).value = source_k24_value
            
            print(f"      ✓ K24 → E14 kopyalandı (dəyər: {source_k24_value})")

        # ================== E15 HESABLAMA: (E13-E14)*0.12+E12 ==================
        print(f"    E15 hesablanır: (E13-E14)*0.12+E12...")

        # E12, E13, E14 dəyərlərini oxu
        e12_value = ws.cell(row=12, column=5).value or 0
        e13_value = ws.cell(row=13, column=5).value or 0
        e14_value = ws.cell(row=14, column=5).value or 0

        # Rəqəmə çevir
        try:
            e12_value = float(e12_value)
        except (ValueError, TypeError):
            e12_value = 0

        try:
            e13_value = float(e13_value)
        except (ValueError, TypeError):
            e13_value = 0

        try:
            e14_value = float(e14_value)
        except (ValueError, TypeError):
            e14_value = 0

        # Hesablama: (E13 - E14) * 0.12 + E12
        result = (e13_value - e14_value) * 0.12 + e12_value
        result = round(result, 2)

        # E15-ə SADƏCƏ DƏYƏRİ yaz (format toxunmur)
        ws.cell(row=15, column=5).value = result

        print(f"      ✓ E12={e12_value:.2f}, E13={e13_value:.2f}, E14={e14_value:.2f}")
        print(f"      ✓ E15 = ({e13_value:.2f} - {e14_value:.2f}) * 0.12 + {e12_value:.2f} = {result:.2f}")

        # ================== E17:E21 HESABLAMA (YANVAR YOXLAMASI) ==================
        print(f"    E17:E21 prosesi başlayır...")

        # Reference date-i parse et
        ref_date = pd.to_datetime(reference_date)

        # Yanvar ayının 1-i yoxla
        is_january_first = (ref_date.month == 1 and ref_date.day == 1)

        if is_january_first:
            print(f"      → Yanvar 1 aşkarlandı, xüsusi hesablama aparılır")
            
            # ================== PREVIOUS-DAN E17:E20 → E18:E21 KOPYALAMA ==================
            previous_file = os.path.join(previous_folder, f"{product}.xlsx")
            
            if not os.path.exists(previous_file):
                print(f"      ⚠ Əvvəlki Excel tapılmadı: {previous_file}")
                print(f"      ⚠ E18:E21 boş qalacaq")
            else:
                try:
                    print(f"      Əvvəlki Excel-dən E17:E20 → E18:E21 kopyalanır...")
                    wb_prev = load_workbook(previous_file, data_only=True)
                    
                    if "Forma8_14" not in wb_prev.sheetnames:
                        print(f"      ⚠ Əvvəlki Excel-də Forma8_14 sheet-i yoxdur")
                        print(f"      ⚠ E18:E21 boş qalacaq")
                    else:
                        ws_prev = wb_prev["Forma8_14"]
                        
                        # E17:E20-dən oxuyub E18:E21-ə yaz
                        for i in range(4):  # 4 sətir: E17→E18, E18→E19, E19→E20, E20→E21
                            prev_row = 17 + i  # 17, 18, 19, 20
                            new_row = 18 + i   # 18, 19, 20, 21
                            
                            prev_value = ws_prev.cell(row=prev_row, column=5).value
                            ws.cell(row=new_row, column=5).value = prev_value
                        
                        print(f"      ✓ E17:E20 → E18:E21 kopyalandı")
                        wb_prev.close()
                
                except Exception as e:
                    print(f"      ⚠ Kopyalama xətası: {e}")
            
            # ================== E17 HESABLAMA: sum(8_7 H21:H24) - sum(8_10 H21:H24) ==================
            print(f"      E17 hesablanır: sum(8_7 H21:H24) - sum(8_10 H21:H24)...")
            
            sum_8_7 = 0
            sum_8_10 = 0
            
            # Forma8_7-dən H21:H24 topla
            if "Forma8_7" in wb.sheetnames:
                ws_8_7 = wb["Forma8_7"]
                for row in range(21, 25):  # 21, 22, 23, 24
                    cell_value = ws_8_7.cell(row=row, column=8).value or 0  # H sütunu (column=8)
                    try:
                        sum_8_7 += float(cell_value)
                    except (ValueError, TypeError):
                        pass
                print(f"      Forma8_7 sum(H21:H24) = {sum_8_7:.2f}")
            else:
                print(f"      ⚠ Forma8_7 sheet-i tapılmadı")
            
            # Forma8_10-dan H21:H24 topla
            if "Forma8_10" in wb.sheetnames:
                ws_8_10 = wb["Forma8_10"]
                for row in range(21, 25):  # 21, 22, 23, 24
                    cell_value = ws_8_10.cell(row=row, column=8).value or 0  # H sütunu (column=8)
                    try:
                        sum_8_10 += float(cell_value)
                    except (ValueError, TypeError):
                        pass
                print(f"      Forma8_10 sum(H21:H24) = {sum_8_10:.2f}")
            else:
                print(f"      ⚠ Forma8_10 sheet-i tapılmadı")
            
            # E17 hesabla və yaz
            e17_result = sum_8_7 - sum_8_10
            e17_result = round(e17_result, 2)
            
            ws.cell(row=17, column=5).value = e17_result
            print(f"      ✓ E17 = {sum_8_7:.2f} - {sum_8_10:.2f} = {e17_result:.2f}")

        else:
            print(f"      → Yanvar 1 deyil, Previous-dan E17:E21 kopyalanır...")
            
            # ================== PREVIOUS-DAN E17:E21 → E17:E21 KOPYALAMA ==================
            previous_file = os.path.join(previous_folder, f"{product}.xlsx")
            
            if not os.path.exists(previous_file):
                print(f"      ⚠ Əvvəlki Excel tapılmadı: {previous_file}")
                print(f"      ⚠ E17:E21 boş qalacaq")
            else:
                try:
                    wb_prev = load_workbook(previous_file, data_only=True)
                    
                    if "Forma8_14" not in wb_prev.sheetnames:
                        print(f"      ⚠ Əvvəlki Excel-də Forma8_14 sheet-i yoxdur")
                        print(f"      ⚠ E17:E21 boş qalacaq")
                    else:
                        ws_prev = wb_prev["Forma8_14"]
                        
                        # E17:E21-dən oxuyub eyni yerlərə yaz
                        for row in range(17, 22):  # 17, 18, 19, 20, 21
                            prev_value = ws_prev.cell(row=row, column=5).value
                            ws.cell(row=row, column=5).value = prev_value
                        
                        print(f"      ✓ E17:E21 → E17:E21 kopyalandı")
                        wb_prev.close()
                
                except Exception as e:
                    print(f"      ⚠ Kopyalama xətası: {e}")
        # ================== E22 HESABLAMA: MIN(MAX(E15-E16;0);MAX(E17:E21)*150%) ==================
        print(f"    E22 hesablanır: MIN(MAX(E15-E16;0);MAX(E17:E21)*150%)...")

        # E15, E16 dəyərlərini oxu
        e15_value = ws.cell(row=15, column=5).value or 0
        e16_value = ws.cell(row=16, column=5).value or 0

        # Rəqəmə çevir
        try:
            e15_value = float(e15_value)
        except (ValueError, TypeError):
            e15_value = 0

        try:
            e16_value = float(e16_value)
        except (ValueError, TypeError):
            e16_value = 0

        # MAX(E15-E16; 0) hesabla
        max_diff = max(e15_value - e16_value, 0)

        # E17:E21 aralığından maksimum dəyəri tap
        e17_e21_values = []
        for row in range(17, 22):  # 17, 18, 19, 20, 21
            cell_value = ws.cell(row=row, column=5).value or 0
            try:
                e17_e21_values.append(float(cell_value))
            except (ValueError, TypeError):
                e17_e21_values.append(0)

        max_e17_e21 = max(e17_e21_values) if e17_e21_values else 0

        # MAX(E17:E21) * 150%
        max_e17_e21_150 = max_e17_e21 * 1.5

        # MIN(MAX(E15-E16;0); MAX(E17:E21)*150%)
        e22_result = min(max_diff, max_e17_e21_150)
        e22_result = round(e22_result, 2)

        # E22-yə yaz
        ws.cell(row=22, column=5).value = e22_result

        print(f"      ✓ E15={e15_value:.2f}, E16={e16_value:.2f}")
        print(f"      ✓ MAX(E15-E16; 0) = MAX({e15_value:.2f}-{e16_value:.2f}; 0) = {max_diff:.2f}")
        print(f"      ✓ MAX(E17:E21) = {max_e17_e21:.2f}")
        print(f"      ✓ MAX(E17:E21)*150% = {max_e17_e21:.2f}*1.5 = {max_e17_e21_150:.2f}")
        print(f"      ✓ E22 = MIN({max_diff:.2f}; {max_e17_e21_150:.2f}) = {e22_result:.2f}")
            
    else:
        print(f"    → Adi məhsul: hesablama aparılmır (default qalır)")
    
    wb.save(excel_file)
    print(f"  ✅ {product}: Forma8_14 tamamlandı")