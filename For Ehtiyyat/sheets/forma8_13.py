from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import pandas as pd
from dateutil.relativedelta import relativedelta

# Forma8_13 Type1 məhsullar
FORMA8_13_TYPE1_PRODUCTS = [
    "(01)FerdiQeza",
    "(02)Tibbi",
    "(03)EmlakYanginDigerRisk",
    "(04)AvtoKasko",
    "(05)DemiryolNeqliyyVasitesi",
    "(06)HavaNeqliyyKasko",
    "(07)SuNeqliyyKasko",
    "(08)Yuk",
    "(09)KendTeserrufBitki",
    "(10)KendTeserrufHeyvan",
    "(11)IshcilerinDeleduzlug",
    "(12)PulvePulSenedSaxtalash",
    "(22)Kredit",
    "(23)Ipoteka",
    "(24)EmlakinDeyerdenDushmesi",
    "(25)IshinDayanmasiRiski",
    "(27)SernishinIcbari",
    "(28)IcbariEkoloji",
    "(29)YanginIcbari",
    "(30)DeputatlarinIcbari",
    "(31)TibbiPersonalinAIDSden",
    "(32)HerbiQulluqcularinIcbari",
    "(33)HuquqMuhafifeIcbari",
    "(34)DovletQulluqcuIcbari",
    "(35)DiplomatlarinIcbari",
    "(37)IcbariDashinmazEmlak",
    "(39)IcbariNVSMMS",
    "(40)IcbariSernishinFerdiQeza",
    "(41)Sefer",
    "(42)Titul",
    "(43)HuquqiXerc"
]

# Forma8_13 Type2 məhsullar
FORMA8_13_TYPE2_PRODUCTS = [
    "(19)PesheMesuliyy",
    "(20)IshegoturenMesuliyy",
    "(21)UmumiMulkiMesuliyy",
    "(13)AvtoKonulluMesuliyy",
    "(14)DemiryolNeqliySahibMesuliyy",
    "(15)HavaNeqliySahibMesuliyy",
    "(16)SuNeqliySahibMesuliyy",
    "(17)YukDashiyanMesuliyy",
    "(18)MulkiMuqavileUzreMesuliyy",
    "(26)AvtoIcbariMesuliyy",
    "(36)AuditorPesheMesuliyyIcbari",
    "(38)IcbariDashinmazEmlakMesul"
]

def run_forma8_13(excel_file: str, reference_date: str, ucot_file: str):
    """Forma8_13 hazırlayır: Bütün product-lar üçün eyni sheet, sadəcə product adı dəyişir"""
    
    wb = load_workbook(excel_file)
    
    # Sheet-lərin mövcudluğunu yoxla
    if "Forma8_1" not in wb.sheetnames:
        raise ValueError(f"{excel_file} faylında 'Forma8_1' sheet-i yoxdur!")
    
    ws_8_1 = wb["Forma8_1"]
    
    # Product adını Forma8_1-dən götür
    product = ws_8_1["C8"].value if ws_8_1["C8"].value else None
    
    if not product:
        print(f"  ⚠ {excel_file}: Forma8_1-də product məlumatı yoxdur (C8 boşdur)")
        wb.save(excel_file)
        return
    
    print(f"    Product: {product}")
    
    # ================== PRODUCT-A GÖRƏ TİP TƏYİN ET ==================
    use_type1 = product in FORMA8_13_TYPE1_PRODUCTS
    use_type2 = product in FORMA8_13_TYPE2_PRODUCTS
    
    if not use_type1 and not use_type2:
        print(f"  ⚠ {product}: Forma8_13 tip məlumatı tapılmadı, skip edilir")
        wb.save(excel_file)
        return
    
    if use_type1:
        print(f"    → Type1")
        source_row = 31
        source_col = 23  # W sütunu
    else:  # use_type2
        print(f"    → Type2")
        source_row = 39
        source_col = 31  # AE sütunu
    
    # ================== FORMA8_13 SHEET YOXLA ==================
    if "Forma8_13" not in wb.sheetnames:
        print(f"  ⚠ Forma8_13 sheet-i tapılmadı!")
        wb.save(excel_file)
        return
    
    ws = wb["Forma8_13"]
    print(f"    ✓ Forma8_13 sheet-i tapıldı")
    
    # ================== STYLES ==================
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    font_style = Font(name="A3 Times AZ Lat", size=10)
    bold_font = Font(name="A3 Times AZ Lat", size=10, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    
    # ================== C8-Ə PRODUCT YAZMA ==================
    ws["C8"].value = product
    ws["C8"].font = font_style
    ws["C8"].alignment = center
    
    print(f"    ✓ C8-ə product yazıldı: {product}")

    # ================== D6-YA TARİX YAZMA ==================
    ref_date = pd.to_datetime(reference_date)
    formatted_date = ref_date.strftime("%d.%m.%Y")
    ws["D6"].value = formatted_date
    print(f"    ✓ D6-ya tarix yazıldı: {formatted_date}")
    
    # ================== FORMA8_12-DƏN DATA KOPYALAMA ==================
    print(f"    Forma8_12-dən data kopyalanır...")

    if "Forma8_12" not in wb.sheetnames:
        print(f"    ⚠ Forma8_12 sheet-i tapılmadı, E13 boş qalacaq")
    else:
        ws_8_12 = wb["Forma8_12"]
        
        # Source cell-dən oxu
        source_cell = ws_8_12.cell(row=source_row, column=source_col)
        
        # Target cell-ə yaz (E13)
        target_cell = ws.cell(row=13, column=5)  # E13
        
        # Dəyəri kopyala
        target_cell.value = source_cell.value
        
        # Formatı kopyala
        if source_cell.font:
            target_cell.font = source_cell.font.copy()
        else:
            target_cell.font = font_style
        
        if source_cell.border:
            target_cell.border = source_cell.border.copy()
        else:
            target_cell.border = thin
        
        if source_cell.fill:
            target_cell.fill = source_cell.fill.copy()
        
        if source_cell.alignment:
            target_cell.alignment = source_cell.alignment.copy()
        else:
            target_cell.alignment = center
        
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        else:
            target_cell.number_format = '#,##0.00'
        
        if use_type1:
            print(f"      ✓ Type1: W31 → E13 kopyalandı (dəyər: {target_cell.value})")
        else:
            print(f"      ✓ Type2: AE39 → E13 kopyalandı (dəyər: {target_cell.value})")

    # ================== FORMA8_11-DƏN DATA KOPYALAMA (25%) ==================
    print(f"    Forma8_11-dən data kopyalanır (25%)...")

    if "Forma8_11" not in wb.sheetnames:
        print(f"    ⚠ Forma8_11 sheet-i tapılmadı, E14 boş qalacaq")
    else:
        ws_8_11 = wb["Forma8_11"]
        
        if use_type1:
            # Type1: G25-dən oxu
            source_8_11_row = 25
            print(f"      → Type1: G25 × 25% → E14")
        else:
            # Type2: G33-dən oxu
            source_8_11_row = 33
            print(f"      → Type2: G33 × 25% → E14")
        
        # Source cell-dən oxu (G sütunu = column 7)
        source_8_11_cell = ws_8_11.cell(row=source_8_11_row, column=7)
        source_8_11_value = source_8_11_cell.value or 0
        
        # 25% hesabla
        try:
            source_8_11_value = float(source_8_11_value)
            result_value = source_8_11_value * 0.25
            result_value = round(result_value, 2)
        except (ValueError, TypeError):
            result_value = 0.00
        
        # Target cell-ə yaz (E14)
        target_8_11_cell = ws.cell(row=14, column=5)  # E14
        target_8_11_cell.value = result_value
        target_8_11_cell.font = font_style
        target_8_11_cell.border = thin
        target_8_11_cell.alignment = center
        target_8_11_cell.number_format = '#,##0.00'
        
        if use_type1:
            print(f"      ✓ Type1: G25 ({source_8_11_value:.2f}) × 25% = {result_value:.2f} → E14")
        else:
            print(f"      ✓ Type2: G33 ({source_8_11_value:.2f}) × 25% = {result_value:.2f} → E14")

    # ================== FORMA8_10-DƏN DATA KOPYALAMA (2.5%) ==================
    print(f"    Forma8_10-dən data kopyalanır (2.5%)...")

    if "Forma8_10" not in wb.sheetnames:
        print(f"    ⚠ Forma8_10 sheet-i tapılmadı, E15 boş qalacaq")
    else:
        ws_8_10 = wb["Forma8_10"]
        
        if use_type1:
            # Type1: G21:G24 toplamı
            sum_start_row = 21
            sum_end_row = 24
            print(f"      → Type1: sum(G21:G24) × 2.5% → E15")
        else:
            # Type2: G29:G32 toplamı
            sum_start_row = 29
            sum_end_row = 32
            print(f"      → Type2: sum(G29:G32) × 2.5% → E15")
        
        # G sütunundakı dəyərləri topla
        total_sum = 0
        for row in range(sum_start_row, sum_end_row + 1):
            cell_value = ws_8_10.cell(row=row, column=7).value or 0  # G sütunu (column=7)
            try:
                total_sum += float(cell_value)
            except (ValueError, TypeError):
                pass
        
        # 2.5% hesabla
        result_value = total_sum * 0.025
        result_value = round(result_value, 2)
        
        # Target cell-ə yaz (E15)
        target_8_10_cell = ws.cell(row=15, column=5)  # E15
        target_8_10_cell.value = result_value
        target_8_10_cell.font = font_style
        target_8_10_cell.border = thin
        target_8_10_cell.alignment = center
        target_8_10_cell.number_format = '#,##0.00'
        
        if use_type1:
            print(f"      ✓ Type1: sum(G21:G24) = {total_sum:.2f} × 2.5% = {result_value:.2f} → E15")
        else:
            print(f"      ✓ Type2: sum(G29:G32) = {total_sum:.2f} × 2.5% = {result_value:.2f} → E15")

    # ================== E16 HESABLAMA (MAX) ==================
    print(f"    E16 hesablanır (MAX(E13:E15))...")

    # E13, E14, E15 dəyərlərini oxu
    e13_value = ws.cell(row=13, column=5).value or 0
    e14_value = ws.cell(row=14, column=5).value or 0
    e15_value = ws.cell(row=15, column=5).value or 0

    # Rəqəmə çevir
    try:
        e13_value = float(e13_value)
    except (ValueError, TypeError):
        e13_value = 0

    try:
        e14_value = float(e14_value)
    except (ValueError, TypeError):
        e14_value = 0

    try:
        e15_value = float(e15_value)
    except (ValueError, TypeError):
        e15_value = 0

    # Maksimum dəyəri tap
    max_value = max(e13_value, e14_value, e15_value)
    max_value = round(max_value, 2)

    # E16-ya yaz
    e16_cell = ws.cell(row=16, column=5)  # E16
    e16_cell.value = max_value
    e16_cell.font = bold_font
    e16_cell.border = thin
    e16_cell.alignment = center
    e16_cell.number_format = '#,##0.00'

    print(f"      ✓ E13={e13_value:.2f}, E14={e14_value:.2f}, E15={e15_value:.2f}")
    print(f"      ✓ MAX(E13:E15) = {max_value:.2f} → E16")

    wb.save(excel_file)
    print(f"  ✅ {product}: Forma8_13 tamamlandı")