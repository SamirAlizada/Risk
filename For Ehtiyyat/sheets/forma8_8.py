from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import pandas as pd
from dateutil.relativedelta import relativedelta

# Forma8_8(1) istifadə edən məhsullar
FORMA8_8_TYPE1_PRODUCTS = [
    "(01)FerdiQeza",
    "(02)Tibbi",
    "(03)EmlakYanginDigerRisk",
    "(04)AvtoKasko",
    "(05)DemiryolNeqliyyVasitesi",
    "(06)HavaNeqliyyKasko",
    "(07)SuNeqliyyKasko",
    "(08)Yuk",
    "(09)KendTeserrufBitki",
    "(10)KendTeserrufHeyvan",
    "(11)IshcilerinDeleduzlug",
    "(12)PulvePulSenedSaxtalash",
    "(22)Kredit",
    "(23)Ipoteka",
    "(24)EmlakinDeyerdenDushmesi",
    "(25)IshinDayanmasiRiski",
    "(27)SernishinIcbari",
    "(28)IcbariEkoloji",
    "(29)YanginIcbari",
    "(30)DeputatlarinIcbari",
    "(31)TibbiPersonalinAIDSden",
    "(32)HerbiQulluqcularinIcbari",
    "(33)HuquqMuhafifeIcbari",
    "(34)DovletQulluqcuIcbari",
    "(35)DiplomatlarinIcbari",
    "(37)IcbariDashinmazEmlak",
    "(39)IcbariNVSMMS",
    "(40)IcbariSernishinFerdiQeza",
    "(41)Sefer",
    "(42)Titul",
    "(43)HuquqiXerc"
]

# Forma8_8(2) istifadə edən məhsullar
FORMA8_8_TYPE2_PRODUCTS = [
    "(19)PesheMesuliyy",
    "(20)IshegoturenMesuliyy",
    "(21)UmumiMulkiMesuliyy",
    "(13)AvtoKonulluMesuliyy",
    "(14)DemiryolNeqliySahibMesuliyy",
    "(15)HavaNeqliySahibMesuliyy",
    "(16)SuNeqliySahibMesuliyy",
    "(17)YukDashiyanMesuliyy",
    "(18)MulkiMuqavileUzreMesuliyy",
    "(26)AvtoIcbariMesuliyy",
    "(36)AuditorPesheMesuliyyIcbari",
    "(38)IcbariDashinmazEmlakMesul"
]

def run_forma8_8(excel_file: str, reference_date: str, ucot_file: str):
    """Forma8_8 hazırlayır: Product-a görə düzgün sheet-i seçir, tarixləri və hesablamaları yazır"""
    
    wb = load_workbook(excel_file)
    
    # Sheet-lərin mövcudluğunu yoxla
    if "Forma8_1" not in wb.sheetnames:
        raise ValueError(f"{excel_file} faylında 'Forma8_1' sheet-i yoxdur!")
    
    ws_8_1 = wb["Forma8_1"]
    
    # Product adını Forma8_1-dən götür
    product = ws_8_1["C8"].value if ws_8_1["C8"].value else None
    
    if not product:
        print(f"  ⚠ {excel_file}: Forma8_1-də product məlumatı yoxdur (C8 boşdur)")
        wb.save(excel_file)
        return
    
    print(f"    Product: {product}")
    
    # ================== PRODUCT-A GÖRƏ TİP TƏYİN ET ==================
    use_type1 = product in FORMA8_8_TYPE1_PRODUCTS
    use_type2 = product in FORMA8_8_TYPE2_PRODUCTS
    
    if not use_type1 and not use_type2:
        print(f"  ⚠ {product}: Forma8_8 tip məlumatı tapılmadı, skip edilir")
        wb.save(excel_file)
        return
    
    if use_type1:
        print(f"    → Type1: Forma8_8(1)")
        sheet_to_keep = "Forma8_8(1)"
        sheet_to_delete = "Forma8_8(2)"
        end_row = 24      # D24 - son tarix (reference date)
        start_row = 13    # D13 - ilk tarix
        period_count = 12  # 12 period (3 il × 4 rüb = 12)
        total_columns = 12  # E-dən P-yə qədər (12 sütun)
        sum_row = 25      # E25 - cəm sətiri
        # Diagonal cəm üçün koordinatlar: E24, F23, G22, H21
        diagonal_coords = [(24, 5), (23, 6), (22, 7), (21, 8)]  # (row, col)
    else:  # use_type2
        print(f"    → Type2: Forma8_8(2)")
        sheet_to_keep = "Forma8_8(2)"
        sheet_to_delete = "Forma8_8(1)"
        end_row = 32      # D32 - son tarix (reference date)
        start_row = 13    # D13 - ilk tarix
        period_count = 20  # 20 period (5 il × 4 rüb = 20)
        total_columns = 20  # E-dən AF-ə qədər (20 sütun)
        sum_row = 33      # E33 - cəm sətiri
        # Diagonal cəm üçün koordinatlar: E32, F31, G30, H29
        diagonal_coords = [(32, 5), (31, 6), (30, 7), (29, 8)]  # (row, col)
    
    # ================== SHEET-LƏRİ İDARƏ ET ==================
    if sheet_to_delete in wb.sheetnames:
        del wb[sheet_to_delete]
        print(f"    ✓ {sheet_to_delete} silindi")
    
    if sheet_to_keep not in wb.sheetnames:
        print(f"  ⚠ {sheet_to_keep} sheet-i tapılmadı!")
        wb.save(excel_file)
        return
    
    ws = wb[sheet_to_keep]
    ws.title = "Forma8_8"
    print(f"    ✓ {sheet_to_keep} → Forma8_8 adına dəyişdirildi")
    
    # ================== STYLES ==================
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    font_style = Font(name="A3 Times AZ Lat", size=14)
    bold_font = Font(name="A3 Times AZ Lat", size=14, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    
    # ================== C8-Ə PRODUCT YAZMA ==================
    ws["C8"].value = product
    ws["C8"].font = font_style
    ws["C8"].alignment = center
    
    # ================== TARİXLƏRİ YAZMA (3 AY INTERVALI) ==================
    print(f"    Tarixlər yazılır (D{start_row}:D{end_row})...")
    
    # Reference date-i parse et
    ref_date = pd.to_datetime(reference_date)
    
    # Son sətirdən (end_row) başlayıb yuxarıya doğru (start_row-a qədər)
    current_date = ref_date
    
    for row in range(end_row, start_row - 1, -1):  # end_row-dan start_row-a qədər geriyə
        # D sütununa tarixi yaz
        cell = ws.cell(row=row, column=4)  # D sütunu (column=4)
        cell.value = current_date
        cell.font = font_style
        cell.border = thin
        cell.alignment = center
        cell.number_format = "DD.MM.YYYY"  # Tarix formatı
        
        # 3 ay geriyə get (növbəti iterasiya üçün)
        current_date = current_date - relativedelta(months=3)
    
    print(f"      ✓ {period_count} tarix yazıldı (son: {ref_date.strftime('%d.%m.%Y')}, ilk: {(current_date + relativedelta(months=3)).strftime('%d.%m.%Y')})")
    
    # ================== HESABLAMALAR (TYPE1 VƏ TYPE2) ==================
    print(f"    Hesablamaları yazılır...")
    
    # Mənbə fayldan DataFrame-ləri oxu
    zerer_df = pd.read_excel(ucot_file, sheet_name='Zerer')
    subraq_df = pd.read_excel(ucot_file, sheet_name='Subraqasiya')
    
    # Tarix sütunlarını datetime tipinə çevir
    zerer_df['Sığоrtа hаdisəsinin bаş verdiyi tarixi'] = pd.to_datetime(
        zerer_df['Sığоrtа hаdisəsinin bаş verdiyi tarixi'], errors='coerce'
    )
    zerer_df['Sığorta ödənişi Tаriхi'] = pd.to_datetime(
        zerer_df['Sığorta ödənişi Tаriхi'], errors='coerce'
    )
    
    subraq_df['Sığоrtа hаdisəsinin bаş vеrdiyi tаriх'] = pd.to_datetime(
        subraq_df['Sığоrtа hаdisəsinin bаş vеrdiyi tаriх'], errors='coerce'
    )
    subraq_df['Subroqasiya gəlirinin daxil olduğu tarix'] = pd.to_datetime(
        subraq_df['Subroqasiya gəlirinin daxil olduğu tarix'], errors='coerce'
    )
    
    start_col_index = 5  # E sütunu
    
    # Üçbucaq formatında doldur
    for row in range(start_row, end_row + 1):
        # Cari sətirin D tarixi (hadisə intervali üçün)
        D_cell_row_value = ws.cell(row=row, column=4).value
        D_cell_row = pd.to_datetime(D_cell_row_value, errors='coerce')
        
        if pd.isna(D_cell_row):
            # Bu sətir üçün bütün sütunları X et
            for col_offset in range(total_columns):
                ws.cell(row=row, column=start_col_index + col_offset, value='X')
            continue
        
        col_offset = 0
        
        # Sağa doğru: D13, D14, D15... (ödəniş tarixi üçün)
        for payment_row in range(row, end_row + 1):
            current_col = start_col_index + col_offset
            
            # Ödəniş tarixini oxu
            D_cell_payment_value = ws.cell(row=payment_row, column=4).value
            D_cell_payment = pd.to_datetime(D_cell_payment_value, errors='coerce')
            
            if pd.isna(D_cell_payment):
                ws.cell(row=row, column=current_col, value='X')
                col_offset += 1
                continue
            
            # Zərər hesabla
            zerer_filter = (
                (zerer_df['Sığоrtа hаdisəsinin bаş verdiyi tarixi'] < D_cell_row) &
                (zerer_df['Sığоrtа hаdisəsinin bаş verdiyi tarixi'] >= D_cell_row - relativedelta(months=3)) &
                (zerer_df['Sığorta ödənişi Tаriхi'] < D_cell_payment) &
                (zerer_df['Sığоrtаnın sinifləri'] == product)
            )
            zerer_sum = zerer_df.loc[zerer_filter, 'Sığorta ödənişi Cəmi'].sum()
            
            # Subroqasiya hesabla
            subraq_filter = (
                (subraq_df['Sığоrtа hаdisəsinin bаş vеrdiyi tаriх'] < D_cell_row) &
                (subraq_df['Sığоrtа hаdisəsinin bаş vеrdiyi tаriх'] >= D_cell_row - relativedelta(months=3)) &
                (subraq_df['Subroqasiya gəlirinin daxil olduğu tarix'] < D_cell_payment) &
                (subraq_df['Sığоrtаnın sinifləri'] == product)
            )
            subraq_sum = subraq_df.loc[subraq_filter, 'Ödənilmiş subroqasiya gəlirinin məbləği'].sum()
            
            # Nəticə
            result = zerer_sum - subraq_sum
            
            # Excel-ə yaz
            cell = ws.cell(row=row, column=current_col)
            cell.value = result
            cell.font = font_style
            cell.border = thin
            cell.alignment = center
            cell.number_format = '#,##0.00'  # Rəqəm formatı
            
            col_offset += 1
        
        # Qalan sütunlara X yaz
        while col_offset < total_columns:
            cell = ws.cell(row=row, column=start_col_index + col_offset)
            cell.value = 'X'
            cell.font = font_style
            cell.border = thin
            cell.alignment = center
            col_offset += 1
    
    if use_type1:
        print(f"      ✓ Type1 hesablamaları tamamlandı (E13:P24)")
    else:
        print(f"      ✓ Type2 hesablamaları tamamlandı (E13:AF32)")
    
    # ================== CƏM SƏTİRİNİ YAZMA ==================
    print(f"    Cəm sətiri yazılır (sətir {sum_row})...")
    
    for col_offset in range(total_columns):
        current_col = start_col_index + col_offset
        
        # Sütundakı bütün rəqəmləri topla (X-ləri skip et)
        column_sum = 0
        for row in range(start_row, end_row + 1):
            cell_value = ws.cell(row=row, column=current_col).value
            
            # Əgər rəqəmdirsə, topla
            if isinstance(cell_value, (int, float)):
                column_sum += cell_value
        
        # Cəm sətrinə yaz
        sum_cell = ws.cell(row=sum_row, column=current_col)
        sum_cell.value = column_sum
        sum_cell.font = bold_font  # Bold font
        sum_cell.border = thin
        sum_cell.alignment = center
        sum_cell.number_format = '#,##0.00'
    
    if use_type1:
        print(f"      ✓ Type1 cəm sətiri yazıldı (E25:P25)")
    else:
        print(f"      ✓ Type2 cəm sətiri yazıldı (E33:AF33)")
    
    # ================== DİAGONAL CƏMİ (B10:C11) ==================
    print(f"    Diagonal cəm hesablanır (B10:C11)...")
    
    diagonal_sum = 0
    for row_idx, col_idx in diagonal_coords:
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        
        if isinstance(cell_value, (int, float)):
            diagonal_sum += cell_value
    
    # B10:C11 artıq merge olunub, sadəcə B10-a yaz
    diagonal_cell = ws["B10"]
    diagonal_cell.value = diagonal_sum
    diagonal_cell.font = bold_font
    diagonal_cell.alignment = center
    diagonal_cell.number_format = '#,##0.00'

    # ================== DİAGONAL CƏMİ (B10:C11) ==================
    print(f"    Diagonal cəm hesablanır (B10:C11)...")
    
    diagonal_sum = 0
    for row_idx, col_idx in diagonal_coords:
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        
        if isinstance(cell_value, (int, float)):
            diagonal_sum += cell_value
    
    # B10:C11 artıq merge olunub, sadəcə B10-a yaz
    diagonal_cell = ws["B10"]
    diagonal_cell.value = diagonal_sum
    diagonal_cell.font = bold_font
    diagonal_cell.alignment = center
    diagonal_cell.number_format = '#,##0.00'
    
    if use_type1:
        print(f"      ✓ Type1 diagonal cəm: {diagonal_sum:.2f} (E24+F23+G22+H21)")
    else:
        print(f"      ✓ Type2 diagonal cəm: {diagonal_sum:.2f} (E32+F31+G30+H29)")
    
    # ================== REZERV ARTIMI SƏTİRİ ==================
    if use_type1:
        reserve_row = 26  # E26:O26
        reserve_columns = 11  # E-dən O-yə qədər (11 sütun)
        ratio_row = 27  # E27:O27
        print(f"    Rezerv artımı yazılır (E26:O26)...")
    else:  # use_type2
        reserve_row = 34  # E34:W34
        reserve_columns = 19  # E-dən W-yə qədər (19 sütun)
        ratio_row = 35  # E35:W35
        print(f"    Rezerv artımı yazılır (E34:W34)...")
    
    for col_offset in range(reserve_columns):
        current_col = start_col_index + col_offset
        
        # Cəm sətirindən dəyər
        sum_value = ws.cell(row=sum_row, column=current_col).value
        
        # Çıxılacaq sətir: end_row-dan başlayıb yuxarıya (col_offset qədər)
        subtract_row = end_row - col_offset
        subtract_value = ws.cell(row=subtract_row, column=current_col).value
        
        # Rezerv artımı hesabla
        if isinstance(sum_value, (int, float)) and isinstance(subtract_value, (int, float)):
            reserve_increase = sum_value - subtract_value
        elif isinstance(sum_value, (int, float)):
            # Əgər subtract_value X-dirsə və ya yoxdursa
            reserve_increase = sum_value
        else:
            reserve_increase = 0
        
        # Rezerv artımı sətrinə yaz
        reserve_cell = ws.cell(row=reserve_row, column=current_col)
        reserve_cell.value = reserve_increase
        reserve_cell.font = font_style
        reserve_cell.border = thin
        reserve_cell.alignment = center
        reserve_cell.number_format = '#,##0.00'
    
    if use_type1:
        print(f"      ✓ Type1 rezerv artımı yazıldı (E26:O26)")
    else:
        print(f"      ✓ Type2 rezerv artımı yazıldı (E34:W34)")
    
    # ================== NİSBƏT SƏTİRİ (RATIO) ==================
    if use_type1:
        print(f"    Nisbət yazılır (E27:O27)...")
    else:
        print(f"    Nisbət yazılır (E35:W35)...")
    
    # ================== 1. ƏVVƏLCƏ REZERV ARTIMINDA 0 VAR MI YOXLA ==================
    has_zero = False
    for col_offset in range(reserve_columns):
        current_col = start_col_index + col_offset
        reserve_value = ws.cell(row=reserve_row, column=current_col).value
        
        if isinstance(reserve_value, (int, float)) and reserve_value == 0:
            has_zero = True
            break
    
    # ================== 2. ƏGƏR 0 VARSA, FAİZ SHEET-DƏN BİRBAŞA GÖTÜR ==================
    if has_zero:
        print(f"      ⚠ Rezerv artımında 0 aşkarlandı, Faiz sheet-dən nisbətlər birbaşa kopyalanır...")
        
        try:
            # Faiz sheet-ni oxu
            faiz_df = pd.read_excel(ucot_file, sheet_name='Faiz')
            
            # Product-a görə filter
            faiz_filtered = faiz_df[faiz_df['Product'] == product]
            
            if not faiz_filtered.empty:
                # İlk sətri götür
                faiz_row = faiz_filtered.iloc[0]
                
                if use_type1:
                    # Type1: D:N sütunlarını E27:O27-yə kopyala
                    print(f"        → Type1: Faiz D:N → E27:O27")
                    
                    for idx in range(11):  # 11 sütun: D, E, F, G, H, I, J, K, L, M, N
                        col_idx = 3 + idx  # D=3, E=4, ..., N=13 (pandas 0-based)
                        target_col = start_col_index + idx
                        
                        if col_idx < len(faiz_row):
                            faiz_value = faiz_row.iloc[col_idx]
                            
                            # DÜZƏLDİLMİŞ YOXLAMA: pd.isna və boş string yoxla
                            if not pd.isna(faiz_value) and faiz_value != '':
                                try:
                                    # Rəqəmə çevir (int64, float64 də daxil)
                                    numeric_value = float(faiz_value)
                                    
                                    ratio_cell = ws.cell(row=ratio_row, column=target_col)
                                    ratio_cell.value = numeric_value
                                    ratio_cell.font = font_style
                                    ratio_cell.border = thin
                                    ratio_cell.alignment = center
                                    ratio_cell.number_format = '0.0000'
                                    print(f"          ✓ Yazıldı: E27+{idx} = {numeric_value}")
                                except (ValueError, TypeError):
                                    # Rəqəmə çevrilə bilmirsə 0 yaz
                                    ratio_cell = ws.cell(row=ratio_row, column=target_col)
                                    ratio_cell.value = 0
                                    ratio_cell.font = font_style
                                    ratio_cell.border = thin
                                    ratio_cell.alignment = center
                                    ratio_cell.number_format = '0.0000'
                                    print(f"          ⚠ Rəqəm deyil, 0 yazıldı: E27+{idx}")
                            else:
                                # NaN və ya boş
                                ratio_cell = ws.cell(row=ratio_row, column=target_col)
                                ratio_cell.value = 0
                                ratio_cell.font = font_style
                                ratio_cell.border = thin
                                ratio_cell.alignment = center
                                ratio_cell.number_format = '0.0000'
                                print(f"          ⚠ NaN/Boş, 0 yazıldı: E27+{idx}")
                        else:
                            # Sütun yoxdursa 0 yaz
                            ratio_cell = ws.cell(row=ratio_row, column=target_col)
                            ratio_cell.value = 0
                            ratio_cell.font = font_style
                            ratio_cell.border = thin
                            ratio_cell.alignment = center
                            ratio_cell.number_format = '0.0000'
                            print(f"          ⚠ Sütun yoxdur, 0 yazıldı: E27+{idx}")
                
                else:  # Type2
                    # Type2: D:V sütunlarını E35:W35-ə kopyala
                    print(f"        → Type2: Faiz D:V → E35:W35")
                    
                    for idx in range(19):  # 19 sütun: D, E, F, ..., V
                        col_idx = 3 + idx  # D=3, E=4, ..., V=21 (pandas 0-based)
                        target_col = start_col_index + idx
                        
                        if col_idx < len(faiz_row):
                            faiz_value = faiz_row.iloc[col_idx]
                            
                            # DÜZƏLDİLMİŞ YOXLAMA: pd.isna və boş string yoxla
                            if not pd.isna(faiz_value) and faiz_value != '':
                                try:
                                    # Rəqəmə çevir (int64, float64 də daxil)
                                    numeric_value = float(faiz_value)
                                    
                                    ratio_cell = ws.cell(row=ratio_row, column=target_col)
                                    ratio_cell.value = numeric_value
                                    ratio_cell.font = font_style
                                    ratio_cell.border = thin
                                    ratio_cell.alignment = center
                                    ratio_cell.number_format = '0.0000'
                                    print(f"          ✓ Yazıldı: E35+{idx} = {numeric_value}")
                                except (ValueError, TypeError):
                                    # Rəqəmə çevrilə bilmirsə 0 yaz
                                    ratio_cell = ws.cell(row=ratio_row, column=target_col)
                                    ratio_cell.value = 0
                                    ratio_cell.font = font_style
                                    ratio_cell.border = thin
                                    ratio_cell.alignment = center
                                    ratio_cell.number_format = '0.0000'
                                    print(f"          ⚠ Rəqəm deyil, 0 yazıldı: E35+{idx}")
                            else:
                                # NaN və ya boş
                                ratio_cell = ws.cell(row=ratio_row, column=target_col)
                                ratio_cell.value = 0
                                ratio_cell.font = font_style
                                ratio_cell.border = thin
                                ratio_cell.alignment = center
                                ratio_cell.number_format = '0.0000'
                                print(f"          ⚠ NaN/Boş, 0 yazıldı: E35+{idx}")
                        else:
                            # Sütun yoxdursa 0 yaz
                            ratio_cell = ws.cell(row=ratio_row, column=target_col)
                            ratio_cell.value = 0
                            ratio_cell.font = font_style
                            ratio_cell.border = thin
                            ratio_cell.alignment = center
                            ratio_cell.number_format = '0.0000'
                            print(f"          ⚠ Sütun yoxdur, 0 yazıldı: E35+{idx}")
                
                print(f"        ✓ Faiz sheet-dən nisbətlər kopyalandı")
            
            else:
                print(f"        ⚠ Faiz sheet-də '{product}' üçün məlumat tapılmadı, normal hesablama edilir")
                
                # Faiz-də məlumat yoxdursa, normal hesabla
                for col_offset in range(reserve_columns):
                    current_col = start_col_index + col_offset
                    next_col = current_col + 1
                    
                    numerator = ws.cell(row=sum_row, column=next_col).value
                    denominator = ws.cell(row=reserve_row, column=current_col).value
                    
                    if isinstance(numerator, (int, float)) and isinstance(denominator, (int, float)) and denominator != 0:
                        ratio = round(numerator / denominator, 4)
                    else:
                        ratio = 0
                    
                    ratio_cell = ws.cell(row=ratio_row, column=current_col)
                    ratio_cell.value = ratio
                    ratio_cell.font = font_style
                    ratio_cell.border = thin
                    ratio_cell.alignment = center
                    ratio_cell.number_format = '0.0000'
        
        except Exception as e:
            print(f"        ✗ Faiz sheet oxunma xətası: {e}")
            
            # Xəta varsa, normal hesabla
            for col_offset in range(reserve_columns):
                current_col = start_col_index + col_offset
                next_col = current_col + 1
                
                numerator = ws.cell(row=sum_row, column=next_col).value
                denominator = ws.cell(row=reserve_row, column=current_col).value
                
                if isinstance(numerator, (int, float)) and isinstance(denominator, (int, float)) and denominator != 0:
                    ratio = round(numerator / denominator, 4)
                else:
                    ratio = 0
                
                ratio_cell = ws.cell(row=ratio_row, column=current_col)
                ratio_cell.value = ratio
                ratio_cell.font = font_style
                ratio_cell.border = thin
                ratio_cell.alignment = center
                ratio_cell.number_format = '0.0000'
    
    else:
        # ================== 3. 0 YOXDURSA, NORMAL NİSBƏT HESABLA ==================
        print(f"      → 0 aşkarlanmadı, normal nisbət hesablanır")
        
        for col_offset in range(reserve_columns):
            current_col = start_col_index + col_offset
            next_col = current_col + 1
            
            numerator = ws.cell(row=sum_row, column=next_col).value
            denominator = ws.cell(row=reserve_row, column=current_col).value
            
            if isinstance(numerator, (int, float)) and isinstance(denominator, (int, float)) and denominator != 0:
                ratio = round(numerator / denominator, 4)
            else:
                ratio = 0
            
            ratio_cell = ws.cell(row=ratio_row, column=current_col)
            ratio_cell.value = ratio
            ratio_cell.font = font_style
            ratio_cell.border = thin
            ratio_cell.alignment = center
            ratio_cell.number_format = '0.0000'
    
    if use_type1:
        print(f"      ✓ Type1 nisbət yazıldı (E27:O27)")
    else:
        print(f"      ✓ Type2 nisbət yazıldı (E35:W35)")
    
    # ================== KUMULATİV NİSBƏT (CUMULATIVE RATIO) ==================
    if use_type1:
        cumulative_row = 28  # E28:O28
        print(f"    Kumulativ nisbət yazılır (E28:O28)...")
    else:
        cumulative_row = 36  # E36:W36
        print(f"    Kumulativ nisbət yazılır (E36:W36)...")
    
    for col_offset in range(reserve_columns):
        current_col = start_col_index + col_offset
        
        # Cari sütundan son sütuna qədər (O27 və ya W35) nisbətləri vur
        cumulative_product = 1.0
        
        for multiply_offset in range(col_offset, reserve_columns):
            multiply_col = start_col_index + multiply_offset
            ratio_value = ws.cell(row=ratio_row, column=multiply_col).value
            
            if isinstance(ratio_value, (int, float)) and ratio_value != 0:
                cumulative_product *= ratio_value
            else:
                # Əgər nisbət 0-dırsa və ya yoxdursa, hasili 0 et
                cumulative_product = 0
                break
        
        # Kumulativ nisbəti yaz (4 onluq)
        cumulative_product = round(cumulative_product, 4)
        
        cumulative_cell = ws.cell(row=cumulative_row, column=current_col)
        cumulative_cell.value = cumulative_product
        cumulative_cell.font = font_style
        cumulative_cell.border = thin
        cumulative_cell.alignment = center
        cumulative_cell.number_format = '0.0000'  # 4 onluq format
    
    if use_type1:
        print(f"      ✓ Type1 kumulativ nisbət yazıldı (E28:O28)")
    else:
        print(f"      ✓ Type2 kumulativ nisbət yazıldı (E36:W36)")
    
    wb.save(excel_file)
    print(f"  ✅ {product}: Forma8_8 tamamlandı")
    
    # ================== TƏRS NİSBƏT (INVERSE RATIO) ==================
    if use_type1:
        inverse_row = 29  # E29:O29
        print(f"    Tərs nisbət yazılır (E29:O29)...")
    else:
        inverse_row = 37  # E37:W37
        print(f"    Tərs nisbət yazılır (E37:W37)...")
    
    for col_offset in range(reserve_columns):
        current_col = start_col_index + col_offset
        
        # Kumulativ nisbətdən dəyər al
        cumulative_value = ws.cell(row=cumulative_row, column=current_col).value
        
        # Tərs nisbət hesabla: 1 / kumulativ_nisbət (4 onluq)
        if isinstance(cumulative_value, (int, float)) and cumulative_value != 0:
            inverse_ratio = round(1.0 / cumulative_value, 4)
        else:
            inverse_ratio = 0  # Sıfıra bölmə xətası varsa 0
        
        # Tərs nisbət sətrinə yaz
        inverse_cell = ws.cell(row=inverse_row, column=current_col)
        inverse_cell.value = inverse_ratio
        inverse_cell.font = font_style
        inverse_cell.border = thin
        inverse_cell.alignment = center
        inverse_cell.number_format = '0.0000'  # 4 onluq format
    
    if use_type1:
        print(f"      ✓ Type1 tərs nisbət yazıldı (E29:O29)")
    else:
        print(f"      ✓ Type2 tərs nisbət yazıldı (E37:W37)")
    
    # ================== FORMA8_7-DƏN G SÜTUNUNU KOPYALAMA ==================
    print(f"    Forma8_7-dən G sütunu kopyalanır...")
    
    # Forma8_7 sheet-ni yoxla
    if "Forma8_7" not in wb.sheetnames:
        print(f"      ⚠ Forma8_7 sheet-i tapılmadı, skip edilir")
    else:
        ws_8_7 = wb["Forma8_7"]
        
        if use_type1:
            # Type1: G13:G24 → Q13:Q24
            source_start = 13
            source_end = 24
            print(f"      → Type1: G13:G24 → Q13:Q24")
        else:
            # Type2: G13:G32 → Q13:Q32
            source_start = 13
            source_end = 32
            print(f"      → Type2: G13:G32 → Q13:Q32")
        
        rows_copied = 0
        for row in range(source_start, source_end + 1):
            # Forma8_7-dən G sütununu oxu
            source_cell = ws_8_7.cell(row=row, column=7)  # G sütunu (column=7)
            
            # Forma8_8-də Q sütununa yaz
            target_cell = ws.cell(row=row, column=17)  # Q sütunu (column=17)
            
            # Dəyəri kopyala
            target_cell.value = source_cell.value
            
            # Formatı kopyala
            if source_cell.font:
                target_cell.font = source_cell.font.copy()
            else:
                target_cell.font = font_style
            
            if source_cell.border:
                target_cell.border = source_cell.border.copy()
            else:
                target_cell.border = thin
            
            if source_cell.fill:
                target_cell.fill = source_cell.fill.copy()
            
            if source_cell.alignment:
                target_cell.alignment = source_cell.alignment.copy()
            else:
                target_cell.alignment = center
            
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            else:
                target_cell.number_format = '#,##0.00'
            
            rows_copied += 1
        
        if use_type1:
            print(f"      ✓ Type1: {rows_copied} sətir kopyalandı (G13:G24 → Q13:Q24)")
        else:
            print(f"      ✓ Type2: {rows_copied} sətir kopyalandı (G13:G32 → Q13:Q32)")
    
    wb.save(excel_file)
    print(f"  ✅ {product}: Forma8_8 tamamlandı")