from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
import pandas as pd


def run_forma8_6(excel_file: str, reference_date: str):
    """Forma8_6 doldurur: Forma8_5-dən qrup toplamlarını oxuyub hesablamalar aparır"""
    
    wb = load_workbook(excel_file)
    
    # Sheet-lərin mövcudluğunu yoxla
    if "Forma8_1" not in wb.sheetnames:
        raise ValueError(f"{excel_file} faylında 'Forma8_1' sheet-i yoxdur!")
    if "Forma8_5" not in wb.sheetnames:
        raise ValueError(f"{excel_file} faylında 'Forma8_5' sheet-i yoxdur!")
    if "Forma8_6" not in wb.sheetnames:
        raise ValueError(f"{excel_file} faylında 'Forma8_6' sheet-i yoxdur!")
    
    ws_8_1 = wb["Forma8_1"]
    ws_8_5 = wb["Forma8_5"]
    ws = wb["Forma8_6"]
    
    # Product adını Forma8_1-dən götür
    product = ws_8_1["C8"].value if ws_8_1["C8"].value else None
    
    if not product:
        print(f"  ⚠ {excel_file}: Forma8_1-də product məlumatı yoxdur (C8 boşdur)")
        wb.save(excel_file)
        return
    
    # Forma8_6-nın C8-nə yaz
    ws["C8"].value = product
    ws["C8"].font = Font(name="A3 Times AZ Lat", size=10)
    ws["C8"].alignment = Alignment(horizontal="center", vertical="center")
    
    print(f"    Forma8_5-dən F sütununda bold toplamları oxuyur...")

    # ================== D6-YA TARİX YAZMA ==================
    ref_date = pd.to_datetime(reference_date)
    formatted_date = ref_date.strftime("%d.%m.%Y")
    ws["D6"].value = formatted_date
    print(f"    ✓ D6-ya tarix yazıldı: {formatted_date}")
    
    # ================== FORMA8_5-DƏN F SÜTUNUNDA BOLD TOPLAMLARI OXUMA ==================
    group_totals = []
    
    # 12-ci sətirdən başlayıb F sütununda (6-cı sütun) bold dəyərləri axtarırıq
    for row in range(12, 200):
        cell_f = ws_8_5.cell(row=row, column=6)
        
        # Əgər F-də dəyər var və bold-dursa
        if cell_f.value is not None and cell_f.font and cell_f.font.bold:
            val = float(cell_f.value)
            group_totals.append(val)
            print(f"      Bold toplam tapıldı (F{row}): {val:.2f}")
            
            # İLK 4 BOLD TOPLAMI götürdük (qrup toplamları)
            # 5-ci bold ümumi cəmdir, onu istəmirik
            if len(group_totals) >= 4:
                break
    
    # Əgər 4-dən az qrup tapdıqsa, qalanları 0 ilə doldur
    while len(group_totals) < 4:
        group_totals.append(0.0)
        print(f"      Qrup {len(group_totals)}: 0.00 (tapılmadı)")
    
    print(f"    Toplanan 4 qrup:")
    for i, val in enumerate(group_totals, 1):
        print(f"      Qrup {i}: {val:.2f}")
    
    # ================== STYLES ==================
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    font = Font(name="A3 Times AZ Lat", size=10)
    bold_font = Font(name="A3 Times AZ Lat", size=10, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    
    # ================== C12-C15: QRUP TOPLAMLARI ==================
    for i in range(4):
        row = 12 + i  # C12, C13, C14, C15
        val = group_totals[i]
        ws.cell(row=row, column=3).value = round(val, 2)
        ws.cell(row=row, column=3).font = font
        ws.cell(row=row, column=3).border = thin
        ws.cell(row=row, column=3).alignment = center
    
    # ================== E12-E15: C * D% ==================
    for row in [12, 13, 14, 15]:
        val_c = ws.cell(row=row, column=3).value
        val_d = ws.cell(row=row, column=4).value
        
        if val_c is not None and val_d is not None:
            val_e = round(float(val_c) * float(val_d) / 100, 2)
        else:
            val_e = 0
        
        ws.cell(row=row, column=5).value = val_e
        ws.cell(row=row, column=5).font = font
        ws.cell(row=row, column=5).border = thin
        ws.cell(row=row, column=5).alignment = center
    
    # ================== C16: C12+C13+C14+C15 ==================
    c16_sum = sum(group_totals)
    ws["C16"].value = round(c16_sum, 2)
    ws["C16"].font = bold_font
    ws["C16"].border = thin
    ws["C16"].alignment = center
    
    # ================== E16: E12+E13+E14+E15 ==================
    e16_sum = sum([ws.cell(row=r, column=5).value or 0 for r in [12, 13, 14, 15]])
    ws["E16"].value = round(e16_sum, 2)
    ws["E16"].font = bold_font
    ws["E16"].border = thin
    ws["E16"].alignment = center
    
    print(f"    Yazılan dəyərlər:")
    print(f"      C12={group_totals[0]:.2f}, C13={group_totals[1]:.2f}, C14={group_totals[2]:.2f}, C15={group_totals[3]:.2f}")
    print(f"      C16={c16_sum:.2f}, E16={e16_sum:.2f}")
    
    wb.save(excel_file)
    print(f"  ✅ {product}: Forma8_6 tamamlandı")