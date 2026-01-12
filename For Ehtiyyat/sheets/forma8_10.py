from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os
import pandas as pd
from dateutil.relativedelta import relativedelta

# Forma8_10(1) istifadə edən məhsullar
FORMA8_10_TYPE1_PRODUCTS = [
    "(01)FerdiQeza",
    "(02)Tibbi",
    "(03)EmlakYanginDigerRisk",
    "(04)AvtoKasko",
    "(05)DemiryolNeqliyyVasitesi",
    "(06)HavaNeqliyyKasko",
    "(07)SuNeqliyyKasko",
    "(08)Yuk",
    "(09)KendTeserrufBitki",
    "(10)KendTeserrufHeyvan",
    "(11)IshcilerinDeleduzlug",
    "(12)PulvePulSenedSaxtalash",
    "(22)Kredit",
    "(23)Ipoteka",
    "(24)EmlakinDeyerdenDushmesi",
    "(25)IshinDayanmasiRiski",
    "(27)SernishinIcbari",
    "(28)IcbariEkoloji",
    "(29)YanginIcbari",
    "(30)DeputatlarinIcbari",
    "(31)TibbiPersonalinAIDSden",
    "(32)HerbiQulluqcularinIcbari",
    "(33)HuquqMuhafifeIcbari",
    "(34)DovletQulluqcuIcbari",
    "(35)DiplomatlarinIcbari",
    "(37)IcbariDashinmazEmlak",
    "(39)IcbariNVSMMS",
    "(40)IcbariSernishinFerdiQeza",
    "(41)Sefer",
    "(42)Titul",
    "(43)HuquqiXerc"
]

# Forma8_10(2) istifadə edən məhsullar
FORMA8_10_TYPE2_PRODUCTS = [
    "(19)PesheMesuliyy",
    "(20)IshegoturenMesuliyy",
    "(21)UmumiMulkiMesuliyy",
    "(13)AvtoKonulluMesuliyy",
    "(14)DemiryolNeqliySahibMesuliyy",
    "(15)HavaNeqliySahibMesuliyy",
    "(16)SuNeqliySahibMesuliyy",
    "(17)YukDashiyanMesuliyy",
    "(18)MulkiMuqavileUzreMesuliyy",
    "(26)AvtoIcbariMesuliyy",
    "(36)AuditorPesheMesuliyyIcbari",
    "(38)IcbariDashinmazEmlakMesul"
]

# 0.01 hesablaması tələb edən məhsullar
CALCULATION_PRODUCTS = {
    "(04)AvtoKasko",
    "(08)Yuk",
    "(03)EmlakYanginDigerRisk",
    "(37)IcbariDashinmazEmlak"
}

def run_forma8_10(excel_file: str, previous_folder: str, reference_date: str):
    """Forma8_10 hazırlayır: Product-a görə düzgün sheet-i seçir və əvvəlki datanı kopyalayır"""
    
    wb = load_workbook(excel_file)
    
    # Sheet-lərin mövcudluğunu yoxla
    if "Forma8_1" not in wb.sheetnames:
        raise ValueError(f"{excel_file} faylında 'Forma8_1' sheet-i yoxdur!")
    
    ws_8_1 = wb["Forma8_1"]
    
    # Product adını Forma8_1-dən götür
    product = ws_8_1["C8"].value if ws_8_1["C8"].value else None
    
    if not product:
        print(f"  ⚠ {excel_file}: Forma8_1-də product məlumatı yoxdur (C8 boşdur)")
        wb.save(excel_file)
        return
    
    print(f"    Product: {product}")
    
    # ================== PRODUCT-A GÖRƏ TİP TƏYİN ET ==================
    use_type1 = product in FORMA8_10_TYPE1_PRODUCTS
    use_type2 = product in FORMA8_10_TYPE2_PRODUCTS
    
    if not use_type1 and not use_type2:
        print(f"  ⚠ {product}: Forma8_10 tip məlumatı tapılmadı, skip edilir")
        wb.save(excel_file)
        return
    
    if use_type1:
        print(f"    → Type1: Forma8_10(1)")
        sheet_to_keep = "Forma8_10(1)"
        sheet_to_delete = "Forma8_10(2)"
        prev_start_row = 14
        prev_end_row = 24
        new_start_row = 13
        d_target_row = 24
        f_target_row = 24
        copy_from_row = 23
        copy_to_row = 24
        calculation_row = 24
    else:  # use_type2
        print(f"    → Type2: Forma8_10(2)")
        sheet_to_keep = "Forma8_10(2)"
        sheet_to_delete = "Forma8_10(1)"
        prev_start_row = 14
        prev_end_row = 32
        new_start_row = 13
        d_target_row = 32
        f_target_row = 32
        copy_from_row = 31
        copy_to_row = 32
        calculation_row = 32
    
    # ================== SHEET-LƏRİ İDARƏ ET ==================
    if sheet_to_delete in wb.sheetnames:
        del wb[sheet_to_delete]
        print(f"    ✓ {sheet_to_delete} silindi")
    
    if sheet_to_keep not in wb.sheetnames:
        print(f"  ⚠ {sheet_to_keep} sheet-i tapılmadı!")
        wb.save(excel_file)
        return
    
    ws = wb[sheet_to_keep]
    ws.title = "Forma8_10"
    print(f"    ✓ {sheet_to_keep} → Forma8_10 adına dəyişdirildi")
    
    # ================== STYLES ==================
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    font_style = Font(name="A3 Times AZ Lat", size=10)
    bold_font = Font(name="A3 Times AZ Lat", size=10, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    
    # ================== C8-Ə PRODUCT YAZMA ==================
    ws["C8"].value = product
    ws["C8"].font = font_style
    ws["C8"].alignment = center

    # ================== D6-YA TARİX YAZMA ==================
    ref_date = pd.to_datetime(reference_date)
    formatted_date = ref_date.strftime("%d.%m.%Y")
    ws["D6"].value = formatted_date
    print(f"    ✓ D6-ya tarix yazıldı: {formatted_date}")
    
    # ================== ƏVVƏLKİ EXCEL-DƏN DATA KOPYALAMA ==================
    previous_file = os.path.join(previous_folder, f"{product}.xlsx")
    
    if not os.path.exists(previous_file):
        print(f"    ⚠ Əvvəlki Excel tapılmadı: {previous_file}")
    else:
        try:
            print(f"    Əvvəlki Excel-dən data kopyalanır: {os.path.basename(previous_file)}")
            # ✅ DÜZƏLDƏ: data_only=True - formulaları deyil, dəyərləri oxu
            wb_prev = load_workbook(previous_file, data_only=True)
            
            if "Forma8_10" not in wb_prev.sheetnames:
                print(f"    ⚠ Əvvəlki Excel-də Forma8_10 sheet-i yoxdur")
            else:
                ws_prev = wb_prev["Forma8_10"]
                
                rows_copied = 0
                for i, prev_row in enumerate(range(prev_start_row, prev_end_row + 1)):
                    new_row = new_start_row + i
                    
                    for col in range(4, 12):  # D-dən K-ya (4-11)
                        prev_cell = ws_prev.cell(row=prev_row, column=col)
                        new_cell = ws.cell(row=new_row, column=col)
                        
                        # ✅ Dəyəri kopyala (formula deyil)
                        new_cell.value = prev_cell.value
                        
                        # Format kopyalama (data_only=True ilə font məlumatı olmaya bilər)
                        if prev_cell.font:
                            new_cell.font = prev_cell.font.copy()
                        if prev_cell.border:
                            new_cell.border = prev_cell.border.copy()
                        if prev_cell.fill:
                            new_cell.fill = prev_cell.fill.copy()
                        if prev_cell.alignment:
                            new_cell.alignment = prev_cell.alignment.copy()
                        if prev_cell.number_format:
                            new_cell.number_format = prev_cell.number_format
                    
                    rows_copied += 1
                
                last_new_row = new_start_row + rows_copied - 1
                print(f"    ✓ {rows_copied} sətir kopyalandı (D{prev_start_row}:K{prev_end_row} → D{new_start_row}:K{last_new_row})")
                wb_prev.close()
        except Exception as e:
            print(f"    ⚠ Kopyalama xətası: {e}")
    
    # ================== FORMA8_4-DƏN SON 3 AYLIK F TOPLAMI ==================
    print(f"    Forma8_4-dən son 3 aylıq F toplamı hesablanır...")
    
    if "Forma8_4" not in wb.sheetnames:
        print(f"    ⚠ Forma8_4 sheet-i yoxdur, D{d_target_row} boş qalacaq")
    else:
        ws_8_4 = wb["Forma8_4"]
        
        ref_date = pd.to_datetime(reference_date)
        period_end = ref_date
        period_start = ref_date - relativedelta(months=3)
        
        print(f"      Period: {period_start.date()} - {period_end.date()}")
        
        forma8_4_data = []
        
        row = 12
        while row <= ws_8_4.max_row:
            val_c = ws_8_4.cell(row=row, column=3).value
            val_f = ws_8_4.cell(row=row, column=6).value
            
            if val_c and val_f:
                try:
                    if isinstance(val_c, str):
                        date_val = pd.to_datetime(val_c, errors='coerce')
                    else:
                        date_val = pd.to_datetime(val_c)
                    
                    if pd.notna(date_val):
                        forma8_4_data.append({
                            'date': date_val,
                            'value': val_f
                        })
                except:
                    pass
            
            row += 1
        
        if forma8_4_data:
            df = pd.DataFrame(forma8_4_data)
            df['value'] = pd.to_numeric(df['value'], errors='coerce').fillna(0)
            
            mask = (df['date'] >= period_start) & (df['date'] < period_end)
            period_sum = df.loc[mask, 'value'].sum()
            
            print(f"      Toplam {len(df.loc[mask])} sətir tapıldı")
        else:
            period_sum = 0
            print(f"      Heç bir data tapılmadı")
        
        ws.cell(row=d_target_row, column=4).value = round(period_sum, 2)
        ws.cell(row=d_target_row, column=4).font = bold_font
        ws.cell(row=d_target_row, column=4).border = thin
        ws.cell(row=d_target_row, column=4).alignment = center
        
        print(f"      ✓ D{d_target_row} = {period_sum:.2f}")
    
    # ================== F→E KOPYALAMA ==================
    print(f"    F{copy_from_row} → E{copy_to_row} kopyalanır...")
    
    source_cell = ws.cell(row=copy_from_row, column=6)
    target_cell = ws.cell(row=copy_to_row, column=5)
    
    target_cell.value = source_cell.value
    
    if source_cell.font:
        target_cell.font = source_cell.font.copy()
    if source_cell.border:
        target_cell.border = source_cell.border.copy()
    if source_cell.fill:
        target_cell.fill = source_cell.fill.copy()
    if source_cell.alignment:
        target_cell.alignment = source_cell.alignment.copy()
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    
    print(f"      ✓ E{copy_to_row} = {target_cell.value}")
    
    # ================== FORMA8_5-DƏN ÜMUMİ F TOPLAMI ==================
    print(f"    Forma8_5-dən ümumi F toplamı götürülür...")
    
    if "Forma8_5" not in wb.sheetnames:
        print(f"    ⚠ Forma8_5 sheet-i yoxdur, F{f_target_row} boş qalacaq")
    else:
        ws_8_5 = wb["Forma8_5"]
        
        f_total = None
        
        for row in range(ws_8_5.max_row, 11, -1):
            val_f = ws_8_5.cell(row=row, column=6).value
            
            if val_f is not None:
                try:
                    f_total = float(val_f)
                    print(f"      Forma8_5 sətir {row}: F = {f_total:.2f}")
                    break
                except (ValueError, TypeError):
                    continue
        
        if f_total is None:
            print(f"      ⚠ Forma8_5-də F toplamı tapılmadı")
            f_total = 0
        
        ws.cell(row=f_target_row, column=6).value = round(f_total, 2)
        ws.cell(row=f_target_row, column=6).font = bold_font
        ws.cell(row=f_target_row, column=6).border = thin
        ws.cell(row=f_target_row, column=6).alignment = center
        
        print(f"      ✓ F{f_target_row} = {f_total:.2f}")
    
    # ================== G SÜTUNU HESABLAMA: D + E - F ==================
    print(f"    G{calculation_row} hesablanır: D + E - F")
    
    d_value = ws.cell(row=calculation_row, column=4).value or 0
    e_value = ws.cell(row=calculation_row, column=5).value or 0
    f_value = ws.cell(row=calculation_row, column=6).value or 0
    
    try:
        d_value = float(d_value)
    except (ValueError, TypeError):
        d_value = 0
    
    try:
        e_value = float(e_value)
    except (ValueError, TypeError):
        e_value = 0
    
    try:
        f_value = float(f_value)
    except (ValueError, TypeError):
        f_value = 0
    
    g_value = d_value + e_value - f_value
    
    target_g_cell = ws.cell(row=calculation_row, column=7)
    target_g_cell.value = round(g_value, 2)
    target_g_cell.font = bold_font
    target_g_cell.border = thin
    target_g_cell.alignment = center
    
    print(f"      ✓ G{calculation_row} = {g_value:.2f}")
    
    # ================== H, I, J, K SÜTUNLARI HESABLAMA ==================
    # ✅ Yalnız xüsusi məhsullar üçün 0.01 hesablaması
    if product in CALCULATION_PRODUCTS:
        print(f"    H, I, J, K sütunları hesablanır (0.01 əmsalı)")
        
        # H = D * 0.01
        h_value = d_value * 0.01
        h_cell = ws.cell(row=calculation_row, column=8)
        h_cell.value = round(h_value, 2)
        h_cell.font = bold_font
        h_cell.border = thin
        h_cell.alignment = center
        
        # I = E * 0.01
        i_value = e_value * 0.01
        i_cell = ws.cell(row=calculation_row, column=9)
        i_cell.value = round(i_value, 2)
        i_cell.font = bold_font
        i_cell.border = thin
        i_cell.alignment = center
        
        # J = F * 0.01
        j_value = f_value * 0.01
        j_cell = ws.cell(row=calculation_row, column=10)
        j_cell.value = round(j_value, 2)
        j_cell.font = bold_font
        j_cell.border = thin
        j_cell.alignment = center
        
        # K = G * 0.01
        k_value = g_value * 0.01
        k_cell = ws.cell(row=calculation_row, column=11)
        k_cell.value = round(k_value, 2)
        k_cell.font = bold_font
        k_cell.border = thin
        k_cell.alignment = center
        
        print(f"      ✓ H{calculation_row} = {h_value:.2f}, I = {i_value:.2f}, J = {j_value:.2f}, K = {k_value:.2f}")
    else:
        print(f"    H, I, J, K sütunları 0.00 yazılır (məhsul {product})")
        
        # H, I, J, K hamısı 0.00
        for col in [8, 9, 10, 11]:  # H, I, J, K
            cell = ws.cell(row=calculation_row, column=col)
            cell.value = 0.00
            cell.font = bold_font
            cell.border = thin
            cell.alignment = center
        
        print(f"      ✓ H{calculation_row} = I = J = K = 0.00")
    
    wb.save(excel_file)
    print(f"  ✅ {product}: Forma8_10 tamamlandı")