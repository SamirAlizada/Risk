import pandas as pd
import shutil
import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment

# Rate konfiqurasiyası
PRODUCT_RATES = {
    "(04)AvtoKasko": 0.02,
    "(08)Yuk": 0.01,
    "(03)EmlakYanginDigerRisk": 0.20,
    "(37)IcbariDashinmazEmlak": 0.20
}

def get_rate(product: str) -> float:
    """Product üçün rate qaytarır"""
    return PRODUCT_RATES.get(product, 0)

def run_forma8_1(ucot_file: str, template_file: str, reference_date: str, output_folder: str):
    """Forma8_1 yaradır: hər məhsul üçün ayrı Excel"""
    
    os.makedirs(output_folder, exist_ok=True)

    # ================== STYLES ==================
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    font = Font(name="A3 Times AZ Lat", size=10)
    bold_font = Font(name="A3 Times AZ Lat", size=10, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    ROW_HEIGHT = 28

    # ================== UCOT OXUMA ==================
    df = pd.read_excel(ucot_file, sheet_name="Simple")
    
    # Lazımi sütunların mövcudluğunu yoxla
    required_cols = ["I", "II", "III", "VII", "XI"]
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        raise ValueError(f"UCOT faylında bu sütunlar tapılmadı: {missing_cols}")
    
    products = df["I"].dropna().unique()
    print(f"  Tapılan məhsullar: {len(products)}")

    # ================== HƏR MƏHSUL ÜÇÜN ==================
    for idx, product in enumerate(products, 1):
        print(f"  [{idx}/{len(products)}] {product} işlənir...", end=" ")
        
        # Yalnız II dolu və VII > 0 olan sətirləri götür
        filtered = df[
            (df["I"] == product) & 
            (df["II"].notna()) &
            (df["VII"].notna()) &
            (df["VII"] > 0)  # ✅ VII sütunu 0-dan böyük olmalıdır
        ][["II", "III", "VII", "XI"]].copy()

        # Excel faylını kopyala
        output_file = os.path.join(output_folder, f"{product}.xlsx")
        shutil.copy(template_file, output_file)

        wb = load_workbook(output_file)
        ws = wb["Forma8_1"]

        # Məhsul adını yaz
        ws["C8"] = product
        ws["C8"].font = font
        ws["C8"].alignment = center

        # ================== D6-YA TARİX YAZMA ==================
        ref_date = pd.to_datetime(reference_date)
        formatted_date = ref_date.strftime("%d.%m.%Y")
        ws["D6"].value = formatted_date
        print(f"    ✓ D6-ya tarix yazıldı: {formatted_date}")

        # Merge-ləri aç
        for m in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(m))

        # Əgər data yoxdursa
        if filtered.empty:
            wb.save(output_file)
            print("(boş)")
            continue

        # Data hazırlığı
        filtered["VII"] = filtered["VII"].fillna(0)
        filtered["XI"] = filtered["XI"].fillna(0)

        start_row = 12
        ws.insert_rows(start_row, amount=len(filtered))

        # Toplam dəyişənləri
        totals = {"D": 0, "E": 0, "F": 0, "G": 0, "H": 0, "I": 0}
        r = start_row

        # Rate al
        rate = get_rate(product)
        has_calculation = rate > 0  # ✅ Hesablama olub-olmadığını yoxla

        # ================== HƏR SƏTİR ÜÇÜN ==================
        for counter, (_, row) in enumerate(filtered.iterrows(), start=1):
            vii = round(float(row["VII"]), 2)
            xi = round(float(row["XI"]), 2)

            # 15% limit tətbiq et
            limit = round(vii * 0.15, 2)
            final_xi = min(xi, limit)  # ✅ min() daha aydındır

            # Hesablamalar (yalnız müəyyən məhsullar üçün)
            val_G = round(vii * rate, 2) if has_calculation else 0
            val_H = round(final_xi * rate, 2) if has_calculation else 0
            val_I = round(val_G - val_H, 2) if has_calculation else 0

            # Sətir məlumatları
            values = [
                f"A{counter}",  # ✅ Sadə counter
                row["II"],  # Policy nömrəsi
                pd.to_datetime(row["III"]).date() if pd.notna(row["III"]) else "",
                vii,
                final_xi,
                round(vii - final_xi, 2),
                val_G,
                val_H,
                val_I
            ]

            # Excel-ə yaz
            for c, v in enumerate(values, start=1):
                cell = ws.cell(row=r, column=c)
                cell.value = v
                cell.font = font
                cell.border = thin
                cell.alignment = center
                if c == 3 and v != "":  # Tarix formatı
                    cell.number_format = "DD.MM.YYYY"

            # Toplamları yığ
            totals["D"] += values[3]
            totals["E"] += values[4]
            totals["F"] += values[5]
            totals["G"] += values[6]
            totals["H"] += values[7]
            totals["I"] += values[8]

            ws.row_dimensions[r].height = ROW_HEIGHT
            r += 1

        # ================== YEKUN SƏTİRİ ==================
        ws.cell(row=r, column=1).value = "AA1"  # ✅ Yekun üçün xüsusi etiket
        ws.cell(row=r, column=1).font = bold_font
        ws.cell(row=r, column=1).border = thin
        ws.cell(row=r, column=1).alignment = center
        
        ws.cell(row=r, column=2).value = "Yekun BSH"
        ws.cell(row=r, column=2).font = bold_font
        ws.cell(row=r, column=2).border = thin
        ws.cell(row=r, column=2).alignment = center

        # Toplam dəyərləri yaz
        for i, col_key in enumerate(["D", "E", "F", "G", "H", "I"], start=4):
            cell = ws.cell(row=r, column=i)
            cell.value = round(totals[col_key], 2)
            cell.font = bold_font
            cell.border = thin
            cell.alignment = center

        # 4 sətir aşağıya kopyala (yalnız D, E, F, G, H, I)
        copy_row = r + 4
        for i, col_key in enumerate(["D", "E", "F", "G", "H", "I"], start=4):
            cell = ws.cell(row=copy_row, column=i)
            cell.value = round(totals[col_key], 2)
            cell.font = bold_font
            cell.border = thin
            cell.alignment = center

        wb.save(output_file)
        print(f"✓ ({len(filtered)} sətir)")