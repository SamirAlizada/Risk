from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
import pandas as pd

# Rate konfiqurasiyası (G sütunu üçün)
PRODUCT_RATES = {
    "(04)AvtoKasko": 0.01,
    "(08)Yuk": 0.01,
    "(03)EmlakYanginDigerRisk": 0.20,
    "(37)IcbariDashinmazEmlak": 0.20
}

def get_rate(product: str) -> float:
    """Product üçün rate qaytarır"""
    return PRODUCT_RATES.get(product, 0)

def run_forma8_4(excel_file: str, ucot_file: str):
    """Forma8_4 doldurur: Təkrarsığortaçıları qruplaşdırır"""
    
    wb = load_workbook(excel_file)
    
    # Sheet-lərin mövcudluğunu yoxla
    if "Forma8_1" not in wb.sheetnames:
        raise ValueError(f"{excel_file} faylında 'Forma8_1' sheet-i yoxdur!")
    if "Forma8_4" not in wb.sheetnames:
        raise ValueError(f"{excel_file} faylında 'Forma8_4' sheet-i yoxdur!")
    
    ws_8_1 = wb["Forma8_1"]
    ws = wb["Forma8_4"]

    # ================== UCOT OXUMA ==================
    df_ucot = pd.read_excel(ucot_file, sheet_name="Simple")
    
    # Lazımi sütunları yoxla
    required_cols = [
        "I", "II", 
        "XVIII", "XIX", "XX", "XXI",           # Qrup identifikasiyası
        "XXVIII", "XXIX", "XXX", "XXXI",       # E sütunu üçün
        "XXXII", "XXXIII", "XXXIV", "XXXV",    # H sütunu üçün
        "XXXVIII", "XXXIX"                      # B, C sütunları üçün
    ]
    missing_cols = [col for col in required_cols if col not in df_ucot.columns]
    if missing_cols:
        raise ValueError(f"UCOT faylında bu sütunlar tapılmadı: {missing_cols}")
    
    # Product adını Forma8_1-in C8-dən götür
    product = ws_8_1["C8"].value if ws_8_1["C8"].value else None
    
    if not product:
        print(f"  ⚠ {excel_file}: Forma8_1-də product məlumatı yoxdur (C8 boşdur)")
        wb.save(excel_file)
        return
    
    # Forma8_4-ün C8-nə də yaz
    ws["C8"].value = product
    ws["C8"].font = Font(name="A3 Times AZ Lat", size=14)
    ws["C8"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Bu məhsula aid dataları götür
    df_product = df_ucot[df_ucot["I"] == product].copy()
    
    if df_product.empty:
        print(f"  ⚠ {product}: UCOT-da data tapılmadı")
        wb.save(excel_file)
        return
    
    # Qrup sütunlarını yoxla və hazırla
    for col in ["XVIII", "XIX", "XX", "XXI"]:
        df_product[col] = pd.to_numeric(df_product[col], errors='coerce').fillna(0)
    
    # ================== QRUPLARA BÖLÜŞDÜRMƏ ==================
    groups = {
        1: df_product[df_product["XVIII"] > 0].copy(),
        2: df_product[df_product["XIX"] > 0].copy(),
        3: df_product[df_product["XX"] > 0].copy(),
        4: df_product[df_product["XXI"] > 0].copy()
    }
    
    # Debug: qrupları yoxla
    print(f"    Debug - Qrup nəticələri:")
    for g_num in [1, 2, 3, 4]:
        print(f"      Qrup {g_num}: {len(groups[g_num])} sətir")
    
    # ✅ Boş qrupları çıxarmırıq - hamısını saxlayırıq
    
    # ================== STYLES ==================
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    font = Font(name="A3 Times AZ Lat", size=14)
    bold_font = Font(name="A3 Times AZ Lat", size=14, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    ROW_HEIGHT = 28
    
    # ================== MERGE-LƏRİ AÇ ==================
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))
    
    # ================== QRUPLARI YAZMAQ ==================
    current_row = 12  # Başlanğıc sətri
    
    # Rate-i al
    rate = get_rate(product)
    
    # Ümumi toplamlar (bütün qruplar üçün)
    grand_totals = {"D": 0, "E": 0, "F": 0, "G": 0, "H": 0, "I": 0}
    
    # Qrup üçün prefix
    group_prefix = {1: "A", 2: "B", 3: "C", 4: "D"}
    
    # Qrup üçün müvafiq sütun adları
    group_col_mapping = {
        1: {"qrup": "XVIII", "e_col": "XXVIII", "h_col": "XXXII"},
        2: {"qrup": "XIX", "e_col": "XXIX", "h_col": "XXXIII"},
        3: {"qrup": "XX", "e_col": "XXX", "h_col": "XXXIV"},
        4: {"qrup": "XXI", "e_col": "XXXI", "h_col": "XXXV"}
    }
    
    # ✅ HƏMIŞƏ 4 QRUPU İŞLƏ
    for group_num in [1, 2, 3, 4]:
        df_group = groups[group_num]
        group_size = len(df_group)
        
        if group_size > 0:
            print(f"    Qrup {group_num}: {group_size} sətir yazılır...")
        else:
            print(f"    Qrup {group_num}: boşdur (0 ilə toplam yazılır)")
        
        col_names = group_col_mapping[group_num]
        
        # Qrup toplamları
        group_totals = {"D": 0, "E": 0, "F": 0, "G": 0, "H": 0, "I": 0}
        
        # ✅ Əgər qrupda data varsa
        if group_size > 0:
            # Sətirləri əlavə et
            ws.insert_rows(current_row, amount=group_size)

            for idx, (_, row) in enumerate(df_group.iterrows(), start=1):
                r = current_row + idx - 1

                # A sütunu - qrup üzrə sıra nömrəsi (A1, B1, C1, D1...)
                cell_a = ws.cell(row=r, column=1)
                prefix = group_prefix.get(group_num, "A")
                cell_a.value = f"{prefix}{idx}"
                cell_a.font = font
                cell_a.border = thin
                cell_a.alignment = center
                
                # B sütunu - XXXVIII
                cell_b = ws.cell(row=r, column=2)
                cell_b.value = row["XXXVIII"] if pd.notna(row["XXXVIII"]) else ""
                cell_b.font = font
                cell_b.border = thin
                cell_b.alignment = center
                
                # C sütunu - XXXIX
                cell_c = ws.cell(row=r, column=3)
                cell_c.value = row["XXXIX"] if pd.notna(row["XXXIX"]) else ""
                cell_c.font = font
                cell_c.border = thin
                cell_c.alignment = center

                # Tarix format
                if pd.notna(row["XXXIX"]):
                    cell_c.number_format = "DD.MM.YYYY"
                
                # D sütunu - Qrupa uyğun sütundan (XVIII, XIX, XX, XXI)
                val_d = row[col_names["qrup"]] if pd.notna(row[col_names["qrup"]]) else 0
                val_d = round(float(val_d), 2)
                cell_d = ws.cell(row=r, column=4)
                cell_d.value = val_d
                cell_d.font = font
                cell_d.border = thin
                cell_d.alignment = center
                group_totals["D"] += val_d
                
                # E sütunu - Qrupa uyğun sütundan (XXVIII, XXIX, XXX, XXXI)
                val_e = row[col_names["e_col"]] if pd.notna(row[col_names["e_col"]]) else 0
                val_e = round(float(val_e), 2)
                cell_e = ws.cell(row=r, column=5)
                cell_e.value = val_e
                cell_e.font = font
                cell_e.border = thin
                cell_e.alignment = center
                group_totals["E"] += val_e
                
                # F sütunu - D - E
                val_f = round(val_d - val_e, 2)
                cell_f = ws.cell(row=r, column=6)
                cell_f.value = val_f
                cell_f.font = font
                cell_f.border = thin
                cell_f.alignment = center
                group_totals["F"] += val_f
                
                # G sütunu - D * rate (məhsula görə)
                val_g = round(val_d * rate, 2)
                cell_g = ws.cell(row=r, column=7)
                cell_g.value = val_g
                cell_g.font = font
                cell_g.border = thin
                cell_g.alignment = center
                group_totals["G"] += val_g
                
                # H sütunu - Qrupa uyğun sütundan (XXXII, XXXIII, XXXIV, XXXV)
                val_h = row[col_names["h_col"]] if pd.notna(row[col_names["h_col"]]) else 0
                val_h = round(float(val_h), 2)
                cell_h = ws.cell(row=r, column=8)
                cell_h.value = val_h
                cell_h.font = font
                cell_h.border = thin
                cell_h.alignment = center
                group_totals["H"] += val_h
                
                # I sütunu - G - H
                val_i = round(val_g - val_h, 2)
                cell_i = ws.cell(row=r, column=9)
                cell_i.value = val_i
                cell_i.font = font
                cell_i.border = thin
                cell_i.alignment = center
                group_totals["I"] += val_i
                
                ws.row_dimensions[r].height = ROW_HEIGHT
        
        # ✅ Qrup toplam sətri (data olsa da, olmasasa - həmişə yazılır)
        total_row = current_row + group_size
        
        # D, E, F, G, H, I toplamları
        for col_idx, col_key in enumerate(["D", "E", "F", "G", "H", "I"], start=4):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.value = round(group_totals[col_key], 2)
            cell.font = bold_font
            cell.border = thin
            cell.alignment = center
            
            # Ümumi toplama əlavə et
            grand_totals[col_key] += group_totals[col_key]
        
        ws.row_dimensions[total_row].height = ROW_HEIGHT
        
        # Növbəti qrup üçün 2 sətir aşağıdan başla
        current_row += group_size + 2
    
    # ✅ ÜMUMİ CƏM (4-cü qrupdan sonra - həmişə)
    grand_total_row = current_row - 1
    
    print(f"    ÜMUMİ CƏM (sətir {grand_total_row}): F={grand_totals['F']:.2f}, I={grand_totals['I']:.2f}")
    
    # D, E, F, G, H, I ümumi toplamları
    for col_idx, col_key in enumerate(["D", "E", "F", "G", "H", "I"], start=4):
        cell = ws.cell(row=grand_total_row, column=col_idx)
        cell.value = round(grand_totals[col_key], 2)
        cell.font = bold_font
        cell.border = thin
        cell.alignment = center
    
    ws.row_dimensions[grand_total_row].height = ROW_HEIGHT
    
    wb.save(excel_file)
    print(f"  ✅ {product}: 4 qrup işləndi")