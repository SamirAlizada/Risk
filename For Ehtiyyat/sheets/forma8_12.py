from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import pandas as pd
from dateutil.relativedelta import relativedelta

# Forma8_12(1) istifadə edən məhsullar
FORMA8_12_TYPE1_PRODUCTS = [
    "(01)FerdiQeza",
    "(02)Tibbi",
    "(03)EmlakYanginDigerRisk",
    "(04)AvtoKasko",
    "(05)DemiryolNeqliyyVasitesi",
    "(06)HavaNeqliyyKasko",
    "(07)SuNeqliyyKasko",
    "(08)Yuk",
    "(09)KendTeserrufBitki",
    "(10)KendTeserrufHeyvan",
    "(11)IshcilerinDeleduzlug",
    "(12)PulvePulSenedSaxtalash",
    "(22)Kredit",
    "(23)Ipoteka",
    "(24)EmlakinDeyerdenDushmesi",
    "(25)IshinDayanmasiRiski",
    "(27)SernishinIcbari",
    "(28)IcbariEkoloji",
    "(29)YanginIcbari",
    "(30)DeputatlarinIcbari",
    "(31)TibbiPersonalinAIDSden",
    "(32)HerbiQulluqcularinIcbari",
    "(33)HuquqMuhafifeIcbari",
    "(34)DovletQulluqcuIcbari",
    "(35)DiplomatlarinIcbari",
    "(37)IcbariDashinmazEmlak",
    "(39)IcbariNVSMMS",
    "(40)IcbariSernishinFerdiQeza",
    "(41)Sefer",
    "(42)Titul",
    "(43)HuquqiXerc"
]

# Forma8_12(2) istifadə edən məhsullar
FORMA8_12_TYPE2_PRODUCTS = [
    "(19)PesheMesuliyy",
    "(20)IshegoturenMesuliyy",
    "(21)UmumiMulkiMesuliyy",
    "(13)AvtoKonulluMesuliyy",
    "(14)DemiryolNeqliySahibMesuliyy",
    "(15)HavaNeqliySahibMesuliyy",
    "(16)SuNeqliySahibMesuliyy",
    "(17)YukDashiyanMesuliyy",
    "(18)MulkiMuqavileUzreMesuliyy",
    "(26)AvtoIcbariMesuliyy",
    "(36)AuditorPesheMesuliyyIcbari",
    "(38)IcbariDashinmazEmlakMesul"
]

def run_forma8_12(excel_file: str, reference_date: str, ucot_file: str):
    """Forma8_12 hazırlayır: Product-a görə düzgün sheet-i seçir, tarixləri və hesablamaları yazır"""
    
    wb = load_workbook(excel_file)
    
    # Sheet-lərin mövcudluğunu yoxla
    if "Forma8_1" not in wb.sheetnames:
        raise ValueError(f"{excel_file} faylında 'Forma8_1' sheet-i yoxdur!")
    
    ws_8_1 = wb["Forma8_1"]
    
    # Product adını Forma8_1-dən götür
    product = ws_8_1["C8"].value if ws_8_1["C8"].value else None
    
    if not product:
        print(f"  ⚠ {excel_file}: Forma8_1-də product məlumatı yoxdur (C8 boşdur)")
        wb.save(excel_file)
        return
    
    print(f"    Product: {product}")
    
    # ================== PRODUCT-A GÖRƏ TİP TƏYİN ET ==================
    use_type1 = product in FORMA8_12_TYPE1_PRODUCTS
    use_type2 = product in FORMA8_12_TYPE2_PRODUCTS
    
    if not use_type1 and not use_type2:
        print(f"  ⚠ {product}: Forma8_12 tip məlumatı tapılmadı, skip edilir")
        wb.save(excel_file)
        return
    
    if use_type1:
        print(f"    → Type1: Forma8_12(1)")
        sheet_to_keep = "Forma8_12(1)"
        sheet_to_delete = "Forma8_12(2)"
        end_row = 24      # D24 - son tarix (reference date)
        start_row = 13    # D13 - ilk tarix
        period_count = 12  # 12 period (3 il × 4 rüb = 12)
        total_columns = 12  # E-dən P-yə qədər (12 sütun)
        sum_row = 25      # E25 - cəm sətiri
        # Diagonal cəm üçün koordinatlar: E24, F23, G22, H21
        diagonal_coords = [(24, 5), (23, 6), (22, 7), (21, 8)]  # (row, col)
    else:  # use_type2
        print(f"    → Type2: Forma8_12(2)")
        sheet_to_keep = "Forma8_12(2)"
        sheet_to_delete = "Forma8_12(1)"
        end_row = 32      # D32 - son tarix (reference date)
        start_row = 13    # D13 - ilk tarix
        period_count = 20  # 20 period (5 il × 4 rüb = 20)
        total_columns = 20  # E-dən AF-ə qədər (20 sütun)
        sum_row = 33      # E33 - cəm sətiri
        # Diagonal cəm üçün koordinatlar: E32, F31, G30, H29
        diagonal_coords = [(32, 5), (31, 6), (30, 7), (29, 8)]  # (row, col)
    
    # ================== SHEET-LƏRİ İDARƏ ET ==================
    if sheet_to_delete in wb.sheetnames:
        del wb[sheet_to_delete]
        print(f"    ✓ {sheet_to_delete} silindi")
    
    if sheet_to_keep not in wb.sheetnames:
        print(f"  ⚠ {sheet_to_keep} sheet-i tapılmadı!")
        wb.save(excel_file)
        return
    
    ws = wb[sheet_to_keep]
    ws.title = "Forma8_12"
    print(f"    ✓ {sheet_to_keep} → Forma8_12 adına dəyişdirildi")

    # ================== STYLES ==================
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    font_style = Font(name="A3 Times AZ Lat", size=10)
    bold_font = Font(name="A3 Times AZ Lat", size=10, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    
    # ================== C8-Ə PRODUCT YAZMA ==================
    ws["C8"].value = product
    ws["C8"].font = font_style
    ws["C8"].alignment = center

    # ================== D6-YA TARİX YAZMA ==================
    ref_date = pd.to_datetime(reference_date)
    formatted_date = ref_date.strftime("%d.%m.%Y")
    ws["D6"].value = formatted_date
    print(f"    ✓ D6-ya tarix yazıldı: {formatted_date}")
    
    # ================== TARİXLƏRİ YAZMA (3 AY INTERVALI) ==================
    print(f"    Tarixlər yazılır (D{start_row}:D{end_row})...")
    
    # Reference date-i parse et
    ref_date = pd.to_datetime(reference_date)
    
    # Son sətirdən (end_row) başlayıb yuxarıya doğru (start_row-a qədər)
    current_date = ref_date
    
    for row in range(end_row, start_row - 1, -1):  # end_row-dan start_row-a qədər geriyə
        # D sütununa tarixi yaz
        cell = ws.cell(row=row, column=4)  # D sütunu (column=4)
        cell.value = current_date
        cell.font = font_style
        cell.border = thin
        cell.alignment = center
        cell.number_format = "DD.MM.YYYY"  # Tarix formatı
        
        # 3 ay geriyə get (növbəti iterasiya üçün)
        current_date = current_date - relativedelta(months=3)
    
    print(f"      ✓ {period_count} tarix yazıldı (son: {ref_date.strftime('%d.%m.%Y')}, ilk: {(current_date + relativedelta(months=3)).strftime('%d.%m.%Y')})")
    
    # ================== HESABLAMALAR (TYPE1 VƏ TYPE2) ==================
    print(f"    Hesablamaları yazılır...")

    # Mənbə fayldan DataFrame-i oxu
    zerer_df = pd.read_excel(ucot_file, sheet_name='Zerer')

    # Tarix sütunlarını datetime tipinə çevir
    zerer_df['Sığоrtа hаdisəsinin bаş verdiyi tarixi'] = pd.to_datetime(
        zerer_df['Sığоrtа hаdisəsinin bаş verdiyi tarixi'], errors='coerce'
    )
    zerer_df['Sığorta ödənişi Tаriхi'] = pd.to_datetime(
        zerer_df['Sığorta ödənişi Tаriхi'], errors='coerce'
    )

    start_col_index = 5  # E sütunu

    # Üçbucaq formatında doldur
    for row in range(start_row, end_row + 1):
        # Cari sətirin D tarixi (hadisə intervali üçün)
        D_cell_row_value = ws.cell(row=row, column=4).value
        D_cell_row = pd.to_datetime(D_cell_row_value, errors='coerce')
        
        if pd.isna(D_cell_row):
            # Bu sətir üçün bütün sütunları X et
            for col_offset in range(total_columns):
                ws.cell(row=row, column=start_col_index + col_offset, value='X')
            continue
        
        col_offset = 0
        
        # Sağa doğru: D13, D14, D15... (ödəniş tarixi üçün)
        for payment_row in range(row, end_row + 1):
            current_col = start_col_index + col_offset
            
            # Ödəniş tarixini oxu
            D_cell_payment_value = ws.cell(row=payment_row, column=4).value
            D_cell_payment = pd.to_datetime(D_cell_payment_value, errors='coerce')
            
            if pd.isna(D_cell_payment):
                ws.cell(row=row, column=current_col, value='X')
                col_offset += 1
                continue
            
            # Zərər hesabla
            zerer_filter = (
                (zerer_df['Sığоrtа hаdisəsinin bаş verdiyi tarixi'] < D_cell_row) &
                (zerer_df['Sığоrtа hаdisəsinin bаş verdiyi tarixi'] >= D_cell_row - relativedelta(months=3)) &
                (zerer_df['Sığorta ödənişi Tаriхi'] < D_cell_payment) &
                (zerer_df['Sığоrtаnın sinifləri'] == product)
            )
            zerer_sum = zerer_df.loc[zerer_filter, 'Sığorta ödənişi Təkrаrsığоrtаçının pаyı'].sum()
            
            # Nəticə (sadəcə zərər)
            result = zerer_sum
            
            # Excel-ə yaz
            cell = ws.cell(row=row, column=current_col)
            cell.value = result
            cell.font = font_style
            cell.border = thin
            cell.alignment = center
            cell.number_format = '#,##0.00'  # Rəqəm formatı
            
            col_offset += 1
        
        # Qalan sütunlara X yaz
        while col_offset < total_columns:
            cell = ws.cell(row=row, column=start_col_index + col_offset)
            cell.value = 'X'
            cell.font = font_style
            cell.border = thin
            cell.alignment = center
            col_offset += 1

    if use_type1:
        print(f"      ✓ Type1 hesablamaları tamamlandı (E13:P24)")
    else:
        print(f"      ✓ Type2 hesablamaları tamamlandı (E13:AF32)")
    
    # ================== CƏM SƏTİRİNİ YAZMA ==================
    print(f"    Cəm sətiri yazılır (sətir {sum_row})...")
    
    for col_offset in range(total_columns):
        current_col = start_col_index + col_offset
        
        # Sütundakı bütün rəqəmləri topla (X-ləri skip et)
        column_sum = 0
        for row in range(start_row, end_row + 1):
            cell_value = ws.cell(row=row, column=current_col).value
            
            # Əgər rəqəmdirsə, topla
            if isinstance(cell_value, (int, float)):
                column_sum += cell_value
        
        # Cəm sətrinə yaz
        sum_cell = ws.cell(row=sum_row, column=current_col)
        sum_cell.value = column_sum
        sum_cell.font = bold_font  # Bold font
        sum_cell.border = thin
        sum_cell.alignment = center
        sum_cell.number_format = '#,##0.00'
    
    if use_type1:
        print(f"      ✓ Type1 cəm sətiri yazıldı (E25:P25)")
    else:
        print(f"      ✓ Type2 cəm sətiri yazıldı (E33:AF33)")
    
    # ================== DİAGONAL CƏMİ (B10:C11) ==================
    print(f"    Diagonal cəm hesablanır (B10:C11)...")
    
    diagonal_sum = 0
    for row_idx, col_idx in diagonal_coords:
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        
        if isinstance(cell_value, (int, float)):
            diagonal_sum += cell_value
    
    # B10:C11 artıq merge olunub, sadəcə B10-a yaz
    diagonal_cell = ws["B10"]
    diagonal_cell.value = diagonal_sum
    diagonal_cell.font = bold_font
    diagonal_cell.alignment = center
    diagonal_cell.number_format = '#,##0.00'

    # ================== DİAGONAL CƏMİ (B10:C11) ==================
    print(f"    Diagonal cəm hesablanır (B10:C11)...")
    
    diagonal_sum = 0
    for row_idx, col_idx in diagonal_coords:
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        
        if isinstance(cell_value, (int, float)):
            diagonal_sum += cell_value
    
    # B10:C11 artıq merge olunub, sadəcə B10-a yaz
    diagonal_cell = ws["B10"]
    diagonal_cell.value = diagonal_sum
    diagonal_cell.font = bold_font
    diagonal_cell.alignment = center
    diagonal_cell.number_format = '#,##0.00'
    
    if use_type1:
        print(f"      ✓ Type1 diagonal cəm: {diagonal_sum:.2f} (E24+F23+G22+H21)")
    else:
        print(f"      ✓ Type2 diagonal cəm: {diagonal_sum:.2f} (E32+F31+G30+H29)")
    
    # ================== REZERV ARTIMI SƏTİRİ ==================
    if use_type1:
        reserve_row = 26  # E26:O26
        reserve_columns = 11  # E-dən O-yə qədər (11 sütun)
        ratio_row = 27  # E27:O27
        print(f"    Rezerv artımı yazılır (E26:O26)...")
    else:  # use_type2
        reserve_row = 34  # E34:W34
        reserve_columns = 19  # E-dən W-yə qədər (19 sütun)
        ratio_row = 35  # E35:W35
        print(f"    Rezerv artımı yazılır (E34:W34)...")
    
    for col_offset in range(reserve_columns):
        current_col = start_col_index + col_offset
        
        # Cəm sətirindən dəyər
        sum_value = ws.cell(row=sum_row, column=current_col).value
        
        # Çıxılacaq sətir: end_row-dan başlayıb yuxarıya (col_offset qədər)
        subtract_row = end_row - col_offset
        subtract_value = ws.cell(row=subtract_row, column=current_col).value
        
        # Rezerv artımı hesabla
        if isinstance(sum_value, (int, float)) and isinstance(subtract_value, (int, float)):
            reserve_increase = sum_value - subtract_value
        elif isinstance(sum_value, (int, float)):
            # Əgər subtract_value X-dirsə və ya yoxdursa
            reserve_increase = sum_value
        else:
            reserve_increase = 0
        
        # Rezerv artımı sətrinə yaz
        reserve_cell = ws.cell(row=reserve_row, column=current_col)
        reserve_cell.value = reserve_increase
        reserve_cell.font = font_style
        reserve_cell.border = thin
        reserve_cell.alignment = center
        reserve_cell.number_format = '#,##0.00'
    
    if use_type1:
        print(f"      ✓ Type1 rezerv artımı yazıldı (E26:O26)")
    else:
        print(f"      ✓ Type2 rezerv artımı yazıldı (E34:W34)")
    
    # ================== NİSBƏT SƏTİRİ (RATIO) ==================
    if use_type1:
        print(f"    Nisbət yazılır (E27:O27)...")
    else:
        print(f"    Nisbət yazılır (E35:W35)...")

    # ================== 1. ƏVVƏLCƏ REZERV ARTIMINDA 0 VAR MI YOXLA ==================
    has_zero = False
    for col_offset in range(reserve_columns):
        current_col = start_col_index + col_offset
        reserve_value = ws.cell(row=reserve_row, column=current_col).value
        
        if isinstance(reserve_value, (int, float)) and reserve_value == 0:
            has_zero = True
            break

    # ================== 2. ƏGƏR 0 VARSA, FAİZ SHEET-DƏN BİRBAŞA GÖTÜR ==================
    if has_zero:
        print(f"      ⚠ Rezerv artımında 0 aşkarlandı, Faiz sheet-dən nisbətlər birbaşa kopyalanır...")
        
        try:
            # Faiz sheet-ni oxu
            faiz_df = pd.read_excel(ucot_file, sheet_name='Faiz')
            
            # Product-a görə filter
            faiz_filtered = faiz_df[faiz_df['Product'] == product]
            
            if not faiz_filtered.empty:
                # İlk sətri götür
                faiz_row = faiz_filtered.iloc[0]
                
                if use_type1:
                    # Type1: AA:AL sütunlarını E27:O27-yə kopyala
                    print(f"        → Type1: Faiz AA:AL → E27:O27")
                    
                    for idx in range(11):  # 11 sütun: AA, AB, AC, ..., AL
                        col_idx = 26 + idx  # AA=26, AB=27, ..., AL=36 (pandas 0-based)
                        target_col = start_col_index + idx
                        
                        if col_idx < len(faiz_row):
                            faiz_value = faiz_row.iloc[col_idx]
                            
                            if not pd.isna(faiz_value) and faiz_value != '':
                                try:
                                    numeric_value = float(faiz_value)
                                    
                                    ratio_cell = ws.cell(row=ratio_row, column=target_col)
                                    ratio_cell.value = numeric_value
                                    ratio_cell.font = font_style
                                    ratio_cell.border = thin
                                    ratio_cell.alignment = center
                                    ratio_cell.number_format = '0.0000'
                                    print(f"          ✓ Yazıldı: E27+{idx} = {numeric_value}")
                                except (ValueError, TypeError):
                                    ratio_cell = ws.cell(row=ratio_row, column=target_col)
                                    ratio_cell.value = 0
                                    ratio_cell.font = font_style
                                    ratio_cell.border = thin
                                    ratio_cell.alignment = center
                                    ratio_cell.number_format = '0.0000'
                                    print(f"          ⚠ Rəqəm deyil, 0 yazıldı: E27+{idx}")
                            else:
                                ratio_cell = ws.cell(row=ratio_row, column=target_col)
                                ratio_cell.value = 0
                                ratio_cell.font = font_style
                                ratio_cell.border = thin
                                ratio_cell.alignment = center
                                ratio_cell.number_format = '0.0000'
                                print(f"          ⚠ NaN/Boş, 0 yazıldı: E27+{idx}")
                        else:
                            ratio_cell = ws.cell(row=ratio_row, column=target_col)
                            ratio_cell.value = 0
                            ratio_cell.font = font_style
                            ratio_cell.border = thin
                            ratio_cell.alignment = center
                            ratio_cell.number_format = '0.0000'
                            print(f"          ⚠ Sütun yoxdur, 0 yazıldı: E27+{idx}")
                
                else:  # Type2
                    # Type2: AA:AS sütunlarını E35:W35-ə kopyala
                    print(f"        → Type2: Faiz AA:AS → E35:W35")
                    
                    for idx in range(19):  # 19 sütun: AA, AB, AC, ..., AS
                        col_idx = 26 + idx  # AA=26, AB=27, ..., AS=44 (pandas 0-based)
                        target_col = start_col_index + idx
                        
                        if col_idx < len(faiz_row):
                            faiz_value = faiz_row.iloc[col_idx]
                            
                            if not pd.isna(faiz_value) and faiz_value != '':
                                try:
                                    numeric_value = float(faiz_value)
                                    
                                    ratio_cell = ws.cell(row=ratio_row, column=target_col)
                                    ratio_cell.value = numeric_value
                                    ratio_cell.font = font_style
                                    ratio_cell.border = thin
                                    ratio_cell.alignment = center
                                    ratio_cell.number_format = '0.0000'
                                    print(f"          ✓ Yazıldı: E35+{idx} = {numeric_value}")
                                except (ValueError, TypeError):
                                    ratio_cell = ws.cell(row=ratio_row, column=target_col)
                                    ratio_cell.value = 0
                                    ratio_cell.font = font_style
                                    ratio_cell.border = thin
                                    ratio_cell.alignment = center
                                    ratio_cell.number_format = '0.0000'
                                    print(f"          ⚠ Rəqəm deyil, 0 yazıldı: E35+{idx}")
                            else:
                                ratio_cell = ws.cell(row=ratio_row, column=target_col)
                                ratio_cell.value = 0
                                ratio_cell.font = font_style
                                ratio_cell.border = thin
                                ratio_cell.alignment = center
                                ratio_cell.number_format = '0.0000'
                                print(f"          ⚠ NaN/Boş, 0 yazıldı: E35+{idx}")
                        else:
                            ratio_cell = ws.cell(row=ratio_row, column=target_col)
                            ratio_cell.value = 0
                            ratio_cell.font = font_style
                            ratio_cell.border = thin
                            ratio_cell.alignment = center
                            ratio_cell.number_format = '0.0000'
                            print(f"          ⚠ Sütun yoxdur, 0 yazıldı: E35+{idx}")
                
                print(f"        ✓ Faiz sheet-dən nisbətlər kopyalandı")
            
            else:
                print(f"        ⚠ Faiz sheet-də '{product}' üçün məlumat tapılmadı, normal hesablama edilir")
                
                for col_offset in range(reserve_columns):
                    current_col = start_col_index + col_offset
                    next_col = current_col + 1
                    
                    numerator = ws.cell(row=sum_row, column=next_col).value
                    denominator = ws.cell(row=reserve_row, column=current_col).value
                    
                    if isinstance(numerator, (int, float)) and isinstance(denominator, (int, float)) and denominator != 0:
                        ratio = round(numerator / denominator, 4)
                    else:
                        ratio = 0
                    
                    ratio_cell = ws.cell(row=ratio_row, column=current_col)
                    ratio_cell.value = ratio
                    ratio_cell.font = font_style
                    ratio_cell.border = thin
                    ratio_cell.alignment = center
                    ratio_cell.number_format = '0.0000'
        
        except Exception as e:
            print(f"        ✗ Faiz sheet oxunma xətası: {e}")
            
            for col_offset in range(reserve_columns):
                current_col = start_col_index + col_offset
                next_col = current_col + 1
                
                numerator = ws.cell(row=sum_row, column=next_col).value
                denominator = ws.cell(row=reserve_row, column=current_col).value
                
                if isinstance(numerator, (int, float)) and isinstance(denominator, (int, float)) and denominator != 0:
                    ratio = round(numerator / denominator, 4)
                else:
                    ratio = 0
                
                ratio_cell = ws.cell(row=ratio_row, column=current_col)
                ratio_cell.value = ratio
                ratio_cell.font = font_style
                ratio_cell.border = thin
                ratio_cell.alignment = center
                ratio_cell.number_format = '0.0000'

    else:
        # ================== 3. 0 YOXDURSA, NORMAL NİSBƏT HESABLA ==================
        print(f"      → 0 aşkarlanmadı, normal nisbət hesablanır")
        
        for col_offset in range(reserve_columns):
            current_col = start_col_index + col_offset
            next_col = current_col + 1
            
            numerator = ws.cell(row=sum_row, column=next_col).value
            denominator = ws.cell(row=reserve_row, column=current_col).value
            
            if isinstance(numerator, (int, float)) and isinstance(denominator, (int, float)) and denominator != 0:
                ratio = round(numerator / denominator, 4)
            else:
                ratio = 0
            
            ratio_cell = ws.cell(row=ratio_row, column=current_col)
            ratio_cell.value = ratio
            ratio_cell.font = font_style
            ratio_cell.border = thin
            ratio_cell.alignment = center
            ratio_cell.number_format = '0.0000'

    if use_type1:
        print(f"      ✓ Type1 nisbət yazıldı (E27:O27)")
    else:
        print(f"      ✓ Type2 nisbət yazıldı (E35:W35)")
    
    # ================== KUMULATİV NİSBƏT (CUMULATIVE RATIO) ==================
    if use_type1:
        cumulative_row = 28  # E28:O28
        print(f"    Kumulativ nisbət yazılır (E28:O28)...")
    else:
        cumulative_row = 36  # E36:W36
        print(f"    Kumulativ nisbət yazılır (E36:W36)...")
    
    for col_offset in range(reserve_columns):
        current_col = start_col_index + col_offset
        
        # Cari sütundan son sütuna qədər (O27 və ya W35) nisbətləri vur
        cumulative_product = 1.0
        
        for multiply_offset in range(col_offset, reserve_columns):
            multiply_col = start_col_index + multiply_offset
            ratio_value = ws.cell(row=ratio_row, column=multiply_col).value
            
            if isinstance(ratio_value, (int, float)) and ratio_value != 0:
                cumulative_product *= ratio_value
            else:
                # Əgər nisbət 0-dırsa və ya yoxdursa, hasili 0 et
                cumulative_product = 0
                break
        
        # Kumulativ nisbəti yaz (4 onluq)
        cumulative_product = round(cumulative_product, 4)
        
        cumulative_cell = ws.cell(row=cumulative_row, column=current_col)
        cumulative_cell.value = cumulative_product
        cumulative_cell.font = font_style
        cumulative_cell.border = thin
        cumulative_cell.alignment = center
        cumulative_cell.number_format = '0.0000'  # 4 onluq format
    
    if use_type1:
        print(f"      ✓ Type1 kumulativ nisbət yazıldı (E28:O28)")
    else:
        print(f"      ✓ Type2 kumulativ nisbət yazıldı (E36:W36)")
    
    # ================== TƏRS NİSBƏT (INVERSE RATIO) ==================
    if use_type1:
        inverse_row = 29  # E29:O29
        print(f"    Tərs nisbət yazılır (E29:O29)...")
    else:
        inverse_row = 37  # E37:W37
        print(f"    Tərs nisbət yazılır (E37:W37)...")
    
    for col_offset in range(reserve_columns):
        current_col = start_col_index + col_offset
        
        # Kumulativ nisbətdən dəyər al
        cumulative_value = ws.cell(row=cumulative_row, column=current_col).value
        
        # Tərs nisbət hesabla: 1 / kumulativ_nisbət (4 onluq)
        if isinstance(cumulative_value, (int, float)) and cumulative_value != 0:
            inverse_ratio = round(1.0 / cumulative_value, 4)
        else:
            inverse_ratio = 0  # Sıfıra bölmə xətası varsa 0
        
        # Tərs nisbət sətrinə yaz
        inverse_cell = ws.cell(row=inverse_row, column=current_col)
        inverse_cell.value = inverse_ratio
        inverse_cell.font = font_style
        inverse_cell.border = thin
        inverse_cell.alignment = center
        inverse_cell.number_format = '0.0000'  # 4 onluq format
    
    if use_type1:
        print(f"      ✓ Type1 tərs nisbət yazıldı (E29:O29)")
    else:
        print(f"      ✓ Type2 tərs nisbət yazıldı (E37:W37)")
    
    # ================== FORMA8_10-DƏN G SÜTUNUNU KOPYALAMA ==================
    print(f"    Forma8_10-dən G sütunu kopyalanır...")

    # Forma8_10 sheet-ni yoxla
    if "Forma8_10" not in wb.sheetnames:
        print(f"      ⚠ Forma8_10 sheet-i tapılmadı, skip edilir")
    else:
        ws_8_10 = wb["Forma8_10"]
        
        if use_type1:
            # Type1: G13:G24 → Q13:Q24
            source_start = 13
            source_end = 24
            target_column = 17  # Q sütunu (column=17)
            print(f"      → Type1: G13:G24 → Q13:Q24")
        else:
            # Type2: G13:G32 → Y13:Y32
            source_start = 13
            source_end = 32
            target_column = 25  # Y sütunu (column=25)
            print(f"      → Type2: G13:G32 → Y13:Y32")
        
        rows_copied = 0
        for row in range(source_start, source_end + 1):
            # Forma8_10-dən G sütununu oxu
            source_cell = ws_8_10.cell(row=row, column=7)  # G sütunu (column=7)
            
            # Forma8_12-də target sütununa yaz
            target_cell = ws.cell(row=row, column=target_column)
            
            # Dəyəri kopyala
            target_cell.value = source_cell.value
            
            # Formatı kopyala
            if source_cell.font:
                target_cell.font = source_cell.font.copy()
            else:
                target_cell.font = font_style
            
            if source_cell.border:
                target_cell.border = source_cell.border.copy()
            else:
                target_cell.border = thin
            
            if source_cell.fill:
                target_cell.fill = source_cell.fill.copy()
            
            if source_cell.alignment:
                target_cell.alignment = source_cell.alignment.copy()
            else:
                target_cell.alignment = center
            
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            else:
                target_cell.number_format = '#,##0.00'
            
            rows_copied += 1
        
        if use_type1:
            print(f"      ✓ Type1: {rows_copied} sətir kopyalandı (G13:G24 → Q13:Q24)")
        else:
            print(f"      ✓ Type2: {rows_copied} sətir kopyalandı (G13:G32 → Y13:Y32)")
    
    # ================== R VƏ Z SÜTUNLARI HESABLAMA ==================
    if use_type1:
        print(f"    R13:R24 hesablanır...")
        
        calculation_start = 13
        calculation_end = 24
        result_column = 18  # R sütunu
        fixed_column = 17   # Q sütunu (sabit)
        multiplier_row = 28 # P28, O28, N28, ... sətiri  ⬅️ DƏYİŞDİ
        
        # Başlanğıc sütun: P (16)
        start_col = 16
        
        for idx, row in enumerate(range(calculation_start, calculation_end + 1)):
            # Hər sətirdə sütun 1 azalır: P, O, N, M, ...
            current_col = start_col - idx
            
            # Pay: P13, O14, N15, ...
            numerator_cell = ws.cell(row=row, column=current_col)
            numerator = numerator_cell.value or 0
            
            # Məxrəc: Q13, Q14, Q15, ... (sabit sütun)
            denominator_cell = ws.cell(row=row, column=fixed_column)
            denominator = denominator_cell.value or 0
            
            # Vurulacaq: P28, O28, N28, ... (eyni sütun, 28-ci sətir)  ⬅️
            multiplier_cell = ws.cell(row=multiplier_row, column=current_col)
            multiplier = multiplier_cell.value or 0
            
            # Hesablama: (numerator / denominator) × multiplier
            try:
                numerator = float(numerator)
                denominator = float(denominator)
                multiplier = float(multiplier)
                
                if denominator != 0:
                    result = (numerator / denominator) * multiplier
                else:
                    result = 0
            except (ValueError, TypeError):
                result = 0
            
            # Nəticəni yaz (4 onluq)
            result_cell = ws.cell(row=row, column=result_column)
            result_cell.value = round(result, 4)
            result_cell.font = font_style
            result_cell.border = thin
            result_cell.alignment = center
            result_cell.number_format = '#,##0.0000'
        
        print(f"      ✓ Type1: R13:R24 hesablandı")
    
    else:  # Type2
        print(f"    Z13:Z32 hesablanır...")
        
        calculation_start = 13
        calculation_end = 32
        result_column = 26  # Z sütunu
        fixed_column = 25   # Y sütunu (sabit)
        multiplier_row = 36 # X36, W36, V36, ... sətiri
        
        # Başlanğıc sütun: X (24)
        start_col = 24
        
        for idx, row in enumerate(range(calculation_start, calculation_end + 1)):
            # Hər sətirdə sütun 1 azalır: X, W, V, U, ...
            current_col = start_col - idx
            
            # Pay: X13, W14, V15, ...
            numerator_cell = ws.cell(row=row, column=current_col)
            numerator = numerator_cell.value or 0
            
            # Məxrəc: Y13, Y14, Y15, ... (sabit sütun)
            denominator_cell = ws.cell(row=row, column=fixed_column)
            denominator = denominator_cell.value or 0
            
            # Vurulacaq: X36, W36, V36, ... (eyni sütun, 36-cı sətir)
            multiplier_cell = ws.cell(row=multiplier_row, column=current_col)
            multiplier = multiplier_cell.value or 0
            
            # Hesablama: (numerator / denominator) × multiplier
            try:
                numerator = float(numerator)
                denominator = float(denominator)
                multiplier = float(multiplier)
                
                if denominator != 0:
                    result = (numerator / denominator) * multiplier
                else:
                    result = 0
            except (ValueError, TypeError):
                result = 0
            
            # Nəticəni yaz (4 onluq)
            result_cell = ws.cell(row=row, column=result_column)
            result_cell.value = round(result, 4)
            result_cell.font = font_style
            result_cell.border = thin
            result_cell.alignment = center
            result_cell.number_format = '#,##0.0000'
        
        print(f"      ✓ Type2: Z13:Z32 hesablandı")
    
    # ================== S VƏ AA SÜTUNLARI (AVERAGE VƏ YA FAİZ12) ==================
    # Əgər əvvəldə Faiz sheet istifadə olubsa, yenidən istifadə et
    if has_zero:  # Nisbət sətirində 0 vardısa, Faiz sheet istifadə olunub
        print(f"    Faiz sheet-dən Faiz12 sütunu oxunur...")
        
        try:
            # Faiz sheet-ni oxu (əgər əvvəl oxunubsa, yenidən oxu)
            faiz_df = pd.read_excel(ucot_file, sheet_name='Faiz')
            
            # Product-a görə filter
            faiz_filtered = faiz_df[faiz_df['Product'] == product]
            
            if not faiz_filtered.empty:
                faiz_row = faiz_filtered.iloc[0]
                
                # Faiz12 sütununu tap
                if 'Faiz12' in faiz_row.index:
                    faiz12_value = faiz_row['Faiz12']
                    
                    # Rəqəmə çevir
                    if not pd.isna(faiz12_value) and faiz12_value != '':
                        try:
                            faiz12_numeric = float(faiz12_value)
                            faiz12_numeric = round(faiz12_numeric, 4)
                            
                            if use_type1:
                                print(f"      → Type1: Faiz12 = {faiz12_numeric:.4f} → S13:S24")
                                
                                # S13:S24-ə Faiz12 dəyərini yaz
                                for row in range(13, 25):
                                    s_cell = ws.cell(row=row, column=19)  # S sütunu
                                    s_cell.value = faiz12_numeric
                                    s_cell.font = font_style
                                    s_cell.border = thin
                                    s_cell.alignment = center
                                    s_cell.number_format = '#,##0.0000'
                                
                                print(f"      ✓ Type1: S13:S24 = {faiz12_numeric:.4f} (Faiz12-dən)")
                            
                            else:  # Type2
                                print(f"      → Type2: Faiz12 = {faiz12_numeric:.4f} → AA13:AA32")
                                
                                # AA13:AA32-yə Faiz12 dəyərini yaz
                                for row in range(13, 33):
                                    aa_cell = ws.cell(row=row, column=27)  # AA sütunu
                                    aa_cell.value = faiz12_numeric
                                    aa_cell.font = font_style
                                    aa_cell.border = thin
                                    aa_cell.alignment = center
                                    aa_cell.number_format = '#,##0.0000'
                                
                                print(f"      ✓ Type2: AA13:AA32 = {faiz12_numeric:.4f} (Faiz12-dən)")
                        
                        except (ValueError, TypeError):
                            print(f"      ⚠ Faiz12 rəqəmə çevrilə bilmədi, ortalama hesablanacaq")
                            # Faiz12 yoxdursa və ya səhvdirsə, ortalama hesabla
                            has_zero = False  # Ortalamanı trigger et
                    else:
                        print(f"      ⚠ Faiz12 boşdur, ortalama hesablanacaq")
                        has_zero = False
                else:
                    print(f"      ⚠ Faiz12 sütunu tapılmadı, ortalama hesablanacaq")
                    has_zero = False
            else:
                print(f"      ⚠ Faiz sheet-də '{product}' tapılmadı, ortalama hesablanacaq")
                has_zero = False
        
        except Exception as e:
            print(f"      ✗ Faiz sheet oxunma xətası: {e}, ortalama hesablanacaq")
            has_zero = False

    # Əgər Faiz sheet istifadə olunmayıbsa və ya Faiz12 tapılmayıbsa, ortalama hesabla
    if not has_zero:
        if use_type1:
            print(f"    S13:S24 hesablanır (R sütununun ortalaması)...")
            
            # R13:R24 dəyərlərini topla
            r_values = []
            for row in range(13, 25):
                r_value = ws.cell(row=row, column=18).value  # R sütunu
                if isinstance(r_value, (int, float)):
                    r_values.append(r_value)
            
            # Ortalama hesabla
            if r_values:
                average = sum(r_values) / len(r_values)
                average = round(average, 4)
            else:
                average = 0
            
            print(f"      R13:R24 ortalaması: {average:.4f}")
            
            # S13:S24-ə yaz
            for row in range(13, 25):
                s_cell = ws.cell(row=row, column=19)
                s_cell.value = average
                s_cell.font = font_style
                s_cell.border = thin
                s_cell.alignment = center
                s_cell.number_format = '#,##0.0000'
            
            print(f"      ✓ Type1: S13:S24 = {average:.4f} (ortalama)")
        
        else:  # Type2
            print(f"    AA13:AA32 hesablanır (Z sütununun ortalaması)...")
            
            # Z13:Z32 dəyərlərini topla
            z_values = []
            for row in range(13, 33):
                z_value = ws.cell(row=row, column=26).value  # Z sütunu
                if isinstance(z_value, (int, float)):
                    z_values.append(z_value)
            
            # Ortalama hesabla
            if z_values:
                average = sum(z_values) / len(z_values)
                average = round(average, 4)
            else:
                average = 0
            
            print(f"      Z13:Z32 ortalaması: {average:.4f}")
            
            # AA13:AA32-yə yaz
            for row in range(13, 33):
                aa_cell = ws.cell(row=row, column=27)
                aa_cell.value = average
                aa_cell.font = font_style
                aa_cell.border = thin
                aa_cell.alignment = center
                aa_cell.number_format = '#,##0.0000'
            
            print(f"      ✓ Type2: AA13:AA32 = {average:.4f} (ortalama)")
    
    # ================== T VƏ AB SÜTUNLARI HESABLAMA ==================
    if use_type1:
        print(f"    T13:T24 hesablanır (Q × S)...")
        
        for row in range(13, 25):  # 13-dən 24-ə qədər
            # Q sütununu oxu
            q_value = ws.cell(row=row, column=17).value or 0  # Q sütunu (column=17)
            
            # S sütununu oxu
            s_value = ws.cell(row=row, column=19).value or 0  # S sütunu (column=19)
            
            # T hesabla: Q × S
            try:
                q_value = float(q_value)
                s_value = float(s_value)
                t_value = q_value * s_value
                t_value = round(t_value, 2)  # 2 onluq
            except (ValueError, TypeError):
                t_value = 0
            
            # T sütununa yaz
            t_cell = ws.cell(row=row, column=20)  # T sütunu (column=20)
            t_cell.value = t_value
            t_cell.font = font_style
            t_cell.border = thin
            t_cell.alignment = center
            t_cell.number_format = '#,##0.00'  # 2 onluq format
        
        print(f"      ✓ Type1: T13:T24 hesablandı (Q × S)")
    
    else:  # Type2
        print(f"    AB13:AB32 hesablanır (Y × AA)...")
        
        for row in range(13, 33):  # 13-dən 32-yə qədər
            # Y sütununu oxu
            y_value = ws.cell(row=row, column=25).value or 0  # Y sütunu (column=25)
            
            # AA sütununu oxu
            aa_value = ws.cell(row=row, column=27).value or 0  # AA sütunu (column=27)
            
            # AB hesabla: Y × AA
            try:
                y_value = float(y_value)
                aa_value = float(aa_value)
                ab_value = y_value * aa_value
                ab_value = round(ab_value, 2)  # 2 onluq
            except (ValueError, TypeError):
                ab_value = 0
            
            # AB sütununa yaz
            ab_cell = ws.cell(row=row, column=28)  # AB sütunu (column=28)
            ab_cell.value = ab_value
            ab_cell.font = font_style
            ab_cell.border = thin
            ab_cell.alignment = center
            ab_cell.number_format = '#,##0.00'  # 2 onluq format
        
        print(f"      ✓ Type2: AB13:AB32 hesablandı (Y × AA)")
    
    # ================== U VƏ AC SÜTUNLARI HESABLAMA ==================
    if use_type1:
        print(f"    U13:U24 hesablanır ((1-inverse_ratio) × T)...")
        
        calculation_start = 13
        calculation_end = 24
        result_column = 21  # U sütunu
        t_column = 20       # T sütunu
        inverse_row = 29    # P29, O29, N29, ... sətiri
        
        # Başlanğıc sütun: P (16)
        start_col = 16
        
        for idx, row in enumerate(range(calculation_start, calculation_end + 1)):
            # Hər sətirdə sütun 1 azalır: P, O, N, M, ...
            current_col = start_col - idx
            
            # Inverse ratio: P29, O29, N29, ...
            inverse_cell = ws.cell(row=inverse_row, column=current_col)
            inverse_value = inverse_cell.value or 0
            
            # T sütunu: T13, T14, T15, ...
            t_cell = ws.cell(row=row, column=t_column)
            t_value = t_cell.value or 0
            
            # Hesablama: (1 - inverse_value) × t_value
            try:
                inverse_value = float(inverse_value)
                t_value = float(t_value)
                result = (1 - inverse_value) * t_value
                result = round(result, 2)  # 2 onluq
            except (ValueError, TypeError):
                result = 0
            
            # Nəticəni yaz
            result_cell = ws.cell(row=row, column=result_column)
            result_cell.value = result
            result_cell.font = font_style
            result_cell.border = thin
            result_cell.alignment = center
            result_cell.number_format = '#,##0.00'  # 2 onluq format
        
        print(f"      ✓ Type1: U13:U24 hesablandı")

    else:  # Type2
        print(f"    AC13:AC32 hesablanır ((1-inverse_ratio) × AB)...")
        
        calculation_start = 13
        calculation_end = 32
        result_column = 29  # AC sütunu
        ab_column = 28      # AB sütunu
        inverse_row = 37    # X37, W37, V37, ... sətiri
        
        # Başlanğıc sütun: X (24)
        start_col = 24
        
        for idx, row in enumerate(range(calculation_start, calculation_end + 1)):
            # Hər sətirdə sütun 1 azalır: X, W, V, U, ...
            current_col = start_col - idx
            
            # Inverse ratio: X37, W37, V37, ...
            inverse_cell = ws.cell(row=inverse_row, column=current_col)
            inverse_value = inverse_cell.value or 0
            
            # AB sütunu: AB13, AB14, AB15, ...
            ab_cell = ws.cell(row=row, column=ab_column)
            ab_value = ab_cell.value or 0
            
            # Hesablama: (1 - inverse_value) × ab_value
            try:
                inverse_value = float(inverse_value)
                ab_value = float(ab_value)
                result = (1 - inverse_value) * ab_value
                result = round(result, 2)  # 2 onluq
            except (ValueError, TypeError):
                result = 0
            
            # Nəticəni yaz
            result_cell = ws.cell(row=row, column=result_column)
            result_cell.value = result
            result_cell.font = font_style
            result_cell.border = thin
            result_cell.alignment = center
            result_cell.number_format = '#,##0.00'  # 2 onluq format
        
        print(f"      ✓ Type2: AC13:AC32 hesablandı")

    # ================== FORMA8_3-DƏN D SÜTUNUNU KOPYALAMA ==================
    print(f"    Forma8_3-dən D sütunu kopyalanır...")

    # Forma8_3 sheet-ni yoxla
    if "Forma8_3" not in wb.sheetnames:
        print(f"      ⚠ Forma8_3 sheet-i tapılmadı, skip edilir")
    else:
        ws_8_3 = wb["Forma8_3"]
        
        if use_type1:
            # Type1: D13:D24 → V13:V24
            source_start = 13
            source_end = 24
            target_column = 22  # V sütunu (column=22)
            print(f"      → Type1: D13:D24 → V13:V24")
        else:
            # Type2: D13:D32 → AD13:AD32
            source_start = 13
            source_end = 32
            target_column = 30  # AD sütunu (column=30)
            print(f"      → Type2: D13:D32 → AD13:AD32")
        
        rows_copied = 0
        for row in range(source_start, source_end + 1):
            # Forma8_3-dən D sütununu oxu
            source_cell = ws_8_3.cell(row=row, column=4)  # D sütunu (column=4)
            
            # Forma8_12-də target sütununa yaz
            target_cell = ws.cell(row=row, column=target_column)
            
            # Dəyəri kopyala
            target_cell.value = source_cell.value
            
            # Formatı kopyala
            if source_cell.font:
                target_cell.font = source_cell.font.copy()
            else:
                target_cell.font = font_style
            
            if source_cell.border:
                target_cell.border = source_cell.border.copy()
            else:
                target_cell.border = thin
            
            if source_cell.fill:
                target_cell.fill = source_cell.fill.copy()
            
            if source_cell.alignment:
                target_cell.alignment = source_cell.alignment.copy()
            else:
                target_cell.alignment = center
            
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            else:
                target_cell.number_format = '#,##0.00'
            
            rows_copied += 1
        
        if use_type1:
            print(f"      ✓ Type1: {rows_copied} sətir kopyalandı (D13:D24 → V13:V24)")
        else:
            print(f"      ✓ Type2: {rows_copied} sətir kopyalandı (D13:D32 → AD13:AD32)")

    # ================== W VƏ AE SÜTUNLARI HESABLAMA ==================
    if use_type1:
        print(f"    W13:W24 hesablanır (U - V, neqativ varsa 0)...")
        
        calculation_start = 13
        calculation_end = 24
        result_column = 23  # W sütunu
        u_column = 21       # U sütunu
        v_column = 22       # V sütunu
        
        for row in range(calculation_start, calculation_end + 1):
            # U sütununu oxu
            u_value = ws.cell(row=row, column=u_column).value or 0
            
            # V sütununu oxu
            v_value = ws.cell(row=row, column=v_column).value or 0
            
            # W hesabla: U - V
            try:
                u_value = float(u_value)
                v_value = float(v_value)
                result = u_value - v_value
                
                # Əgər neqativdirsə 0 yaz
                if result < 0:
                    result = 0.00
                else:
                    result = round(result, 2)  # 2 onluq
            except (ValueError, TypeError):
                result = 0.00
            
            # Nəticəni yaz
            result_cell = ws.cell(row=row, column=result_column)
            result_cell.value = result
            result_cell.font = font_style
            result_cell.border = thin
            result_cell.alignment = center
            result_cell.number_format = '#,##0.00'  # 2 onluq format
        
        print(f"      ✓ Type1: W13:W24 hesablandı")

    else:  # Type2
        print(f"    AE13:AE32 hesablanır (AC - AD, neqativ varsa 0)...")
        
        calculation_start = 13
        calculation_end = 32
        result_column = 31  # AE sütunu
        ac_column = 29      # AC sütunu
        ad_column = 30      # AD sütunu
        
        for row in range(calculation_start, calculation_end + 1):
            # AC sütununu oxu
            ac_value = ws.cell(row=row, column=ac_column).value or 0
            
            # AD sütununu oxu
            ad_value = ws.cell(row=row, column=ad_column).value or 0
            
            # AE hesabla: AC - AD
            try:
                ac_value = float(ac_value)
                ad_value = float(ad_value)
                result = ac_value - ad_value
                
                # Əgər neqativdirsə 0 yaz
                if result < 0:
                    result = 0.00
                else:
                    result = round(result, 2)  # 2 onluq
            except (ValueError, TypeError):
                result = 0.00
            
            # Nəticəni yaz
            result_cell = ws.cell(row=row, column=result_column)
            result_cell.value = result
            result_cell.font = font_style
            result_cell.border = thin
            result_cell.alignment = center
            result_cell.number_format = '#,##0.00'  # 2 onluq format
        
        print(f"      ✓ Type2: AE13:AE32 hesablandı")

    # ================== W VƏ AE SÜTUNLARI TOPLAMI ==================
    if use_type1:
        print(f"    W30 hesablanır (W13:W24 toplamı)...")
        
        sum_row = 30
        w_column = 23  # W sütunu
        
        # W13:W24 toplamını hesabla
        total = 0
        for row in range(13, 25):  # 13-dən 24-ə qədər
            w_value = ws.cell(row=row, column=w_column).value or 0
            try:
                total += float(w_value)
            except (ValueError, TypeError):
                pass
        
        total = round(total, 2)
        
        # W30-a yaz
        sum_cell = ws.cell(row=sum_row, column=w_column)
        sum_cell.value = total
        sum_cell.font = bold_font
        sum_cell.border = thin
        sum_cell.alignment = center
        sum_cell.number_format = '#,##0.00'
        
        print(f"      ✓ Type1: W30 = {total:.2f}")

    else:  # Type2
        print(f"    AE38 hesablanır (AE13:AE32 toplamı)...")
        
        sum_row = 38
        ae_column = 31  # AE sütunu
        
        # AE13:AE32 toplamını hesabla
        total = 0
        for row in range(13, 33):  # 13-dən 32-yə qədər
            ae_value = ws.cell(row=row, column=ae_column).value or 0
            try:
                total += float(ae_value)
            except (ValueError, TypeError):
                pass
        
        total = round(total, 2)
        
        # AE38-ə yaz
        sum_cell = ws.cell(row=sum_row, column=ae_column)
        sum_cell.value = total
        sum_cell.font = bold_font
        sum_cell.border = thin
        sum_cell.alignment = center
        sum_cell.number_format = '#,##0.00'
        
        print(f"      ✓ Type2: AE38 = {total:.2f}")

    # ================== W31 VƏ AE39 HESABLAMA (×1.03) ==================
    if use_type1:
        print(f"    W31 hesablanır (W30 × 1.03)...")
        
        source_row = 30
        target_row = 31
        w_column = 23  # W sütunu
        
        # W30 dəyərini oxu
        w30_value = ws.cell(row=source_row, column=w_column).value or 0
        
        # W31 hesabla: W30 × 1.03
        try:
            w30_value = float(w30_value)
            result = w30_value * 1.03
            result = round(result, 2)
        except (ValueError, TypeError):
            result = 0.00
        
        # W31-ə yaz
        result_cell = ws.cell(row=target_row, column=w_column)
        result_cell.value = result
        result_cell.font = bold_font
        result_cell.border = thin
        result_cell.alignment = center
        result_cell.number_format = '#,##0.00'
        
        print(f"      ✓ Type1: W31 = {result:.2f} (W30 × 1.03)")

    else:  # Type2
        print(f"    AE39 hesablanır (AE38 × 1.03)...")
        
        source_row = 38
        target_row = 39
        ae_column = 31  # AE sütunu
        
        # AE38 dəyərini oxu
        ae38_value = ws.cell(row=source_row, column=ae_column).value or 0
        
        # AE39 hesabla: AE38 × 1.03
        try:
            ae38_value = float(ae38_value)
            result = ae38_value * 1.03
            result = round(result, 2)
        except (ValueError, TypeError):
            result = 0.00
        
        # AE39-a yaz
        result_cell = ws.cell(row=target_row, column=ae_column)
        result_cell.value = result
        result_cell.font = bold_font
        result_cell.border = thin
        result_cell.alignment = center
        result_cell.number_format = '#,##0.00'
        
        print(f"      ✓ Type2: AE39 = {result:.2f} (AE38 × 1.03)")

    wb.save(excel_file)
    print(f"  ✅ {product}: Forma8_12 tamamlandı")