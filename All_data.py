import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment

# Fayl yolları
ucot_file = r"C:\Users\samir.alizade\Desktop\Automantion\Risk\UcotA.xlsx"
target_file = r"C:\Users\samir.alizade\Desktop\Automantion\Risk\ALL.xlsx"

# 1. Ucot -> Simple sheet oxu
df = pd.read_excel(ucot_file, sheet_name="Simple")

# 2. I sütununa görə filter
filtered_df = df[df["I"] == "(04)AvtoKasko"]

# 3. Lazım olan sütunları seç
result_df = filtered_df[["II", "III", "VII", "XI"]].copy()

# 🔹 NaN dəyərləri 0 ilə əvəz et
result_df["VII"] = result_df["VII"].fillna(0)
result_df["XI"] = result_df["XI"].fillna(0)

print(f"📊 Total filtered rows: {len(result_df)}")

# 4. Target excel aç
book = load_workbook(target_file)
sheet = book["Forma8_1"]

# 🔹 Merge-ləri aç
for merged_range in list(sheet.merged_cells.ranges):
    sheet.unmerge_cells(str(merged_range))

# 🔹 Border və Font tərifi
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
cell_font = Font(name='A3 Times AZ Lat', size=14)
bold_font = Font(name='A3 Times AZ Lat', size=14, bold=True)
center_align = Alignment(horizontal='center', vertical='center')

# A11 başlığı
sheet["A11"] = "A"
sheet["A11"].border = thin_border
sheet["A11"].font = cell_font
sheet["A11"].alignment = center_align

start_row = 12
data_rows = len(result_df)

# 🔹 B12-dən əvvəl boş sətirlər əlavə et
sheet.insert_rows(start_row, amount=data_rows)

# 🔹 Toplamlar
total_D = total_E = total_F = total_G = total_H = total_I = 0
row_count = 0

# 🔹 Data sətirləri
for idx, row in result_df.iterrows():
    try:
        vii = float(row["VII"])
        xi = float(row["XI"])

        limit_15 = round(vii * 0.15, 2)
        final_xi = round(limit_15 if xi > limit_15 else xi, 2)

        # A sütunu
        sheet[f"A{start_row}"] = f"A{start_row-11}"
        sheet[f"A{start_row}"].border = thin_border
        sheet[f"A{start_row}"].font = cell_font
        sheet[f"A{start_row}"].alignment = center_align

        # B sütunu
        sheet[f"B{start_row}"] = row["II"]
        sheet[f"B{start_row}"].border = thin_border
        sheet[f"B{start_row}"].font = cell_font
        sheet[f"B{start_row}"].alignment = center_align

        # C sütunu (short date)
        cell = sheet[f"C{start_row}"]
        if pd.notna(row["III"]):
            date_value = pd.to_datetime(row["III"]).date()
            cell.value = date_value
            cell.number_format = "DD.MM.YYYY"
        else:
            cell.value = ""
        cell.border = thin_border
        cell.font = cell_font
        cell.alignment = center_align

        # D–I sütunları
        val_D = round(vii, 2)
        val_E = round(final_xi, 2)
        val_F = round(vii - final_xi, 2)
        val_G = round(vii * 0.02, 2)
        val_H = round(final_xi * 0.02, 2)
        val_I = round(val_G - val_H, 2)

        for col, val in zip(['D','E','F','G','H','I'], [val_D, val_E, val_F, val_G, val_H, val_I]):
            sheet[f"{col}{start_row}"] = val
            sheet[f"{col}{start_row}"].border = thin_border
            sheet[f"{col}{start_row}"].font = cell_font
            sheet[f"{col}{start_row}"].alignment = center_align

        # Toplamlar
        total_D += val_D
        total_E += val_E
        total_F += val_F
        total_G += val_G
        total_H += val_H
        total_I += val_I

        row_count += 1
        start_row += 1

    except Exception as e:
        print(f"❌ Error at row {idx}: {e}")
        continue

print(f"✅ {row_count} sətir yazıldı (expected: {len(result_df)})")

# 🔹 Toplamları yuvarlaqlaşdır
total_D = round(total_D, 2)
total_E = round(total_E, 2)
total_F = round(total_F, 2)
total_G = round(total_G, 2)
total_H = round(total_H, 2)
total_I = round(total_I, 2)

# 🔹 Yekun sətir
sheet.cell(row=start_row, column=1).value = "AA1"
sheet.cell(row=start_row, column=1).border = thin_border
sheet.cell(row=start_row, column=1).font = cell_font
sheet.cell(row=start_row, column=1).alignment = center_align

sheet.cell(row=start_row, column=2).value = "Yekun BSH"
sheet.cell(row=start_row, column=2).border = thin_border
sheet.cell(row=start_row, column=2).font = cell_font
sheet.cell(row=start_row, column=2).alignment = center_align

for col, val in zip(range(4,10), [total_D, total_E, total_F, total_G, total_H, total_I]):
    sheet.cell(row=start_row, column=col).value = val
    sheet.cell(row=start_row, column=col).border = thin_border
    sheet.cell(row=start_row, column=col).font = bold_font  # Bold yekun
    sheet.cell(row=start_row, column=col).alignment = center_align

# 4 sətir aşağı totals
copy_row = start_row + 4

for col, val in zip(range(4,10), [total_D, total_E, total_F, total_G, total_H, total_I]):
    sheet.cell(row=copy_row, column=col).value = val
    sheet.cell(row=copy_row, column=col).border = thin_border
    sheet.cell(row=copy_row, column=col).font = bold_font
    sheet.cell(row=copy_row, column=col).alignment = center_align

# 🔹 Bütün əlavə etdiyimiz sətirlərin height-ni eyni et
ROW_HEIGHT = 28  # istədiyin ölçü (Excel default ~15-dir)

first_data_row = 12
last_data_row = copy_row  # 4 sətir aşağı totals daxil

for r in range(first_data_row, last_data_row + 1):
    sheet.row_dimensions[r].height = ROW_HEIGHT

# 6. Yadda saxla
book.save(target_file)
print("\n✅ Fayl saxlanıldı!")
